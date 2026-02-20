#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FTTH Automation for EMR Project
Merupakan sebuah tool automation yang digunakan untuk membantu pekerjaan
para engineer FTTH di EMR Project. Terkhusus untuk drafter FTTH di PT Fiberhome Technologies Indonesia.

Tools ini dibuat oleh Galih Prima Aditya Firdaus
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import math
import os
import xml.etree.ElementTree as ET
from datetime import datetime
import csv
import re

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill
    import base64
    from io import BytesIO
    _PANDAS_OK = True
except ImportError:
    _PANDAS_OK = False

import subprocess
import sys

# Windows taskbar icon support
try:
    from ctypes import windll
    myappid = 'ftth.automationapp.1.0'
    windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except Exception:
    pass


class ModernButton(tk.Canvas):
    """Custom button widget with modern styling, supports state='disabled'/'normal'"""
    _DISABLED_FG  = '#6b7280'
    _DISABLED_BG  = '#3d4f5e'

    def __init__(self, parent, text, command, bg_color, hover_color, **kwargs):
        super().__init__(parent, highlightthickness=0, cursor='hand2', **kwargs)
        self.bg_color    = bg_color
        self.hover_color = hover_color
        self.command     = command
        self.text        = text
        self._state      = 'normal'

        super().configure(bg=bg_color, height=35)
        self.text_id = self.create_text(
            0, 0, text=text, fill='white',
            font=('Segoe UI', 9, 'bold'), anchor='center'
        )

        self.bind('<Button-1>', self._on_click)
        self.bind('<Enter>',    self._on_enter)
        self.bind('<Leave>',    self._on_leave)
        self.bind('<Configure>', self._on_configure)

    def _on_click(self, e):
        if self._state == 'normal':
            self.command()

    def _on_configure(self, event):
        self.coords(self.text_id, event.width // 2, event.height // 2)

    def _on_enter(self, event):
        if self._state == 'normal':
            super().configure(bg=self.hover_color)

    def _on_leave(self, event):
        if self._state == 'normal':
            super().configure(bg=self.bg_color)
        # when disabled: bg already set by configure, don't reset

    def configure(self, **kw):
        state = kw.pop('state', None)
        if kw:
            super().configure(**kw)
        if state is not None:
            self._state = state
            if hasattr(self, 'text_id'):          # guard: text_id may not exist yet
                if state == 'disabled':
                    super().configure(bg=self._DISABLED_BG, cursor='arrow')
                    self.itemconfigure(self.text_id, fill=self._DISABLED_FG)
                else:
                    super().configure(bg=self.bg_color, cursor='hand2')
                    self.itemconfigure(self.text_id, fill='white')

    config = configure


class FTTHAutomationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FTTH Automation for EMR Project")
        self.root.geometry("1000x750")
        self.root.resizable(False, False)
        
        # Set window icon
        try:
            self.root.iconbitmap('app.ico')
        except Exception:
            pass
        
        # Center window on screen
        self.root.update_idletasks()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = 1000
        window_height = 750
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Figma Color Palette
        self.colors = {
            'bg_main': '#2d3e50',
            'bg_section': '#3a4f63',
            'bg_input': '#4a5f73',
            'border': '#5a6f83',
            'title_bar': '#0066cc',
            'title_bar_hover': '#0052a3',
            'output_bg': '#1a2f1a',
            'output_border': '#2d4a2d',
            'output_text': '#4ade80',
            'text_primary': '#ffffff',
            'text_secondary': '#d1d5db',
            'text_label': '#9ca3af',
            'button_blue': '#0066cc',
            'button_blue_hover': '#0052a3',
            'button_gray': '#4a5f73',
            'button_gray_hover': '#5a6f83',
            'button_green': '#28a745',
            'button_green_hover': '#218838',
            'gradient_start': '#3a4f63',
            'gradient_end': '#2d3e50',
        }
        
        self.root.configure(bg=self.colors['bg_main'])
        
        # Main container
        main_container = tk.Frame(root, bg=self.colors['bg_main'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # Header (no title bar)
        header = tk.Frame(main_container, bg=self.colors['gradient_start'], height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        title_label = tk.Label(header, text="FTTH Automation for EMR Project",
                      bg=self.colors['gradient_start'], fg=self.colors['text_primary'],
                      font=('Segoe UI', 18, 'bold'))
        title_label.pack(pady=(14, 2))

        credit_label = tk.Label(header, text="Created by Galih Prima",
                            bg=self.colors['gradient_start'], fg=self.colors['text_secondary'],
                            font=('Segoe UI', 9))
        credit_label.pack()
        
        # Tab Navigation
        tab_frame = tk.Frame(main_container, bg=self.colors['bg_main'])
        tab_frame.pack(fill=tk.X, padx=16, pady=(16, 0))
        
        

        self.active_tab = 'cable'
        self.tab_buttons = {}

        tabs = [
            ('cable',   'Cable Name\nGenerator'),
            ('ci',      'Cluster\nDescription'),
            ('feeder',  'Feeder\nDescription'),
            ('hp',      'HP Grouping\nby FAT for KMZ'),
            ('csv',     'CSV → KML\nConverter'),
            ('counter', 'Homepass\nCounter'),
            ('kmlext',  'KML Extractor\nFor HPDB'),
            ('boq',     'BoQ Generator\nFor FDDP'),
        ]

        tab_row = tk.Frame(tab_frame, bg=self.colors['bg_main'])
        tab_row.pack(anchor='w')

        for tab_id, tab_name in tabs:
            btn = tk.Label(tab_row, text=tab_name,
                           font=('Segoe UI', 8),
                           padx=14, pady=5,
                           justify='center',
                           cursor='hand2')
            btn.pack(side=tk.LEFT, padx=(0, 6))
            btn.bind('<Button-1>', lambda e, t=tab_id: self.switch_tab(t))
            self.tab_buttons[tab_id] = btn

        
        # Content area
        self.content_frame = tk.Frame(main_container, bg=self.colors['bg_main'])
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)
        
        # Create all tab panels
        self.panels = {}
        self.panels['cable'] = self.create_cable_panel()
        self.panels['ci'] = self.create_ci_panel()
        self.panels['feeder'] = self.create_feeder_panel()
        self.panels['hp'] = self.create_hp_panel()
        self.panels['csv'] = self.create_csv_panel()
        self.panels['counter'] = self.create_counter_panel()
        self.panels['kmlext'] = self.create_kmlext_panel()
        self.panels['boq'] = self.create_boq_panel()
        
        # Show initial tab
        self.switch_tab('cable')
    
    def switch_tab(self, tab_id):
        """Switch between tabs"""
        self.active_tab = tab_id
        
        # Update tab button styles
        for tid, btn in self.tab_buttons.items():
            if tid == tab_id:
                btn.configure(bg=self.colors['button_blue'], fg='white')
            else:
                btn.configure(bg=self.colors['bg_section'], fg=self.colors['text_secondary'])
        
        # Show/hide panels
        for pid, panel in self.panels.items():
            if pid == tab_id:
                panel.pack(fill=tk.BOTH, expand=True)
            else:
                panel.pack_forget()
    
    def create_input_section(self, parent, title="Input Parameters"):
        """Create styled input section"""
        section = tk.Frame(parent, bg=self.colors['bg_section'], 
                          highlightbackground=self.colors['border'],
                          highlightthickness=1)
        section.pack(fill=tk.BOTH, expand=True, padx=(0, 12))
        
        # Header
        header = tk.Label(section, text=title, bg=self.colors['bg_section'],
                         fg=self.colors['text_primary'], font=('Segoe UI', 9),
                         anchor='w')
        header.pack(fill=tk.X, padx=16, pady=(12, 8))
        
        separator = tk.Frame(section, bg=self.colors['border'], height=1)
        separator.pack(fill=tk.X, padx=16, pady=(0, 12))
        
        # Input container
        input_container = tk.Frame(section, bg=self.colors['bg_section'])
        input_container.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 32))
        
        return input_container
    
    def create_output_section(self, parent):
        """Create styled output section"""
        section = tk.Frame(parent, bg=self.colors['bg_section'],
                          highlightbackground=self.colors['border'],
                          highlightthickness=1)
        section.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header = tk.Label(section, text="Result Output", bg=self.colors['bg_section'],
                         fg=self.colors['text_primary'], font=('Segoe UI', 9),
                         anchor='w')
        header.pack(fill=tk.X, padx=16, pady=(12, 8))
        
        separator = tk.Frame(section, bg=self.colors['border'], height=1)
        separator.pack(fill=tk.X, padx=16, pady=(0, 12))
        
        # Output text
        output_text = tk.Text(section, bg=self.colors['output_bg'],
                             fg=self.colors['output_text'],
                             font=('Consolas', 9),
                             highlightbackground=self.colors['output_border'],
                             highlightthickness=1,
                             relief=tk.FLAT,
                             padx=16, pady=16,
                             wrap=tk.WORD,
                             state='disabled')
        output_text.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 60))
        
        return output_text
    
    def create_input_field(self, parent, label_text, row):
        """Create styled input field"""
        label = tk.Label(parent, text=label_text, bg=self.colors['bg_section'],
                        fg=self.colors['text_primary'], font=('Segoe UI', 9),
                        anchor='w')
        label.grid(row=row, column=0, sticky='w', pady=(0, 12))
        
        entry = tk.Entry(parent, bg=self.colors['bg_input'],
                        fg=self.colors['text_primary'],
                        font=('Segoe UI', 9),
                        relief=tk.FLAT,
                        highlightbackground=self.colors['border'],
                        highlightthickness=1,
                        insertbackground=self.colors['text_primary'])
        entry.grid(row=row+1, column=0, sticky='ew', pady=(0, 12), ipady=6)
        
        return entry
    
    def create_cable_panel(self):
        """Cable Name Generator Panel"""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])
        
        # Two column layout
        left_col = tk.Frame(panel, bg=self.colors['bg_main'])
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))
        
        # Pack buttons FIRST at bottom so they always get full height
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(8, 0))
        
        ModernButton(btn_frame, "GENERATE", self.generate_cable,
                    self.colors['button_blue'], self.colors['button_blue_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        
        ModernButton(btn_frame, "RESET", self.reset_cable,
                    self.colors['button_gray'], self.colors['button_gray_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)
        
        ModernButton(btn_frame, "COPY", lambda: self.copy_result(self.cable_output),
                    self.colors['button_green'], self.colors['button_green_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        
        # Input section fills remaining space above buttons
        input_container = self.create_input_section(left_col)
        input_container.grid_columnconfigure(0, weight=1)
        
        # Cable Category
        category_label = tk.Label(input_container, text="Cable Category:",
                                 bg=self.colors['bg_section'],
                                 fg=self.colors['text_primary'],
                                 font=('Segoe UI', 9), anchor='w')
        category_label.grid(row=0, column=0, sticky='w', pady=(0, 8))
        
        category_frame = tk.Frame(input_container, bg=self.colors['bg_section'])
        category_frame.grid(row=1, column=0, sticky='ew', pady=(0, 12))
        
        self.cable_category = tk.StringVar(value='cluster')
        self.category_buttons = {}
        
        for value, text in [('cluster', 'Cluster Cable'), ('feeder', 'Feeder Cable')]:
            btn = tk.Label(category_frame, text=text,
                          font=('Segoe UI', 9),
                          cursor='hand2',
                          padx=16, pady=8)
            btn.pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
            btn.bind('<Button-1>', lambda e, v=value: self.select_cable_category(v))
            self.category_buttons[value] = btn
        
        # Dynamic input fields
        self.cable_inputs_frame = tk.Frame(input_container, bg=self.colors['bg_section'])
        self.cable_inputs_frame.grid(row=2, column=0, sticky='ew')
        self.cable_inputs_frame.grid_columnconfigure(0, weight=1)
        
        self.cable_entries = {}
        
        # Initialize with cluster category
        self.select_cable_category('cluster')
        
        # Output section
        self.cable_output = self.create_output_section(right_col)
        
        return panel
    
    def select_cable_category(self, category):
        """Select cable category and update button colors"""
        self.cable_category.set(category)
        
        # Update button colors
        for cat, btn in self.category_buttons.items():
            if cat == category:
                btn.configure(bg=self.colors['button_green'], fg='white')
            else:
                btn.configure(bg=self.colors['bg_input'], fg=self.colors['text_secondary'])
        
        self.update_cable_fields()
    
    def update_cable_fields(self):
        """Update cable input fields based on category"""
        for widget in self.cable_inputs_frame.winfo_children():
            widget.destroy()
        
        self.cable_entries.clear()
        category = self.cable_category.get()
        
        row = 0
        
        if category == 'feeder':
            self.cable_entries['olt'] = self.create_input_field(
                self.cable_inputs_frame, "OLT Code:", row)
            row += 2
        
        self.cable_entries['fdt'] = self.create_input_field(
            self.cable_inputs_frame, "FDT Code:", row)
        row += 2
        
        if category == 'cluster':
            self.cable_entries['line'] = self.create_input_field(
                self.cable_inputs_frame, "Line Code:", row)
            row += 2
        
        # Feeder Type dropdown (only for feeder)
        if category == 'feeder':
            label = tk.Label(self.cable_inputs_frame, text="Feeder Type:",
                            bg=self.colors['bg_section'],
                            fg=self.colors['text_primary'],
                            font=('Segoe UI', 9), anchor='w')
            label.grid(row=row, column=0, sticky='w', pady=(0, 12))
            
            feeder_types = ['SUBFEEDER', 'HUBFEEDER', 'MAINFEEDER']
            self.feeder_type_var = tk.StringVar(value=feeder_types[0])
            
            combo_feeder = ttk.Combobox(self.cable_inputs_frame, textvariable=self.feeder_type_var,
                                values=feeder_types, state='readonly', font=('Segoe UI', 9))
            combo_feeder.grid(row=row+1, column=0, sticky='ew', pady=(0, 12), ipady=4)
            row += 2
        
        # Cable Type dropdown
        label = tk.Label(self.cable_inputs_frame, text="Cable Type:",
                        bg=self.colors['bg_section'],
                        fg=self.colors['text_primary'],
                        font=('Segoe UI', 9), anchor='w')
        label.grid(row=row, column=0, sticky='w', pady=(0, 12))
        
        cable_types = ['24C/2T', '36C/3T', '48C/4T'] if category == 'cluster' else \
                     ['24C/2T', '48C/4T', '96C/8T', '144C/12T', '288C/24T']
        
        self.cable_type_var = tk.StringVar(value=cable_types[0])
        
        combo = ttk.Combobox(self.cable_inputs_frame, textvariable=self.cable_type_var,
                            values=cable_types, state='readonly', font=('Segoe UI', 9))
        combo.grid(row=row+1, column=0, sticky='ew', pady=(0, 12), ipady=4)
        row += 2
        
        self.cable_entries['length'] = self.create_input_field(
            self.cable_inputs_frame, "Length by OTDR (m):", row)
    
    def generate_cable(self):
        """Generate cable name"""
        category = self.cable_category.get()
        
        try:
            if category == 'cluster':
                fdt = self.cable_entries['fdt'].get().strip()
                line = self.cable_entries['line'].get().strip()
                ctype = self.cable_type_var.get()
                length = self.cable_entries['length'].get().strip()
                
                result = f"{fdt} - CABLE LINE {line} (FO {ctype}) - AE - {length} M"
            else:
                olt = self.cable_entries['olt'].get().strip()
                fdt = self.cable_entries['fdt'].get().strip()
                feeder_type = self.feeder_type_var.get()
                ctype = self.cable_type_var.get()
                length = self.cable_entries['length'].get().strip()
                
                result = f"{olt} - {fdt} ({feeder_type} CABLE FO {ctype}) - AE - {length} M"
            
            self.set_output(self.cable_output, result.upper())
        except (KeyError, ValueError) as e:
            self.set_output(self.cable_output, f"Error: {str(e)}")
    
    def reset_cable(self):
        """Reset cable inputs"""
        for entry in self.cable_entries.values():
            if isinstance(entry, tk.Entry):
                entry.delete(0, tk.END)
        self.set_output(self.cable_output, "")
    
    def create_ci_panel(self):
        """CI Description Generator Panel"""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])
        
        # Two column layout
        left_col = tk.Frame(panel, bg=self.colors['bg_main'])
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))
        
        # Input section
        input_container = self.create_input_section(left_col)
        input_container.grid_columnconfigure(0, weight=1)
        
        self.ci_entries = {}
        row = 0
        
        for key, label in [('route', 'Route (m):'), ('fdt', 'Slack FDT (unit):'),
                          ('fat', 'Slack FAT (unit):'), ('otdr', 'By OTDR (m):')]:
            self.ci_entries[key] = self.create_input_field(input_container, label, row)
            row += 2
        
        # Buttons
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(fill=tk.X, pady=(24, 0))
        
        ModernButton(btn_frame, "GENERATE", self.generate_ci,
                    self.colors['button_blue'], self.colors['button_blue_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        
        ModernButton(btn_frame, "RESET", self.reset_ci,
                    self.colors['button_gray'], self.colors['button_gray_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)
        
        ModernButton(btn_frame, "COPY", lambda: self.copy_result(self.ci_output),
                    self.colors['button_green'], self.colors['button_green_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        
        # Output section
        self.ci_output = self.create_output_section(right_col)
        
        return panel
    
    def generate_ci(self):
        """Generate CI description"""
        try:
            route = float(self.ci_entries['route'].get() or 0)
            fdt = float(self.ci_entries['fdt'].get() or 0)
            fat = float(self.ci_entries['fat'].get() or 0)
            otdr = self.ci_entries['otdr'].get() or '0'
            
            total_slack = fdt + fat
            route_plus_slack = route + (total_slack * 20)
            total_length = math.ceil(route_plus_slack * 1.05)
            
            result = f"""Total Route : {route} m
Total Slack : {total_slack} unit ({fdt} slack FDT & {fat} slack FAT) @20 m
Toleransi : 5%
Total Length Cable : {route} + {total_slack * 20} = {route_plus_slack} m + ({route_plus_slack} m x 5%) = {total_length} m
By OTDR : {otdr} m"""
            
            self.set_output(self.ci_output, result.upper())
        except (ValueError, KeyError) as e:
            self.set_output(self.ci_output, f"Error: {str(e)}")
    
    def reset_ci(self):
        """Reset CI inputs"""
        for entry in self.ci_entries.values():
            entry.delete(0, tk.END)
        self.set_output(self.ci_output, "")
    
    def create_feeder_panel(self):
        """Feeder Description Generator Panel"""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])
        
        # Two column layout
        left_col = tk.Frame(panel, bg=self.colors['bg_main'])
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))
        
        # Input section
        input_container = self.create_input_section(left_col)
        input_container.grid_columnconfigure(0, weight=1)
        
        self.feeder_entries = {}
        row = 0
        
        for key, label in [('route', 'Route (m):'), ('slack', 'Slack (unit):'),
                          ('otdr', 'By OTDR (m):')]:
            self.feeder_entries[key] = self.create_input_field(input_container, label, row)
            row += 2
        
        # Buttons
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(fill=tk.X, pady=(24, 0))
        
        ModernButton(btn_frame, "GENERATE", self.generate_feeder,
                    self.colors['button_blue'], self.colors['button_blue_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        
        ModernButton(btn_frame, "RESET", self.reset_feeder,
                    self.colors['button_gray'], self.colors['button_gray_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)
        
        ModernButton(btn_frame, "COPY", lambda: self.copy_result(self.feeder_output),
                    self.colors['button_green'], self.colors['button_green_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        
        # Output section
        self.feeder_output = self.create_output_section(right_col)
        
        return panel
    
    def generate_feeder(self):
        """Generate feeder description"""
        try:
            route = float(self.feeder_entries['route'].get() or 0)
            slack = float(self.feeder_entries['slack'].get() or 0)
            otdr = self.feeder_entries['otdr'].get() or '0'
            
            route_plus_slack = route + (slack * 20)
            total_length = math.ceil(route_plus_slack * 1.05)
            
            result = f"""Total Route : {route} m
Total Slack : {slack} unit @20 m
Toleransi : 5%
Total Length Cable : {route} + {slack * 20} = {route_plus_slack} m + ({route_plus_slack} m x 5%) = {total_length} m
By OTDR : {otdr} m"""
            
            self.set_output(self.feeder_output, result.upper())
        except (ValueError, KeyError) as e:
            self.set_output(self.feeder_output, f"Error: {str(e)}")
    
    def reset_feeder(self):
        """Reset feeder inputs"""
        for entry in self.feeder_entries.values():
            entry.delete(0, tk.END)
        self.set_output(self.feeder_output, "")
    
    def set_output(self, text_widget, content):
        """Set text in output widget"""
        text_widget.config(state='normal')
        text_widget.delete(1.0, tk.END)
        text_widget.insert(1.0, content if content else "FTTH Automation for EMR Project Ready...")
        text_widget.config(state='disabled')
    
    def copy_result(self, text_widget):
        """Copy result to clipboard"""
        content = text_widget.get(1.0, tk.END).strip()
        if content and content != "FTTH Automation for EMR Project Ready...":
            self.root.clipboard_clear()
            self.root.clipboard_append(content)

    # =========================
    # TAB BARU: HOME PASS GROUPING (KML)
    # =========================

    def create_hp_panel(self):
        """Homepass Grouping Panel (standalone, no QGIS)"""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])

        # Two column layout (mirip tab lain)
        left_col = tk.Frame(panel, bg=self.colors['bg_main'])
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        input_container = self.create_input_section(left_col, title="KML Grouping Parameters")
        input_container.grid_columnconfigure(0, weight=1)

        # Vars
        self.hp_points_path = tk.StringVar()
        self.hp_polys_path = tk.StringVar()
        self.hp_out_path = tk.StringVar()
        self.hp_use_border = tk.BooleanVar(value=False)

        row = 0
        row = self._hp_file_picker(input_container, "Homepass KML (Points):", self.hp_points_path, row)
        row = self._hp_file_picker(input_container, "Boundary FAT KML (Polygons):", self.hp_polys_path, row)
        row = self._hp_save_picker(input_container, "Output KML (Single file):", self.hp_out_path, row)

        chk = tk.Checkbutton(
            input_container,
            text="Border tolerant (anggap titik di garis boundary = masuk FAT)",
            variable=self.hp_use_border,
            bg=self.colors['bg_section'],
            fg=self.colors['text_secondary'],
            activebackground=self.colors['bg_section'],
            activeforeground=self.colors['text_primary'],
            selectcolor=self.colors['bg_input']
        )
        chk.grid(row=row, column=0, sticky='w', pady=(8, 0))
        row += 1

        # Buttons
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(fill=tk.X, pady=(24, 0))

        ModernButton(btn_frame, "RUN GROUPING", self.run_hp_grouping,
                    self.colors['button_blue'], self.colors['button_blue_hover'],
                    width=180).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))

        ModernButton(btn_frame, "RESET", self.reset_hp_grouping,
                    self.colors['button_gray'], self.colors['button_gray_hover'],
                    width=150).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)

        ModernButton(btn_frame, "OPEN OUTPUT FOLDER", self.open_hp_output_folder,
                    self.colors['button_green'], self.colors['button_green_hover'],
                    width=200).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))

        # Output section
        self.hp_output = self.create_output_section(right_col)
        self.set_output(self.hp_output, "Pilih 2 file KML (homepass points + boundary FAT polygons), lalu RUN GROUPING.")

        return panel

    def _hp_file_picker(self, parent, label, var, row):
        tk.Label(parent, text=label, bg=self.colors['bg_section'],
                fg=self.colors['text_primary'], font=('Segoe UI', 9),
                anchor='w').grid(row=row, column=0, sticky='w', pady=(0, 8))
        row += 1

        frame = tk.Frame(parent, bg=self.colors['bg_section'])
        frame.grid(row=row, column=0, sticky='ew', pady=(0, 12))
        frame.grid_columnconfigure(0, weight=1)

        entry = tk.Entry(frame, textvariable=var, bg=self.colors['bg_input'],
                        fg=self.colors['text_primary'], font=('Segoe UI', 9),
                        relief=tk.FLAT, highlightbackground=self.colors['border'],
                        highlightthickness=1, insertbackground=self.colors['text_primary'])
        entry.grid(row=0, column=0, sticky='ew', ipady=6)

        def browse():
            path = filedialog.askopenfilename(
                title=label,
                filetypes=[("KML files", "*.kml"), ("All files", "*.*")]
            )
            if path:
                var.set(path)

        ModernButton(frame, "BROWSE", browse,
                    self.colors['button_gray'], self.colors['button_gray_hover'],
                    width=90).grid(row=0, column=1, padx=(8, 0))

        return row + 1

    def _hp_save_picker(self, parent, label, var, row):
        tk.Label(parent, text=label, bg=self.colors['bg_section'],
                fg=self.colors['text_primary'], font=('Segoe UI', 9),
                anchor='w').grid(row=row, column=0, sticky='w', pady=(0, 8))
        row += 1

        frame = tk.Frame(parent, bg=self.colors['bg_section'])
        frame.grid(row=row, column=0, sticky='ew', pady=(0, 12))
        frame.grid_columnconfigure(0, weight=1)

        entry = tk.Entry(frame, textvariable=var, bg=self.colors['bg_input'],
                        fg=self.colors['text_primary'], font=('Segoe UI', 9),
                        relief=tk.FLAT, highlightbackground=self.colors['border'],
                        highlightthickness=1, insertbackground=self.colors['text_primary'])
        entry.grid(row=0, column=0, sticky='ew', ipady=6)

        def browse():
            path = filedialog.asksaveasfilename(
                title=label,
                defaultextension=".kml",
                filetypes=[("KML files", "*.kml")]
            )
            if path:
                var.set(path)

        ModernButton(frame, "SAVE AS", browse,
                    self.colors['button_gray'], self.colors['button_gray_hover'],
                    width=90).grid(row=0, column=1, padx=(8, 0))

        return row + 1

    def reset_hp_grouping(self):
        self.hp_points_path.set("")
        self.hp_polys_path.set("")
        self.hp_out_path.set("")
        self.hp_use_border.set(False)
        self.set_output(self.hp_output, "Pilih 2 file KML (homepass points + boundary FAT polygons), lalu RUN GROUPING.")

    def open_hp_output_folder(self):
        out_path = self.hp_out_path.get().strip()
        if not out_path:
            messagebox.showinfo("Info", "Output belum dipilih.")
            return
        folder = os.path.dirname(out_path)
        if folder and os.path.isdir(folder):
            try:
                if sys.platform == 'win32':
                    os.startfile(folder)
                elif sys.platform == 'darwin':
                    subprocess.run(['open', folder])
                else:
                    subprocess.run(['xdg-open', folder])
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def run_hp_grouping(self):
        points_path = self.hp_points_path.get().strip()
        polys_path = self.hp_polys_path.get().strip()
        out_path = self.hp_out_path.get().strip()
        include_border = bool(self.hp_use_border.get())

        if not points_path or not os.path.isfile(points_path):
            messagebox.showerror("Error", "Homepass KML (Points) belum benar.")
            return
        if not polys_path or not os.path.isfile(polys_path):
            messagebox.showerror("Error", "Boundary FAT KML (Polygons) belum benar.")
            return
        if not out_path:
            messagebox.showerror("Error", "Output KML belum dipilih.")
            return

        try:
            self.set_output(self.hp_output, "Processing...\n- Parse polygons\n- Parse points\n- Grouping\n- Export single KML\n")
            self.root.update_idletasks()

            fat_polygons = self._parse_kml_polygons(polys_path)  # [(fat_name, [(lon,lat),...]),...]
            points = self._parse_kml_points(points_path)         # [(point_name, lon, lat),...]

            grouped = self._group_points(points, fat_polygons, include_border=include_border)

            self._write_single_kml_with_folders(grouped, out_path, root_folder_name="HOME PASS by FAT")

            total_points = sum(len(v) for v in grouped.values())
            fat_count = len(grouped)
            no_fat = len(grouped.get("NO_FAT", []))

            msg = (
                f"DONE ✅\n\n"
                f"Output: {out_path}\n"
                f"Total Homepass: {total_points}\n"
                f"Total FAT Folder: {fat_count}\n"
                f"NO_FAT (tidak masuk polygon): {no_fat}\n\n"
                f"Tips: buka KML di Google Earth → folder induk → subfolder per FAT."
            )
            self.set_output(self.hp_output, msg)

        except Exception as e:
            self.set_output(self.hp_output, f"ERROR:\n{str(e)}")
    
    # ---------- KML parsing ----------

    def _kml_ns(self, root):
        # returns namespace like "{http://www.opengis.net/kml/2.2}" or ""
        if root.tag.startswith("{") and "}" in root.tag:
            return root.tag.split("}")[0] + "}"
        return ""

    def _parse_kml_points(self, kml_path):
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = self._kml_ns(root)

        points = []
        for pm in root.findall(f".//{ns}Placemark"):
            name_el = pm.find(f"{ns}name")
            pname = name_el.text.strip() if (name_el is not None and name_el.text) else ""

            coord_el = pm.find(f".//{ns}Point/{ns}coordinates")
            if coord_el is None or not coord_el.text:
                continue

            coord_text = coord_el.text.strip()
            first = coord_text.split()[0]  # lon,lat,alt
            parts = first.split(",")
            if len(parts) < 2:
                continue
            lon = float(parts[0])
            lat = float(parts[1])
            points.append((pname, lon, lat))

        return points

    def _parse_kml_polygons(self, kml_path):
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = self._kml_ns(root)

        polygons = []
        for pm in root.findall(f".//{ns}Placemark"):
            name_el = pm.find(f"{ns}name")
            fat_name = name_el.text.strip() if (name_el is not None and name_el.text) else "UNKNOWN"

            coord_el = pm.find(f".//{ns}Polygon/{ns}outerBoundaryIs/{ns}LinearRing/{ns}coordinates")
            if coord_el is None or not coord_el.text:
                continue

            coords = []
            for part in coord_el.text.strip().split():
                p = part.split(",")
                if len(p) < 2:
                    continue
                coords.append((float(p[0]), float(p[1])))

            # polygon minimal 4 titik (awal=akhir)
            if len(coords) >= 4:
                polygons.append((fat_name, coords))

        return polygons

    # ---------- Geometry (point in polygon) ----------

    def _point_on_segment(self, p, a, b, eps=1e-12):
        (x, y) = p
        (x1, y1) = a
        (x2, y2) = b
        cross = (x - x1) * (y2 - y1) - (y - y1) * (x2 - x1)
        if abs(cross) > eps:
            return False
        dot = (x - x1) * (x2 - x1) + (y - y1) * (y2 - y1)
        if dot < -eps:
            return False
        sq_len = (x2 - x1) ** 2 + (y2 - y1) ** 2
        if dot - sq_len > eps:
            return False
        return True

    def _point_in_poly(self, point, poly, include_border=False):
        # Ray casting
        x, y = point
        inside = False
        n = len(poly)
        # iterate edges (assume poly closed or not; we handle by looping n)
        for i in range(n):
            a = poly[i]
            b = poly[(i + 1) % n]

            if include_border and self._point_on_segment((x, y), a, b):
                return True

            xi, yi = a
            xj, yj = b

            # check ray intersection
            if ((yi > y) != (yj > y)):
                denom = (yj - yi) if (yj - yi) != 0 else 1e-30
                x_int = (xj - xi) * (y - yi) / denom + xi
                if x < x_int:
                    inside = not inside

        return inside

    def _group_points(self, points, fat_polygons, include_border=False):
        grouped = {}
        for pname, lon, lat in points:
            assigned = "NO_FAT"
            for fat_name, poly_coords in fat_polygons:
                if self._point_in_poly((lon, lat), poly_coords, include_border=include_border):
                    assigned = str(fat_name)
                    break
            grouped.setdefault(assigned, []).append((pname, lon, lat))
        return grouped

    # ---------- KML writing ----------

    def _kml_escape(self, s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _write_single_kml_with_folders(self, grouped, out_path, root_folder_name="HOME PASS by FAT"):
        out_dir = os.path.dirname(out_path)
        if out_dir and not os.path.isdir(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with open(out_path, "w", encoding="utf-8") as f:
            f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
            f.write('<kml xmlns="http://www.opengis.net/kml/2.2">\n')
            f.write("  <Document>\n")
            f.write(f"    <name>{self._kml_escape(root_folder_name)}</name>\n")
            f.write(f"    <description>Generated: {self._kml_escape(now)}</description>\n")
            f.write("    <Folder>\n")
            f.write(f"      <name>{self._kml_escape(root_folder_name)}</name>\n")

            # Sort FAT names; put NO_FAT last
            keys = sorted([k for k in grouped.keys() if k != "NO_FAT"], key=lambda x: str(x))
            if "NO_FAT" in grouped:
                keys.append("NO_FAT")

            for fat in keys:
                f.write("      <Folder>\n")
                f.write(f"        <name>{self._kml_escape(str(fat))}</name>\n")

                for pname, lon, lat in grouped[fat]:
                    f.write("        <Placemark>\n")
                    if pname:
                        f.write(f"          <name>{self._kml_escape(pname)}</name>\n")
                    f.write("          <Point>\n")
                    f.write(f"            <coordinates>{lon},{lat},0</coordinates>\n")
                    f.write("          </Point>\n")
                    f.write("        </Placemark>\n")

                f.write("      </Folder>\n")

            f.write("    </Folder>\n")
            f.write("  </Document>\n")
            f.write("</kml>\n")

    # =========================
    # TAB BARU: CSV TO KML CONVERTER (KML)
    # =========================

    def create_csv_panel(self):
        """CSV to KML Converter Panel"""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])

        # Two column layout
        left_col = tk.Frame(panel, bg=self.colors['bg_main'])
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        # Input section
        input_container = self.create_input_section(left_col, title="CSV to KML")
        input_container.grid_columnconfigure(0, weight=1)

        self.csv_path = tk.StringVar()

        row = 0

        # File picker
        tk.Label(input_container, text="CSV File:",
                 bg=self.colors['bg_section'], fg=self.colors['text_primary'],
                 font=('Segoe UI', 9), anchor='w').grid(row=row, column=0, sticky='w', pady=(0, 8))
        row += 1

        file_frame = tk.Frame(input_container, bg=self.colors['bg_section'])
        file_frame.grid(row=row, column=0, sticky='ew', pady=(0, 12))
        file_frame.grid_columnconfigure(0, weight=1)

        tk.Entry(file_frame, textvariable=self.csv_path,
                 bg=self.colors['bg_input'], fg=self.colors['text_primary'],
                 font=('Segoe UI', 9), relief=tk.FLAT,
                 highlightbackground=self.colors['border'], highlightthickness=1,
                 insertbackground=self.colors['text_primary']).grid(row=0, column=0, sticky='ew', ipady=6)

        def browse():
            path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
            if path:
                self.csv_path.set(path)
                self._preview_csv(path)

        ModernButton(file_frame, "BROWSE", browse,
                     self.colors['button_gray'], self.colors['button_gray_hover'],
                     width=90).grid(row=0, column=1, padx=(8, 0))

        # Buttons
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(fill=tk.X, pady=(24, 0))

        ModernButton(btn_frame, "CONVERT", self.run_csv_to_kml,
                     self.colors['button_blue'], self.colors['button_blue_hover'],
                     width=150).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Output section
        self.csv_output = self.create_output_section(right_col)
        self.set_output(self.csv_output, "Load CSV lalu convert ke KML")

        return panel

    def _detect_delimiter(self, path):
        """Auto-detect CSV delimiter (comma or semicolon)"""
        with open(path, newline='', encoding='utf-8') as f:
            sample = f.read(2048)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            return dialect.delimiter
        except Exception:
            # Fallback: hitung mana yang lebih banyak
            return ';' if sample.count(';') > sample.count(',') else ','

    def _preview_csv(self, path):
        """Preview CSV content (name, latitude, longitude) in the output box"""
        try:
            delim = self._detect_delimiter(path)
            with open(path, newline='', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=delim)
                rows = list(reader)

            if not rows:
                self.set_output(self.csv_output, "CSV kosong atau tidak valid.")
                return

            # Cek apakah baris pertama adalah header (tidak mengandung angka di kolom 1/2)
            first = rows[0]
            try:
                float(first[1])
                has_header = False  # baris pertama sudah data
            except (ValueError, IndexError):
                has_header = True   # baris pertama adalah header

            data_rows = rows[1:] if has_header else rows

            # Header display
            separator = "-" * 58
            lines = [
                f"{'NO':<5} {'NAME':<26} {'LATITUDE':>12} {'LONGITUDE':>12}",
                separator
            ]

            for i, r in enumerate(data_rows, start=1):
                if len(r) < 3:
                    continue
                name = str(r[0]).strip()
                lat  = str(r[1]).strip()
                lon  = str(r[2]).strip()
                # Truncate nama panjang
                if len(name) > 24:
                    name = name[:22] + '..'
                lines.append(f"{i:<5} {name:<26} {lat:>12} {lon:>12}")

            lines.append(separator)
            lines.append(f"Delimiter: '{delim}'  |  Total: {len(data_rows)} baris")
            self.set_output(self.csv_output, '\n'.join(lines))
        except Exception as e:
            self.set_output(self.csv_output, f"Gagal baca CSV:\n{str(e)}")

    def run_csv_to_kml(self):
        path = self.csv_path.get().strip()
        if not os.path.isfile(path):
            messagebox.showerror("Error", "CSV tidak valid")
            return

        save = filedialog.asksaveasfilename(defaultextension=".kml", filetypes=[("KML", "*.kml")])
        if not save:
            return

        delim = self._detect_delimiter(path)
        count = 0
        with open(path, newline='', encoding='utf-8') as f, open(save, "w", encoding="utf-8") as out:
            reader = csv.reader(f, delimiter=delim)
            rows = list(reader)

            # Skip header jika ada
            try:
                float(rows[0][1])
                data_rows = rows       # baris pertama sudah data
            except (ValueError, IndexError):
                data_rows = rows[1:]   # skip header

            out.write('<?xml version="1.0" encoding="UTF-8"?>\n<kml xmlns="http://www.opengis.net/kml/2.2">\n<Document>\n')

            for r in data_rows:
                try:
                    name = str(r[0]).strip()
                    lat  = float(r[1])
                    lon  = float(r[2])
                    out.write(f"<Placemark><name>{name}</name><Point><coordinates>{lon},{lat},0</coordinates></Point></Placemark>\n")
                    count += 1
                except (IndexError, ValueError):
                    continue

            out.write("</Document></kml>")

        self.set_output(self.csv_output, f"Berhasil convert {count} titik\n{save}")

    # =========================
    # TAB BARU: HOMEPASS COUNTER
    # =========================

    def create_counter_panel(self):
        """Homepass Counter Panel"""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])

        # Two column layout
        left_col = tk.Frame(panel, bg=self.colors['bg_main'])
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        # Input section
        input_container = self.create_input_section(left_col, title="Count Placemark")
        input_container.grid_columnconfigure(0, weight=1)

        tk.Label(input_container, text="Paste KML Content:",
                 bg=self.colors['bg_section'], fg=self.colors['text_primary'],
                 font=('Segoe UI', 9), anchor='w').grid(row=0, column=0, sticky='w', pady=(0, 8))

        self.counter_text = tk.Text(
            input_container,
            bg=self.colors['bg_input'],
            fg=self.colors['text_primary'],
            font=('Consolas', 9),
            relief=tk.FLAT,
            highlightbackground=self.colors['border'],
            highlightthickness=1,
            insertbackground=self.colors['text_primary'],
            wrap=tk.WORD
        )
        self.counter_text.grid(row=1, column=0, sticky='nsew', pady=(0, 12))
        input_container.grid_rowconfigure(1, weight=1)

        # Auto-count saat konten berubah (paste/ketik)
        self.counter_text.bind('<<Modified>>', self._on_counter_modified)

        # Buttons
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(fill=tk.X, pady=(24, 0))

        ModernButton(btn_frame, "RESET", self.reset_counter,
                     self.colors['button_gray'], self.colors['button_gray_hover'],
                     width=150).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Output section
        self.counter_output = self.create_output_section(right_col)
        self.set_output(self.counter_output, "Paste KML disini, hasil otomatis muncul...")

        return panel

    def _on_counter_modified(self, event=None):
        """Triggered setiap kali isi counter_text berubah"""
        # Reset flag agar event bisa terpicu lagi
        self.counter_text.edit_modified(False)
        self.count_kml()

    def reset_counter(self):
        """Reset counter input dan output"""
        self.counter_text.edit_modified(False)
        self.counter_text.delete('1.0', tk.END)
        self.set_output(self.counter_output, "Paste KML disini, hasil otomatis muncul...")
    
    def count_kml(self):
        text = self.counter_text.get("1.0", tk.END)
        if not text.strip():
            self.set_output(self.counter_output, "Paste KML disini, hasil otomatis muncul...")
            return
        placemarks = len(re.findall(r"<Placemark", text, re.IGNORECASE))
        coords     = len(re.findall(r"<coordinates>", text, re.IGNORECASE))

        self.set_output(self.counter_output, f"RESULT\n\n    Placemark   : {placemarks}\n    Coordinates : {coords}")

    # =========================
    # TAB: KML EXTRACTOR
    # =========================

    def create_kmlext_panel(self):
        """KML Extractor Panel — extract Placemarks + export Excel with FAT highlight"""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])

        # ── Left column: controls ──
        left_col = tk.Frame(panel, bg=self.colors['bg_main'])
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, ipadx=0)
        left_col.configure(width=320)
        left_col.pack_propagate(False)

        input_container = self.create_input_section(left_col, title="KML Extractor")
        input_container.grid_columnconfigure(0, weight=1)

        self.kmlext_path = tk.StringVar()
        self.kmlext_status = tk.StringVar(value="Pilih file KML untuk mulai.")
        self.kmlext_rows_cache = []

        # File picker
        tk.Label(input_container, text="File KML:",
                 bg=self.colors['bg_section'], fg=self.colors['text_primary'],
                 font=('Segoe UI', 9), anchor='w').grid(row=0, column=0, sticky='w', pady=(0, 8))

        fp_frame = tk.Frame(input_container, bg=self.colors['bg_section'])
        fp_frame.grid(row=1, column=0, sticky='ew', pady=(0, 12))
        fp_frame.grid_columnconfigure(0, weight=1)

        tk.Entry(fp_frame, textvariable=self.kmlext_path,
                 bg=self.colors['bg_input'], fg=self.colors['text_primary'],
                 font=('Segoe UI', 9), relief=tk.FLAT,
                 highlightbackground=self.colors['border'], highlightthickness=1,
                 insertbackground=self.colors['text_primary']).grid(row=0, column=0, sticky='ew', ipady=6)

        ModernButton(fp_frame, "BROWSE", self._kmlext_pick_file,
                     self.colors['button_gray'], self.colors['button_gray_hover'],
                     width=90).grid(row=0, column=1, padx=(8, 0))

        # Status label
        tk.Label(input_container, textvariable=self.kmlext_status,
                 bg=self.colors['bg_section'], fg=self.colors['text_secondary'],
                 font=('Segoe UI', 8), anchor='w', wraplength=280, justify='left'
                 ).grid(row=2, column=0, sticky='w', pady=(8, 0))

        # Buttons
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(fill=tk.X, pady=(16, 0))

        ModernButton(btn_frame, "EXTRACT / RELOAD", self._kmlext_run_extract,
                     self.colors['button_blue'], self.colors['button_blue_hover'],
                     width=160).pack(fill=tk.X, pady=(0, 6))

        ModernButton(btn_frame, "EXPORT EXCEL (.xlsx)", self._kmlext_export_excel,
                     self.colors['button_green'], self.colors['button_green_hover'],
                     width=160).pack(fill=tk.X)

        # ── Right column: Treeview preview ──
        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        preview_section = tk.Frame(right_col, bg=self.colors['bg_section'],
                                   highlightbackground=self.colors['border'],
                                   highlightthickness=1)
        preview_section.pack(fill=tk.BOTH, expand=True)

        hdr = tk.Label(preview_section, text="Preview Placemarks",
                       bg=self.colors['bg_section'], fg=self.colors['text_primary'],
                       font=('Segoe UI', 9), anchor='w')
        hdr.pack(fill=tk.X, padx=16, pady=(12, 8))

        tk.Frame(preview_section, bg=self.colors['border'], height=1).pack(fill=tk.X, padx=16, pady=(0, 8))

        tree_frame = tk.Frame(preview_section, bg=self.colors['bg_section'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 12))

        # Style Treeview agar cocok dengan tema gelap
        style = ttk.Style()
        style.theme_use('default')
        style.configure('KMLExt.Treeview',
                        background=self.colors['output_bg'],
                        foreground=self.colors['output_text'],
                        fieldbackground=self.colors['output_bg'],
                        rowheight=22,
                        font=('Consolas', 8))
        style.configure('KMLExt.Treeview.Heading',
                        background=self.colors['bg_input'],
                        foreground=self.colors['text_primary'],
                        font=('Segoe UI', 8, 'bold'),
                        relief='flat')
        style.map('KMLExt.Treeview', background=[('selected', self.colors['button_blue'])])

        cols = ("FolderName", "Name", "Latitude", "Longitude", "GeometryType")
        self.kmlext_tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                        style='KMLExt.Treeview')

        col_conf = {
            "FolderName":    ("FAT ID",       116),
            "Name":          ("Home Number",   116),
            "Latitude":      ("Latitude",      116),
            "Longitude":     ("Longitude",     116),
            "GeometryType": ("Geometry",       116),
        }
        for c, (heading, width) in col_conf.items():
            self.kmlext_tree.heading(c, text=heading)
            self.kmlext_tree.column(c, anchor="center", width=width, stretch=True)

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.kmlext_tree.yview)
        xscroll = ttk.Scrollbar(tree_frame, orient="horizontal",
                                command=self.kmlext_tree.xview)
        self.kmlext_tree.configure(yscrollcommand=yscroll.set,
                                   xscrollcommand=xscroll.set)

        self.kmlext_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        return panel

    # ── KML Extractor helpers ──

    def _kmlext_strip_ns(self, tag: str) -> str:
        return tag.split("}", 1)[-1] if "}" in tag else tag

    def _kmlext_parse_first_coord(self, text: str):
        if not text or not text.strip():
            return None, None
        parts = re.split(r"\s+", text.strip())
        first = parts[0].strip()
        vals = first.split(",")
        if len(vals) < 2:
            return None, None
        try:
            return float(vals[0]), float(vals[1])  # lon, lat
        except ValueError:
            return None, None

    def _kmlext_extract_rows(self, kml_path: str):
        try:
            tree = ET.parse(kml_path)
        except Exception as e:
            raise RuntimeError(f"Gagal baca KML: {e}")

        root = tree.getroot()
        rows = []

        def traverse(node, folder_stack):
            tag = self._kmlext_strip_ns(node.tag)

            if tag == "Folder":
                folder_name = ""
                for child in node:
                    if self._kmlext_strip_ns(child.tag) == "name" and (child.text or "").strip():
                        folder_name = child.text.strip()
                        break
                new_stack = folder_stack[:]
                if folder_name:
                    new_stack.append(folder_name)
                for child in node:
                    traverse(child, new_stack)
                return

            if tag == "Placemark":
                placemark_name = ""
                coords_text = None
                geom_type = ""

                for child in node:
                    if self._kmlext_strip_ns(child.tag) == "name" and (child.text or "").strip():
                        placemark_name = child.text.strip()
                        break

                for child in node.iter():
                    ctag = self._kmlext_strip_ns(child.tag)
                    if ctag in ("Point", "LineString", "Polygon"):
                        geom_type = ctag
                    if ctag == "coordinates" and (child.text or "").strip():
                        coords_text = child.text.strip()
                        break

                lon, lat = self._kmlext_parse_first_coord(coords_text)
                subfolder = folder_stack[-1] if folder_stack else ""
                rows.append({
                    "FolderName": subfolder,
                    "Name": placemark_name,
                    "Latitude": lat,
                    "Longitude": lon,
                    "GeometryType": geom_type or "",
                })
                return

            for child in node:
                traverse(child, folder_stack)

        traverse(root, [])
        return rows

    def _kmlext_pick_file(self):
        path = filedialog.askopenfilename(
            title="Pilih file KML",
            filetypes=[("KML Files", "*.kml"), ("All Files", "*.*")]
        )
        if not path:
            return
        self.kmlext_path.set(path)
        self.kmlext_status.set(f"File: {os.path.basename(path)} | Extracting...")
        self.root.update_idletasks()
        self._kmlext_run_extract()

    def _kmlext_clear_tree(self):
        for item in self.kmlext_tree.get_children():
            self.kmlext_tree.delete(item)

    def _kmlext_run_extract(self):
        path = self.kmlext_path.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showerror("Error", "File KML belum dipilih atau tidak ditemukan.")
            return

        try:
            rows = self._kmlext_extract_rows(path)
        except Exception as e:
            messagebox.showerror("Gagal Extract", str(e))
            self.kmlext_status.set("Gagal extract.")
            return

        self.kmlext_rows_cache = rows
        self._kmlext_clear_tree()

        if not rows:
            self.kmlext_status.set("Tidak ada Placemark terdeteksi.")
            messagebox.showwarning("Kosong", "Tidak ada Placemark yang terdeteksi di KML ini.")
            return

        preview_limit = 8000
        for r in rows[:preview_limit]:
            self.kmlext_tree.insert("", "end", values=(
                r["FolderName"],
                r["Name"],
                "" if r["Latitude"] is None else r["Latitude"],
                "" if r["Longitude"] is None else r["Longitude"],
                r["GeometryType"],
            ))

        extra = f" (preview {preview_limit} dari {len(rows)})" if len(rows) > preview_limit else ""
        self.kmlext_status.set(
            f"✅ {len(rows)} Placemark terdeteksi{extra}. Siap export Excel."
        )

    def _kmlext_export_excel(self):
        if not _PANDAS_OK:
            messagebox.showerror(
                "Library Tidak Ada",
                "Fitur ini membutuhkan pandas dan openpyxl.\n"
                "Install dulu:\n  pip install pandas openpyxl"
            )
            return

        if not self.kmlext_rows_cache:
            messagebox.showinfo("Info", "Belum ada data. Pilih file KML dulu.")
            return

        src_path = self.kmlext_path.get().strip()
        base_name = os.path.splitext(os.path.basename(src_path))[0] if src_path else "kml"
        # Output selalu di folder yang sama dengan KML input, overwrite otomatis
        save_path = os.path.join(os.path.dirname(src_path), "KMLExtractorResult.xlsx")

        df = pd.DataFrame(
            self.kmlext_rows_cache,
            columns=["FolderName", "Name", "Latitude", "Longitude", "GeometryType"]
        )

        try:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Extract")
                ws = writer.book["Extract"]

                highlight_fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
                seen_fat = set()

                for row in range(2, ws.max_row + 1):
                    fat = ws[f"A{row}"].value
                    if fat is None:
                        continue
                    fat = str(fat).strip()
                    if not fat:
                        continue
                    if fat not in seen_fat:
                        seen_fat.add(fat)
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = highlight_fill

        except Exception as e:
            messagebox.showerror("Gagal Export", f"Gagal simpan Excel: {e}")
            return

        self.kmlext_status.set(f"Export sukses ✅: {os.path.basename(save_path)}")
        messagebox.showinfo(
            "Sukses",
            f"Excel tersimpan (overwrite):\n{save_path}"
        )


    # ─────────────────────────────────────────────────────────────────────────
    # BoQ Uploader Panel
    # ─────────────────────────────────────────────────────────────────────────

    def create_boq_panel(self):
        """BoQ Uploader Panel — load BoQ Excel, preview Material/Service, export."""
        panel = tk.Frame(self.content_frame, bg=self.colors['bg_main'])

        if not _BOQ_OK:
            err_lbl = tk.Label(panel,
                text="⚠  BoQ Generator tidak tersedia.\n"
                     "Pastikan dependensi (pandas, openpyxl) sudah terinstall.",
                bg=self.colors['bg_main'], fg='#f87171',
                font=('Segoe UI', 10), justify='center')
            err_lbl.pack(expand=True)
            return panel

        # ── state variables ──────────────────────────────────────────────────
        self.boq_path_var       = tk.StringVar()
        self.boq_cluster_var    = tk.StringVar(value="none")  # 'none' agar tidak tristate
        self.boq_homepass_var   = tk.StringVar(value="")
        self.boq_status_var     = tk.StringVar(value="Pilih file BoQ Excel untuk mulai.")
        self.boq_mat_df         = None
        self.boq_svc_df         = None
        self.boq_out_dir_var    = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop", "BOQ_OUTPUT"))

        # ── Layout: left (controls) + right (preview) ────────────────────────
        left_col = tk.Frame(panel, bg=self.colors['bg_main'], width=300)
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)
        left_col.pack_propagate(False)

        right_col = tk.Frame(panel, bg=self.colors['bg_main'])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        # ── Left: input section ──────────────────────────────────────────────
        inp = self.create_input_section(left_col, title="BoQ Uploader")
        inp.grid_columnconfigure(0, weight=1)

        row_i = 0

        # Cluster toggle — di atas file picker
        tk.Label(inp, text="Tipe BoQ:", bg=self.colors['bg_section'],
                 fg=self.colors['text_primary'], font=('Segoe UI', 9, 'bold'),
                 anchor='w').grid(row=row_i, column=0, sticky='w', pady=(0, 2))
        row_i += 1

        rb_frame = tk.Frame(inp, bg=self.colors['bg_section'])
        rb_frame.grid(row=row_i, column=0, sticky='w', pady=(0, 8))

        for val, lbl in [("tidak", "Feeder"), ("ya", "Cluster")]:
            rb = tk.Radiobutton(rb_frame, text=lbl, variable=self.boq_cluster_var, value=val,
                                bg=self.colors['bg_section'], fg=self.colors['text_primary'],
                                selectcolor=self.colors['bg_input'], activebackground=self.colors['bg_section'],
                                activeforeground=self.colors['text_primary'],
                                font=('Segoe UI', 9), cursor='hand2',
                                command=self._boq_update_state)
            rb.pack(side=tk.LEFT, padx=(0, 12))
        row_i += 1

        # File picker
        tk.Label(inp, text="File BoQ Excel:", bg=self.colors['bg_section'],
                 fg=self.colors['text_primary'], font=('Segoe UI', 9),
                 anchor='w').grid(row=row_i, column=0, sticky='w', pady=(0, 4))
        row_i += 1

        fp = tk.Frame(inp, bg=self.colors['bg_section'])
        fp.grid(row=row_i, column=0, sticky='ew', pady=(0, 10))
        fp.grid_columnconfigure(0, weight=1)

        self.boq_file_entry = tk.Entry(fp, textvariable=self.boq_path_var,
                 bg=self.colors['bg_input'], fg=self.colors['text_secondary'],
                 font=('Segoe UI', 8), relief=tk.FLAT,
                 highlightbackground=self.colors['border'], highlightthickness=1,
                 insertbackground=self.colors['text_primary'],
                 state='disabled')
        self.boq_file_entry.grid(row=0, column=0, sticky='ew', ipady=5)
        self.boq_browse_btn = ModernButton(fp, "BROWSE", self._boq_pick_file,
                      self.colors['button_gray'], self.colors['button_gray_hover'],
                      width=80)
        self.boq_browse_btn.grid(row=0, column=1, padx=(6, 0))
        self.boq_browse_btn.configure(state='disabled')
        row_i += 1

        # Homepass entry (cluster only)
        tk.Label(inp, text="Homepass (RT/RW Permit qty):", bg=self.colors['bg_section'],
                 fg=self.colors['text_secondary'], font=('Segoe UI', 8),
                 anchor='w').grid(row=row_i, column=0, sticky='w', pady=(0, 2))
        row_i += 1

        self.boq_hp_entry = tk.Entry(inp, textvariable=self.boq_homepass_var,
                                     bg=self.colors['bg_input'], fg=self.colors['text_primary'],
                                     font=('Segoe UI', 9), relief=tk.FLAT,
                                     highlightbackground=self.colors['border'], highlightthickness=1,
                                     insertbackground=self.colors['text_primary'],
                                     state='disabled')
        self.boq_hp_entry.grid(row=row_i, column=0, sticky='ew', ipady=5, pady=(0, 10))
        self.boq_homepass_var.trace_add("write", lambda *_: self._boq_update_state())
        row_i += 1

        # Status label
        self.boq_hint_lbl = tk.Label(inp, textvariable=self.boq_status_var,
                                     bg=self.colors['bg_section'], fg=self.colors['text_secondary'],
                                     font=('Segoe UI', 8), anchor='w',
                                     wraplength=260, justify='left')
        self.boq_hint_lbl.grid(row=row_i, column=0, sticky='w')
        row_i += 1

        # Output folder
        tk.Label(inp, text="Output Folder:", bg=self.colors['bg_section'],
                 fg=self.colors['text_primary'], font=('Segoe UI', 9),
                 anchor='w').grid(row=row_i, column=0, sticky='w', pady=(12, 2))
        row_i += 1

        of = tk.Frame(inp, bg=self.colors['bg_section'])
        of.grid(row=row_i, column=0, sticky='ew', pady=(0, 10))
        of.grid_columnconfigure(0, weight=1)
        tk.Entry(of, textvariable=self.boq_out_dir_var,
                 bg=self.colors['bg_input'], fg=self.colors['text_primary'],
                 font=('Segoe UI', 8), relief=tk.FLAT,
                 highlightbackground=self.colors['border'], highlightthickness=1,
                 insertbackground=self.colors['text_primary']
                 ).grid(row=0, column=0, sticky='ew', ipady=5)
        ModernButton(of, "BROWSE", self._boq_pick_out_dir,
                     self.colors['button_gray'], self.colors['button_gray_hover'],
                     width=80).grid(row=0, column=1, padx=(6, 0))
        row_i += 1

        # Action buttons
        btn_frame = tk.Frame(left_col, bg=self.colors['bg_main'])
        btn_frame.pack(fill=tk.X, pady=(12, 0))

        self.boq_load_btn = ModernButton(btn_frame, "LOAD BOQ", self._boq_load,
                                          self.colors['button_blue'], self.colors['button_blue_hover'],
                                          width=160)
        self.boq_load_btn.pack(fill=tk.X, pady=(0, 6))
        self.boq_load_btn.configure(state='disabled')

        self.boq_export_btn = ModernButton(btn_frame, "EXPORT EXCEL (.xlsx)", self._boq_export,
                                            self.colors['button_green'], self.colors['button_green_hover'],
                                            width=160)
        self.boq_export_btn.pack(fill=tk.X)
        self.boq_export_btn.configure(state='disabled')

        # ── Right: preview notebook tabs ─────────────────────────────────────
        # Material preview
        mat_sec = tk.Frame(right_col, bg=self.colors['bg_section'],
                           highlightbackground=self.colors['border'], highlightthickness=1)
        mat_sec.pack(fill=tk.BOTH, expand=True, pady=(0, 6))

        self.boq_lbl_mat = tk.Label(mat_sec, text="Preview Material", bg=self.colors['bg_section'],
                 fg=self.colors['text_primary'], font=('Segoe UI', 9), anchor='w')
        self.boq_lbl_mat.pack(fill=tk.X, padx=12, pady=(8, 4))
        tk.Frame(mat_sec, bg=self.colors['border'], height=1).pack(fill=tk.X, padx=12)

        self.boq_tree_mat = self._boq_make_tree(mat_sec)

        # Service preview
        svc_sec = tk.Frame(right_col, bg=self.colors['bg_section'],
                           highlightbackground=self.colors['border'], highlightthickness=1)
        svc_sec.pack(fill=tk.BOTH, expand=True, pady=(0, 6))

        self.boq_lbl_svc = tk.Label(svc_sec, text="Preview Service", bg=self.colors['bg_section'],
                 fg=self.colors['text_primary'], font=('Segoe UI', 9), anchor='w')
        self.boq_lbl_svc.pack(fill=tk.X, padx=12, pady=(8, 4))
        tk.Frame(svc_sec, bg=self.colors['border'], height=1).pack(fill=tk.X, padx=12)

        self.boq_tree_svc = self._boq_make_tree(svc_sec)

        # Warnings
        warn_sec = tk.Frame(right_col, bg=self.colors['bg_section'],
                            highlightbackground=self.colors['border'], highlightthickness=1)
        warn_sec.pack(fill=tk.BOTH, expand=False)
        warn_sec.configure(height=100)

        tk.Label(warn_sec, text="Warnings / Mismatch", bg=self.colors['bg_section'],
                 fg=self.colors['text_primary'], font=('Segoe UI', 9), anchor='w'
                 ).pack(fill=tk.X, padx=12, pady=(8, 4))
        tk.Frame(warn_sec, bg=self.colors['border'], height=1).pack(fill=tk.X, padx=12)

        self.boq_txt_warn = tk.Text(warn_sec, bg=self.colors['output_bg'],
                                    fg='#f87171',
                                    font=('Consolas', 8),
                                    highlightbackground=self.colors['output_border'],
                                    highlightthickness=1, relief=tk.FLAT,
                                    padx=10, pady=8, wrap=tk.WORD,
                                    height=5, state='disabled')
        self.boq_txt_warn.pack(fill=tk.BOTH, expand=True, padx=12, pady=(4, 10))

        # Pastikan state awal konsisten (load btn disabled sampai tipe dipilih)
        self._boq_update_state()

        return panel

    def _boq_make_tree(self, parent):
        """Create a styled Treeview with scrollbars inside parent."""
        style = ttk.Style()
        style.configure('BOQ.Treeview',
                        background=self.colors['output_bg'],
                        foreground=self.colors['output_text'],
                        fieldbackground=self.colors['output_bg'],
                        rowheight=20, font=('Consolas', 8))
        style.configure('BOQ.Treeview.Heading',
                        background=self.colors['bg_input'],
                        foreground=self.colors['text_primary'],
                        font=('Segoe UI', 8, 'bold'))
        style.map('BOQ.Treeview', background=[('selected', self.colors['button_blue'])])

        frm = tk.Frame(parent, bg=self.colors['bg_section'])
        frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=(4, 8))

        tree = ttk.Treeview(frm, style='BOQ.Treeview', show='headings')
        vsb = ttk.Scrollbar(frm, orient='vertical', command=tree.yview)
        hsb = ttk.Scrollbar(frm, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)
        return tree

    def _boq_update_state(self, *_):
        """Enable/disable Load BoQ, browse, dan homepass entry berdasarkan tipe BoQ."""
        choice = self.boq_cluster_var.get()

        if choice == "none":
            # Belum pilih tipe — kunci semua input
            self.boq_file_entry.configure(state='disabled')
            self.boq_browse_btn.configure(state='disabled')
            self.boq_load_btn.configure(state='disabled')
            self.boq_hp_entry.configure(state='disabled')
            self.boq_status_var.set("← Pilih tipe BoQ (Cluster / Feeder) dulu.")
            self.boq_hint_lbl.configure(fg='#f87171')

        elif choice == "ya":
            # Cluster — aktifkan browse & homepass
            self.boq_file_entry.configure(state='normal')
            self.boq_browse_btn.configure(state='normal')
            self.boq_hp_entry.configure(state='normal')
            hp_ok = bool(self.boq_homepass_var.get().strip())
            if hp_ok:
                self.boq_load_btn.configure(state='normal')
                self.boq_status_var.set("✓ Cluster — homepass terisi. Siap Load BoQ.")
                self.boq_hint_lbl.configure(fg='#4ade80')
            else:
                self.boq_load_btn.configure(state='disabled')
                self.boq_status_var.set("← Isi nilai Homepass dulu (RT/RW Permit qty).")
                self.boq_hint_lbl.configure(fg='#fb923c')

        else:  # tidak (Feeder)
            # Feeder — aktifkan browse, homepass tidak perlu
            self.boq_file_entry.configure(state='normal')
            self.boq_browse_btn.configure(state='normal')
            self.boq_homepass_var.set("")
            self.boq_hp_entry.configure(state='disabled')
            self.boq_load_btn.configure(state='normal')
            self.boq_status_var.set("✓ Feeder — siap Load BoQ.")
            self.boq_hint_lbl.configure(fg='#4ade80')

    def _boq_pick_file(self):
        path = filedialog.askopenfilename(
            title="Pilih file BoQ Excel",
            filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
        )
        if path:
            self.boq_path_var.set(path)

    def _boq_pick_out_dir(self):
        d = filedialog.askdirectory(title="Pilih folder output")
        if d:
            self.boq_out_dir_var.set(d)

    def _boq_fill_tree(self, tree, df, max_rows=300):
        tree.delete(*tree.get_children())
        tree['columns'] = []
        if df is None or df.empty:
            return
        cols = list(df.columns)
        # Hapus kolom ke-6 (index 5) jika ada — mengulang deskripsi
        if len(cols) > 5:
            cols = cols[:5] + cols[6:]
        tree['columns'] = cols

        # Kumpulkan semua nilai untuk auto-sizing
        row_vals_all = []
        for i, (_, row) in enumerate(df.iterrows()):
            if i >= max_rows:
                break
            vals = [str(row[c]) if row[c] is not None else '' for c in cols]
            row_vals_all.append(vals)
            tree.insert('', 'end', values=vals)

        # Auto-size tiap kolom berdasarkan konten
        char_px = 7  # approx px per karakter (Consolas 8pt)
        for ci, c in enumerate(cols):
            header_w = len(str(c)) * char_px + 24
            max_w = header_w
            for rv in row_vals_all:
                if ci < len(rv):
                    cw = len(rv[ci]) * char_px + 16
                    if cw > max_w:
                        max_w = cw
            tree.column(c, width=min(max_w, 380), anchor='w', minwidth=50)
            tree.heading(c, text=str(c), anchor='w')

    def _boq_fill_warnings(self, boq_path, missing_mat, missing_svc):
        self.boq_txt_warn.configure(state='normal')
        self.boq_txt_warn.delete('1.0', 'end')
        lines = [f"BoQ: {boq_path}", ""]
        if missing_mat:
            lines.append("MATERIAL — ada di BoQ (qty>0) tapi TIDAK ada di template:")
            lines.extend([f"  • {d}" for d in missing_mat])
            lines.append("")
        else:
            lines.append("MATERIAL: OK — semua item ada di template.")
            lines.append("")
        if missing_svc:
            lines.append("SERVICE — ada di BoQ (qty>0) tapi TIDAK ada di template:")
            lines.extend([f"  • {d}" for d in missing_svc])
        else:
            lines.append("SERVICE: OK — semua item ada di template.")
        self.boq_txt_warn.insert('1.0', "\n".join(lines))
        self.boq_txt_warn.configure(state='disabled')

    def _boq_load(self):
        path = self.boq_path_var.get().strip()
        if not path:
            messagebox.showwarning("File kosong", "Pilih file BoQ Excel terlebih dahulu.")
            return

        try:
            self.boq_status_var.set("Memproses BoQ…")
            self.root.update_idletasks()

            boq_mat, boq_svc = _boq_process_boq(path)
            mat_out, svc_out, missing_mat, missing_svc = _boq_build_outputs(boq_mat, boq_svc)

            # Inject homepass jika Cluster
            hp_str = self.boq_homepass_var.get().strip()
            if hp_str:
                try:
                    hp = float(hp_str)
                    if hp > 0:
                        svc_out = _boq_inject_homepass(svc_out, hp)
                except ValueError:
                    messagebox.showwarning("Homepass",
                        f"Nilai homepass '{hp_str}' bukan angka — diabaikan.")

            self.boq_mat_df = mat_out
            self.boq_svc_df = svc_out

            self._boq_fill_tree(self.boq_tree_mat, mat_out)
            self._boq_fill_tree(self.boq_tree_svc, svc_out)
            self._boq_fill_warnings(path, missing_mat, missing_svc)

            # Update header label dengan jumlah baris
            self.boq_lbl_mat.configure(text=f"Preview Material  ({len(mat_out)} baris)")
            self.boq_lbl_svc.configure(text=f"Preview Service  ({len(svc_out)} baris)")

            self.boq_export_btn.configure(state='normal')
            self.boq_status_var.set(
                f"✓ Selesai. Material: {len(mat_out)} baris | Service: {len(svc_out)} baris"
            )
            self.boq_hint_lbl.configure(fg='#4ade80')

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.boq_status_var.set("Error saat memproses BoQ.")
            self.boq_hint_lbl.configure(fg='#f87171')

    def _boq_export(self):
        if self.boq_mat_df is None or self.boq_svc_df is None:
            messagebox.showwarning("Belum siap", "Load BoQ terlebih dahulu.")
            return

        out_dir = self.boq_out_dir_var.get().strip()
        if not out_dir:
            messagebox.showwarning("Folder kosong", "Pilih folder output terlebih dahulu.")
            return

        try:
            self.boq_status_var.set("Mengekspor…")
            self.root.update_idletasks()

            out_mat = os.path.join(out_dir, "material_upload_FDDP.xlsx")
            out_svc = os.path.join(out_dir, "service_upload_FDDP.xlsx")

            _boq_write_value_only(_BOQMAT_B64, self.boq_mat_df, out_mat)
            _boq_write_value_only(_BOQSVC_B64, self.boq_svc_df, out_svc)

            self.boq_status_var.set("✓ Export selesai (overwrite).")
            self.boq_hint_lbl.configure(fg='#4ade80')
            messagebox.showinfo("Sukses",
                f"File berhasil disimpan (overwrite):\n• {out_mat}\n• {out_svc}")
        except Exception as e:
            messagebox.showerror("Export gagal", str(e))
            self.boq_status_var.set("Export gagal.")
            self.boq_hint_lbl.configure(fg='#f87171')


# ── BoQ Uploader: embedded logic (no external dependency) ───────────────────

# ---- Embedded Excel templates (base64) ----
_BOQMAT_B64 = """UEsDBBQABgAIAAAAIQBYcSD4iQEAACIGAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACslE1PAjEQhu8m/odNr4YteDDGsHBAPRklEX9AbWfZhm7bdArCv3e2fIQYZEPYyzb9mPd9Ojud4Xhdm2wFAbWzBRvkfZaBlU5pOy/Y1+y198gyjMIqYZyFgm0A2Xh0ezOcbTxgRtEWC1bF6J84R1lBLTB3HiztlC7UItI0zLkXciHmwO/7/QcunY1gYy82Gmw0fIZSLE3MXta0vCUJYJBlk+3BxqtgwnujpYhEyldW/XHp7RxyikxnsNIe7wiD8ZMOzc7/Bru4D0pN0AqyqQjxXdSEwdeG/7iw+HZukZ8XOUHpylJLUE4ua8pAjj6AUFgBxNrkacxroe2e+4x/Oow8DYOOQZr7JeEWDqAfFqwwb9oukB/PuiY61m6BilSEwNP3eook02KIcWMAu66FJNrmXIkA6jMGeq6dAxxrt3BIYeSkorrtOAkH3XP+9JimwXmkthLgcoB932iie56EIEQNh85x6gUeHKklXX3jprKtAnWpt1xidPXV9luZE+Y8dfjRLwAAAP//AwBQSwMEFAAGAAgAAAAhABNevmUCAQAA3wIAAAsACAJfcmVscy8ucmVscyCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACskk1LAzEQhu+C/yHMvTvbKiLSbC9F6E1k/QExmf1gN5mQpLr990ZBdKG2Hnqcr3eeeZn1ZrKjeKMQe3YSlkUJgpxm07tWwkv9uLgHEZNyRo3sSMKBImyq66v1M40q5aHY9T6KrOKihC4l/4AYdUdWxYI9uVxpOFiVchha9EoPqiVcleUdht8aUM00xc5ICDtzA6I++Lz5vDY3Ta9py3pvyaUjK5CmRM6QWfiQ2ULq8zWiVqGlJMGwfsrpiMr7ImMDHida/Z/o72vRUlJGJYWaA53m+ew4BbS8pEVzE3/cmUZ85zC8Mg+nWG4vyaL3MbE9Y85XzzcSzt6y+gAAAP//AwBQSwMEFAAGAAgAAAAhAEt69vaPAwAA4QgAAA8AAAB4bC93b3JrYm9vay54bWykVvFvozYU/n3S/gdm9VcKToEkqOSUBNhlau6yNGs3qVLkgFO8AmbGNFSn+9/3DCFtLqcp60WJjf3sz99773sm1x/qLNWeqSgZzz2EL02k0TziMcsfPfTHKtQHSCslyWOS8px66IWW6MPo55+ud1w8bTh/0gAgLz2USFm4hlFGCc1IeckLmoNly0VGJAzFo1EWgpK4TCiVWWr0TNMxMsJy1CK44hwMvt2yiPo8qjKayxZE0JRIoF8mrCg7tCw6By4j4qkq9IhnBUBsWMrkSwOKtCxyZ485F2STgts1trVawNeBHzah6XUngenkqIxFgpd8Ky8B2mhJn/iPTQPjoxDUpzE4D8kyBH1mKocHVsJ5JyvngOW8gmHzh9EwSKvRigvBeyeafeDWQ6PrLUvpXStdjRTFJ5KpTKVIS0kpg5hJGnuoD0O+o0cToiomFUvB2hs6vQEyRgc5LwQMIPfjVFKRE0mnPJcgtT31H5VVgz1NOIhYW9J/KiYo1A5ICNyBlkQu2ZQLIhOtEqmHfnUfwtkkWH78PA/0xfLzb8F09bAQ/G8aSW3+sqRFtUlZ9LCiWQFFQLXf5/7DG2GS0yr4H9IkkYqMAdFoGbfP30YGiAu3k99CCg2eZ/4NpOCWPENCIO3xvl5nEHF8tc4j4eL1l/5kOp0ObEvH476lW75l6UPTD3S/P7D6ljOwg+nVV3BGOG7ESSWTfa4VtIcsSOyJaU7qzoJNt2LxK40v5v6jq/6bprN9VQ6rW+2O0V35qgo11Op7lsd85yEd98Cpl+PhrjHes1gmSlamBUvauY+UPSbAGNt9tQ/Ur5h56IiR3zIK4aOr5oiR8YZSc38CtabX8kbzt+pOxXBRq74JMtKEq84Qsxg3Sey20brRdbqkWyrgpqcAdTL3urnXbP7epphuWU5jVXQA8Wa057Su0zy7XIdMFZJPJNmQkqpajEja8FU8IRwJi2OqXj1o1Lrxy8X4ArsX4UXfujbe4IL6js8EoGghNNU1Pg+x2RsqvkD3ppRND2XEINITezAxr4Y93QpxqFt4aOqTiWPpth9e2X3sTwM7VFJTrzK3Vojbd95QA6PZTYmsoLRVVTdjV7XhfvYwuW0n9hE7qkx36StX9rv/a+EtvKpTeubi8O7MhdNP89X8zLU3wWp9H567eDyf+OPz14+Xy/Ffq+DP7gjjuwE1IOcgkC7zRvfvZPQvAAAA//8DAFBLAwQUAAYACAAAACEAtqgYShcBAADZAwAAGgAIAXhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArJNBS8QwEIXvgv8hzN2mXXUR2XQPirDgSdcfENJpG5omJRN1++8NFdsuLPXS43tD3nwwebv9qTXsCz1pZwVkSQoMrXKFtpWAj+PLzQMwCtIW0jiLAnok2OfXV7s3NDLER1TrjlhMsSSgDqF75JxUja2kxHVo46R0vpUhSl/xTqpGVsg3abrlfp4B+VkmOxQC/KG4BXbsu7j5/2xXllrhs1OfLdpwYQUPkQtjoPQVBgGD/DWzJIICv8ywWZMBTwG9leZV22ZCmbvE52qRLFuT7Nv5hmrEMGGNFvFhsgizXRNGSaOeaqntBDNaS7e6XxOCaumxeA8+1oEmkDN7CeZuVZjQm9i+8ffSoP/W87NC5j8AAAD//wMAUEsDBBQABgAIAAAAIQCcbdIz+Q0AAPZVAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1spJxvc9vGEcbfd6bfgeWbtDO2iP8kNZIyJm3WntpJpkrSdjp9QVOQxTEpMCRkO+n0u3cPEIl9du8OVs6TiBIXe1g83Lv7cXG4i2+/bDeDT+X+sK7uL4fxWTQclPer6mZ9/+Fy+NOPi+eT4eBQL+9vlpvqvrwc/loeht9e/fEPF5+r/cfDXVnWA2rh/nA5vKvr3flodFjdldvl4azalfdkua3222VNf+4/jA67fbm8aZy2m1ESRcVou1zfD9sWzvdf00Z1e7telS+r1cO2vK/bRvblZllT/Ie79e5wbG27+prmtsv9x4fd81W13VET79ebdf1r0+hwsF2dv/lwX+2X7zd03V/ibLkafNnTfwn9nx5P07yvzrRdr/bVobqtz6jlURuzvvzpaDpark4t6ev/qmbibLQvP63NB9g1lfy+kOL81FbSNZb+zsaKU2NGrv35w/rmcvjf6PHfc3qNzY+o+3G0/W94ddHkyQ/7we16U5f7d9UNfQrxcHR1cbOmj95c7mBf3l4OX8Tnr8eZMTQeP6/Lzwf2+6Bevr8uN+WqLuns8XDwW1Vtr1dL86FOKNtPf35nMnXTvmmS+31VfTSNvSG3yMTTNGJOu1zV60/lvNzQ0QvS6fBLG0hy/s/Fy7yJcXSK5eqi+/0Y16LpFHRtN+Xt8mFTz6vNP9Y39d3lcDo8vvf36vPrcv3hrqagc3N1q2pD/vRzsF2brkoZuvzSvH5ufePJWZbk40mc5MPB6uFQV9vHVpuQTp4UceNJr4+e0+hskudZMRmT5/vyUC/W5rTeVugjbVqh18dWEt6K5/zZoye9HiNPz+IsKnriptiaM9Lr0S85y8dR2nfBxaMjvXYn9Ek1arVuPsOXy3p5dbGvPg+of5Ikh93SjHbxObVmPpw0O0tPep8+McqXlXF4YTwaPzr0QO9+uoouRp8oKVaPR8z0ETEeMddHJHjES31Eike80kdkeMRCH5GfjhjR9Z9EMBnfiXC80KS5OBrUTbeO5EWSz0mAQlwet43FhXHbRFySieNyaM57e/Xz2++//9tPP/z5RfLsm3/H/3nx9u3gzY+v3g3efDe4/tc1/fbNn16cL54VzxYv3l6/+svF6Lb5KM6mRT7l/8T5Te+2xA1yUAeAnFjf3JRN/zwqkzJlpuNO1CZBZuR+OsNUKMNtUhluk8qYkJQy6ZOUEclhjxJ0oO5sSYusu/g4SkWgM/I5XXwscmbOjfLquU1evQlEXX32lKuPz8bJ9DgXNQkt5HDEDXrQoODPi5xJMx3LHkPunTRyROBGKQ23SWlMTEqaPCQxHGGCEjRS+pUomBJFKpUg904JMfLNuVEqwW1SCROTUqIIUcIRJigx7lNiDN1FKkHunRJihJ9zo1SC26QSJialxDhECUeYoIT5WsFnUjVqTrgSsRw1yb1TQsxkc26USnCbVMLEpJSYhCjhCBOUIOTzKzGFnBBzxIzcOyWETHNulEqAo5iTX5mglBTTECkccYIUMfG4ZQ6ht81s3bBFHGUKoMirk0Bc5ty0ebJKDcAYSxGacJQKcfSkqURSliNY1EFypuodcQySiPlhRmYmicjzOViVJOCqJGkpUUBXHIdkhitYlASpc3CnJWEESlkiJopZzFEulqQFViUJuCpJrBwaPw1EZZY4gkVJeskzZuhJkogZY0bmLksSiV9gVZJwV91xrAAaBxGoK1iUBCHUliUApJmYOmYxJ7tEfRHzISm4akmsVBo/CUvFJ7RwBYuS9HIolRfM8Epf2sxXt6kaSzjiJZK/yNkzvAIdqo5jpdE4CEchHBYsStILpHFLpK0kk0REPiMz6zgSxMCqOg5gopLEiqVxEJe6gkVJesk0BjTNJJCRmUkiiQysShLgRSWJlU/jIEB1BYuS9CJqDIyaqSzh9JdINCNnT8cBcFSSWEE1DiJVCIcFi5Igq55qXEComfh8ZzFHP9mX5mBVqeGFVOOqKTV+EqYmovcuXNGCEAmCqmWeoSO6QXWSyHnGNHCi0kQyK1ilJmBU80wTma6KPYlZ5TzjChYl6WXWhDFrFBdyBCEzk0QyK1iVJF5mNa6WQmEQs7qCRUl6mTWBqulEMiuZmSSSWcGqJPEyq3G1SBLErK5gUZJeZk2QWWXFkMydJLJUNAerksTLrMbVIkkQs7qCRUl6mTVhzBrFuawAkJlJIpkVrEoSKGTKeca4WiQJYlZXsChJL7PSraJTSSCKx2p45eCZSmY195mcJQEw6uHVyqxJELPCGVmwKEkvsyasikqSqCzh4JlKZiVnjyReZjWuliwJYlYIhwWLkvQya8KYNYonEtDIzDqOZFawqo7jZVbjapEkiFldwaIkvcyaMGalsURNwhw8U8ms5OzJEi+zGleLJEHMCuGwYFGS3vpqwvCVskTduuQMmqqbl74SK7XMynBqeLXiaxJUZYUzsmBBkrSXXukIXnGVY4lp4DSAppJewSo7DhjV8NpEpu9cBtGrK1iUpJdeU6RXOeOQmUki6RWsShIvvRpXy83cIHp1BYuS9NJrCvQ6lqhGZiaJurftu+0PrjpLrPSaBtGrK1iUpP9eP7/ZH4/ljJNyBJU3MuZgVVnipVfjasmSsHv+jmBRkl56TZFeVZZwBM0kvZKze8YBo84SK72mQfQKZ2TBoiS99EoLhhi95ipLOJ/KGxlzs9rISa9g1JJY6TUNolc4IwsWJeml17Sl17RddTFVWcIRVN7ImJOzRxIvvRpXS8cJolcIhwWLkvTSawoV10hNwhxB5Y2MOTl7JPHSq3G1SBJErxAOCxYl6aXXtKXX3GRJSgv/xFI6MncXLUvUc7Cq4RVcJaoZV4skQfTqChYl6aXXtKXXoySyhERmJolCNR+9oquSxEqvaRC9uoIFSbJeeqUjzPB6lESimmngNIBmEtXAKrMEXaUkTWR65VkQvbqCRUl66TWD9QKxLNGTmUkiUQ2sShJwVZJY6TULoldXsChJL71mnF6nU5UlnE9ztTbRR6/Usudrn7HqsSQLolc4IwsWJeml14zTay4XZM7I3F1XLlENrCpLgCVVlljpNQuiV1ewKEkvvWacXhNVoSczk0QWGsGqJPHWXo2rJUuC6NUVLErSS68Z0OtETsJkZpLIQiNYlSQAvipLrPSaBdGrK1iUpJdeM6i9TtTwyhE0l4VGcnajGhgV0BurJUuC6BXOyIJFSXrpNYPaayHplcwsS2ShEawqS7z0alwtkgTRqytYlKSXXjNYL6A7DkdQ+aVwTs6eLPHWXo2rRZIgeoVwWLAoCadXc/HtMzEZXy8Qax04lOYSWcnZo4O34GpcLTo8CVnFaLeAcFiwoEPOkbXTgd5mVVZVizdeJ07NJaeCVXYRMKpRowlHPwbwtHWt4pQLV7QoBAfVNitp6m+fFWsfijPPSh3zhB6lO+mTFOrJGTJ3+sgsmoNV6eMtuRpXnSf5k6A1EQPeAqMVpPDX9qnBpnu8fnyCkD9ylXOaJd306lc6oispTaeypGQa6HJJAj5YlVZemm0i07kURLOuYDGVOM3aJcEHr+QMlHMkLSTgg1VJ4q3FGldL+gTRrCtYlITTrF0SWP0aSU7JOZIW6qEjXy0WXPWIY6XZPIhmXcGiJJxm7ZLwx7Bo/asoKeUcSQsJ+GBVWeJd/WpcLVkS9jCWI1iUhNOsXRL+PFYUyXvEOefVQgI+WJUk3lqscbVIEkSzrmBREk6zdknwwSw1vHIkLSTg575aLBh1x7HSbB5Es3BGFixKwmnWLgnQbCQLjzlHUrnYbw5WlSVemjWuliwJollXsCgJ0qxtEoaFsKp+knM6LeRKArAqSbxga1wtkgTVYl3BgiQFgq1FEjqCMa58yGRmGjhxSSFZH6xSEjCqjtNEph/wDKrFuoJFSQTialQroBYrl0/MyMwkkdgPViWJF2uNq86SIqgW6woWJeml1wKe3ZLsPCMzk0TSK1iVJF56Na4WSYLo1RUsStJLrwWsg1WPoJCZfQ2W9IpWsacCGHXHsdJrEUSvrmBRkl56LXgtdjpVz4kDn0p6JWd33QCMWhIrvRZB9IrhdMGiJL212ALoVT7WOCMz6zhyxgGr6jheejWulo4TRK+uYFGS3lpswZ/dilUtlswsD9R2Ar5aLLjqLLHSaxG2p4AjWJSktxZbtPT6+ITfRC6uIDOTRAI9WuVY4q3FGldLlgTRqytYlITTK2lomYRbej1KIosDBVRbJdCjVUripVfjapEkiF5dwaIkvSsJipZejw89quEVCq+yYk/OnuHVS6/G1SJJEL1iOI6NjMa9KwnoCP4UlywOmAa6pVhyeEWryBIwqrGkiUxvyhFEr65gIUvGvSsJ6Iiu0Bip+zqmgU4SCfRolZJ46bWJTEsSRK+uYFGS3pUEY6RXWS8hM5NEAj1apSReejWuuuOMg+jVFSxK0ruSYIz0KhdXkJlJIoEerVISb+3VuFokCaJXV7AoSe9KAtqjsOs4iVw+MSMzu5shgR6sEtXAqMcSK72Og+jVFWwrSbvhYbtx3vKhrhbNDo6n3RoXxvvJ20G220DSTokP2/sB7c9ntmU0irbv086I7S+DT2YPx4g2divMpon8zWYXRHjDbPOljhrLd+gpPtFS6zU6nfrxtzY42tuvu+iri93yQ/luuf+wvj8MNuUtjY3RGWHRvt3asfm9rnbNu8TU76uatm08/nVH26WWtItgdEZpfVtVtA9m+wfFY9q9LuuH3WC33JX76/VvtKElzaTVfk0bozb7oV4Od9W+3i/XtbmA0WnH1qv/AwAA//8DAFBLAwQUAAYACAAAACEApI+SbIAGAACuGwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54bWzsWU9vE0cUv1fqdxjtHWwndogjHBQ7NmkhECWGiuN4Pd4dPLuzmhkn+IbgiFSpKq24VKp66aFqiwRSK5V+mYZSUSrxFfpmZm3vxGOSlEj9RxQl9uxv3v/35r3Zi5fuJAztEyEpTxtB5Xw5QCQNeZ+mUSO40e2cWw2QVDjtY8ZT0gjGRAaX1t9/7yJeUzFJCIL9qVzDjSBWKlsrlWQIy1ie5xlJ4dmAiwQr+CqiUl/gA6CbsNJSubxSSjBNA5TiBMheHwxoSNDzH396+dWjX+4+gN9gfcKjzYBRqqReCJnY0xyIs9Fg+8OKRsixbDGB9jFrBMCuzw+65I4KEMNSwYNGUDY/QWn9Ygmv5ZuYWrC3sK9jfvJ9+Yb+cMnwFFFvyrTSqdYvbE7pGwBT87h2u91qV6b0DACHIWhqZSnSrHZWK80JzQLIfpyn3SrXylUXX6C/PCdzvdls1uq5LJaoAdmP1Tn8anmlurHk4A3I4mtz+Gpzo9VacfAGZPErc/jOhfpK1cUbUMxoOpxDa4d2Ojn1KWTA2ZYXvgrw1XIOn6EgGqbRpVkMeKoWxVqCb3PRAYAGMqxoitQ4IwMcQjC3cNITFGsGeI3gwhO7FMq5Jc0LyVDQTDWCDzMMiTGj9/rZt6+fPUGvnz0+vPf08N4Ph/fvH9773tJyNm7hNCpufPX1J398cRf9/uTLVw8/8+NlEf/rdw+e//ypHwgZNJPoxeePf3v6+MWjj19+89AD3xC4V4R3aUIkukYO0C5PQDdjGFdy0hOn29GNMXV24Bhoe0i3VewAr40x8+GaxDXeTQHFwwe8PLrtyLoXi5GiHs5X4sQBbnPOmlx4DXBF8ypYuDtKIz9zMSridjHe9/Fu4dRxbXuUQdWcBKVj+1ZMHDF3GE4VjkhKFNLP+JAQj3a3KHXsuk1DwSUfKHSLoiamXpN0ac8JpNmmLZqAX8Y+ncHVjm22b6ImZz6tN8m+i4SEwMwjfJcwx4yX8UjhxEeyixNWNPhVrGKfkHtjERZxbanA0xFhHLX7RErfnusC9C04/QqGeuV1+zYbJy5SKDr00byKOS8iN/mwFeMk88pM07iI/UAOIUQx2uHKB9/mbobo7+AHnC50901KHHcfXwhu0MgRaRYg+slIeHx5mXA3H8dsgImpMlDSnUqd0PRNZZtRqNvvyvbkHNuAQ8yXPFtHivUi3L+wRG/iUbpDICvmj6h3FfpdhQ7+8xV6US6ffV2elWKo0rNe23TeycLGe0AZ21NjRq5K03tLOID6HVg0Q4GZDKeDWBbDx7zNd3CRwGYPElx9RFW8F+MM+vaKGSMjmZOOJMq4hHnRLJu5lhyhbUZUCq27mTZreg6xlUNitc37dnm5OG9OyZjpMzIz7YTRsiZwUmbLF96OWcVKtdBsrmoVI5opio5qU5XBh/OqweLUmtDZIOiHwMorMP1r2WHewYz0td3tLD5xi2Z9pi6SMe6T3Eda73kfVYyTJrEyCSOPj/TseIyPCtzqmuxbcDuJk4rsqgvYTbz3Nl6aDMwzL+m8PZKOLC0mJ0vRQSOo15ZqAQpx1ggGMCrDxyQDr0vdTGIWwbVTqIQN+2OT2YTrzJt1f1hW4PbD2n1OYacOZEKqTSxjGxrmUR4CLDWDvZF/qQZmPSsFbKT/BSmWVyEY/jYpwI6ua8lgQEJVdHZhxdx7GEBeSvlIEbEX9w9Qj43ELgb361AFffpUwo2HqQj6C1zPaWubR25xzpOueClmcHYdsyzGebnVKTrJZAs3eTyVwXyz0hrxQDev7Ea506tiUv6MVCmG8f9MFX2ewBXEcl97IIRLYoGRztdGwIWKOVShLKZhR8DFmakdEC1wxQuPIajgqtr8F2Rf/7c5Z2mYtIZJUu3SCAkK55GKBSE7UJZM9B1DrJKfXZYkywmZiCqIKzMrdo/sE9bVNXBFn+0BiiHUTTXJy4DBHY0/93ueQb1INzn/1M7HJvNp2wPdHdgWy+4/YS9SLRT9wlFQ9559pqealoM3HOynPGptxZrTeKl24qM2g4skpP/A+UdFyIgJY32gdvku1FYE7y9se4Ugqs/ZxgPpAmnLYw8aJ7tog0mTsg1L3t2eeRsFt9x5pzvlC1n6VzrdUxp72py57JxcfHP3eTpj5xZ2bF3sdD2mhqQ9mqK6PZoMMsYx5oVZ8WUW790GR2/Ca4MRU9K+LrgD14YwZdgXD5D81rlm6/qfAAAA//8DAFBLAwQUAAYACAAAACEAsI1PNu4DAACUDAAADQAAAHhsL3N0eWxlcy54bWy0V92OozYUvq/Ud7C4ZwwE0iQKrDaTQVppW1WaqdRbB0xi1djImCnZap+gl32PvkHfpn2PHhsIjHYyyUy3uQDbHH/n5zvn2Fm/a0uOHqmqmRSx4994DqIikzkT+9j56SF1Fw6qNRE54VLQ2DnS2nmXfPvNutZHTu8PlGoEEKKOnYPW1QrjOjvQktQ3sqICvhRSlUTDVO1xXSlK8tpsKjkOPG+OS8KE0yGsyuwakJKoX5rKzWRZEc12jDN9tFgOKrPVh72Qiuw4mNr6IclQ689VgFo1KLGrX+gpWaZkLQt9A7hYFgXL6JfmLvESk2xEAuS3IfkR9oInvrfqjUghVvSRGfqcZF1IoWuUyUbo2An7hWRdf0KPhAO9voOTdSa5VEgDSxAkuyJISTuJW8LZTjErdiCqBna7nbPQrFlue9GSQaTNIjZaO93Pq2Iipy3NY2dhxEdl//z5+99//XFO10XYix4UpGT82DkQvNJ876mpk7i8CtU6UUNwGOcnYmaGGFhI1pDBmiqRwgT144djBbQIKLYutlbugvRekaMfRNdvqCVnubFifztNBuv0rl87sTa31OOJrYZza5d9gXs7qXLoIUPmBQDdLSVrTgsNrCu2P5i3lhU8d1JrqLNknTOyl4Jwk0bDjulO6D3QZmJHH6BNDKlLGi37zMUGvke/KGttsCZcFAUzBysvynbOPO9L7xSEKKOc3xtnfi5OcYrApbZAoinTUn+AAoHua4ppGEKM+2EXk26SrKFI96KkAoqTKs0yU9sZTGlXj21xHtYHDf8DbHAWFpg95+LX34SnUe5iPgm3aYivDzdqi68Qd8Ps83E/4SNSVfz4fuDWZPiLVKNfFakeaAut3nZx/BLvs1E/hH1MJ8iHzj8YWP0p5J8B7Gdg8zjb2DIf52dtPUjFPgHQKxLzbQEyfcNwDixPSuxJgZ1SApmTJ3Z+MBcSPrgNancN45qJjmRzPg6V+mQDChAErosVBLNVq4ZBxf7m9T8X3pF5eONjNus+frZn3DnY2QhsNFwFDHcn87sIPEKH10KHl6ENli0Kkz1X2etPQMciBc7ydmyH9vjR5vJmG+WJReAopwVpuH44fYydcfw9zVlTQux6qR/Zo9QWInbG8UdzAvlzQwXUzMcajg14o0YxYPFu891ye5cG7sLbLNxwRiN3GW22bhTebrbbdOkF3u3nyRXyP1wg7Y0XGpEfrmoO10zVO9sbfz+uxc5k0plvj3gwe2r7Mph77yPfc9OZ57vhnCzcxXwWuWnkB9t5uLmL0mhie/TGi6aHfb+7shrjo5VmJeVMDFwNDE1XgSSYvuAEHpjA49+J5F8AAAD//wMAUEsDBBQABgAIAAAAIQDmq8IA9wQAABsSAAAUAAAAeGwvc2hhcmVkU3RyaW5ncy54bWysWE1vG0UYvlvKf3i1B+RKiddeO/4IiSvXH0loTJbYFuI42Z3ag3dntzOzLTnCBSFUcmolLogrF1QuXADxZ0gC/4J31naa7Nppsu4hkuN3nuf9mvdjvPv4K9+DF1RIFvA9o1QoGkC5E7iMj/eM0bC3VTdAKsJd4gWc7hlnVBqPmxu5XSkVIJbLPWOiVLhjmtKZUJ/IQhBSjpJngfCJwn/F2JShoMSVE0qV75lWsVg1fcK4AU4QcbVnWA1UE3H2PKLt2Tf1stHclay5q5pX3/1y+ebHqz/fXP389UYOOalgxEOoS3dN1dw19bGbRy/Pz//9++1GrkOlI1io0LXkwYsfvrn49Xfz6rc/Ln76fiPXJzx6RhwVCSpMdRamiC9evf7nr1cbuRFnKsl1+frtf9+eb+Q+iwhXTJ2l5HNzTjA6YiqT4t4xuiIoaLUw6MN+obptFTpbrc5gAFYlFqZBB9BvDbsnh62jJN/T/oM0lKsrNNxlVqWeAVSzMoAa7zFvedhKlVVxuzPY9RVuLdwFh4TEwRRDGHgUfH1VqQu9zjAZcrud/Gbh/f05Fq7fH3Ht9jvIWKCR7l2mWguvHwLars3z8hDQIPSYwuqF0k4lGZ4bsjrk+4EbeUQ8Sp7qM0cEcBdP6kQ9lZqbqQsi5QaBmN2ifKmKiRVKxv+mlNvUpdgKPUjENC7c92BH3KVijrvmGdgHDasIAzbmRLceGGBnoxKGuhMw7niRS4E4DpUyiAUfET/8GIuIO4IqCqdEphpVj51igI+x6TnwScC4grYXSE0esx7yLY9xqvtKG5tOqlPcD409Yw00VtMaaKyjNdBYU2ugdX1lgs9ijoWWAd0J/Kz5iqHZkhVDs2UqhmZLUwzNlqMYmjFBsxCvSI/NxoowT4/mQdtspTt7O+CcOgq7yHK5TfmYuIRwcPHP1jsSwW84lGAQnbqRo2C7aFYsCPGUnixBLMXGjTXeQXGqCya/wIVrPQ2bgNpZSFK9kk6p+FKbblNxStgUP51c2wUmHHTs7iY6osGFBFqvhzsSByaujbj/SSpeUKOJLsdOzcFQtkyrDi+J4AQmTBF/E8q4ceLehgvhOILUoAicKVWx4hk2dSChoFI0y5W5ghEyav5BzG9Hik3u4o+hyQOfo6m4HYMi2KXzw8hjmE542nrSPYL+F3DStUdPjg7bS8bHIklDhvmHGvjYxAVsG5swGCL4aKtYLEOp0aim07CAwgw7xNCMcRI1snDcVl/5AOofxHFbfVkPulQShtTHMaajlA+pmEYEXyCg4rBhyLCAoilIbKrjTSAIx9LTpUWmks4mZSr4M6UymGJtQX7G1PDvdayWPjbQDyJ9B/RsxQXwhDhTTJwBL5mawL5HpASbcOrdryIWbMdYx++40vfO86CvV85bamNbiHCX6oeSBSPIV7aL4PvQoaGapHyOdwW8Sx2mq+4UayLg0BPEp/lZP3sEW1CuV+arUf6408vGoXfG2XqVnaOG+9K6HPXq+r7gZFzXjFJ1fQ69u99hB2Y/flouni3JKzV/XV7v8En5/N2zUj5/06yUz18wq+TLB4RNlDNBs93FY7jWsuaTFS8iXsmR3cYPJbzSyeGwFmH5g/HhjvBhDZwRLjGwaevHjDmfCiOXCIJzpR/PFauwnZojx0ddffqw9ek+jDqtk5Y+3cXfEeDWaRN/3Gn+DwAA//8DAFBLAwQUAAYACAAAACEAH0pW2zo7AAApigEAIgAAAHhsL2V4dGVybmFsTGlua3MvZXh0ZXJuYWxMaW5rMS54bWzMfVuPHDmW3nsD/R8IPSzUWEmVEZmVl95uLVQXtTSjatVUlaaxfovKCqlylJdCZpZa2jf7xbAX43naBQwYxvrRL4P1g/3iG+Df4rn4X/iQwcjkOcmPZESqG9uY6a6oIoPkF4fkx0PyfN/89cfZVH0ol6vJYv7tg+xJ54Eq5+PFzWT+7tsHb66ePx4+UKt1Mb8ppot5+e2DT+XqwV8//fKLb8qP63I5L6avJvP3il4yX3374Ha9vvv64GA1vi1nxerJ4q6c01/eLpazYk2Py3cHq7tlWdysbstyPZse5J1O/2BWTOYPqjd8PRunvGRWLN/f3z0eL2Z3xXpyPZlO1p/Mux6o2fjrl+/mi2VxPaXKfsx66uPHaZ7V76df7BQwm4yXi9Xi7foJvfBg8fbtZFzu1nN0MGI1Na9t9648OyDwpoTbKu9Q3Z5usDxaLCyWXy9TgKgqe7IY38/K+bqCc1lOCZTFfHU7uVs9UMuvJzffPli+vNHlmEp/XUzNl1uXb5bT1eaX16vF9N78bpsnf3Dw9JsDby7zCb8vZiW9YfOz+lBMv31w+ebo+PX36vy1zi3+dvT6V+r07EJdvVbPX3j+/uzVK/Xy6vRMvfxeXf7NJf3kSXR5fz2brBUZVbmcFFM1mVcmRq3GqVfl8gN92VhibZmZafZOC0+KdXFZrm2b9JMyaV4SwJ1Nwfz35lUmFf99Tl9jufhR0WfWH2ZcTqf652dkqetvH6zWS/rlh6d//v3/+vPv/9Mf/+vf/enf/Oc//sO//9P/+Ic//eO//PKLTcOpn5bfHHygb6RfsH3NEX/NH//Db//wb1u85pi/pqrEH3/3uz//73/68ouTcjVeTu60qfmqcMLz/uHf/as//P6/Hfzpv/z3P/zHv/vyi7Nifv+2GK/vl+XyYP3pztuKU/GK3/79//2fv/3yizfzydpX4nPR6L//p//3r3/35Re/ui/maxog3CwHBP0G/76Lv3748PSwQ/90h52Or5zjPvtG5+X8XTEr5sVc/UUxu/srda4H0+Ku8MPCMz9/8fj5C18hp6KQY2+Lq9pmsGkDt2n6YdO0kb9pA9a0q2UxX90tlutiNXmkzhZ6pF3Rz7ahN+Vs8yuvCfC34bbydOf+tlbVx20dum3VDx+e0vRC/2TgM9LE5nS156/VeLEslTZGdXmmvnvSP8yfnDx+dnJ5qfKe+ePK20rxnhfq7NnV6cXLZ6+835Wn/uWZ97tW1e88GfUPR+4/A/ilR27r9YP90nnn0Pulj0as9ZdneSfLMo1VJ3v88sRr9zzLyznRgqliuMXB4u8gk7CTxeXpxa9fHp96IeN5AGRVm5tAlnXYwEvExzEHMwJo8B4zi6vH6SPK7Cav8euH8BN52gEoXpKGoMgEIDSAfHjK7IWNlBmfqvRT3clGg0OvzVCW5F7W7Qd6mXxRpJuJ5KjFVRMCLdbz9HZy1k/bjgUsI4c9K0c9K+N5oGWEEeIvSbQMngnhVDU8gFOX4dT19qUM9SWe3O1LGDGepyVi/CWJiPFMCDENSLAv9Rhi+sn2pYyYh78v9dL7Um8Y6kviRbG+xJOjFldNyJ4M8pGZd+2/8IydHTII9NO2c+U+CI4oi2/Y1dNWF3YunsdnKmG0eP5EK+GZEGZVm5tgxthqxlliPWnlqKPx5G5Hw+jxPM3R2yG8CTO+aBdCr6K/gVGJ8d+sYpCGFGajgZ8AU6Lk+WqQh/qYJL9hWijKRS2umhBoMWPBWcUjq9VMLwMt5ozU94HDDZX8N4XSUc1cmFFzLQ+GrDdjtJeedihcl5rt7w3HInnzhrfhsqJQ1HDLZmHDc8ZdzVM9e/S7/u9MiZItexRiYvJFkdlDJActtk3Alm38ZBsmZp7qyWIIFrPa7edQet8HDjdUcNekxYooFDU3RjxzRjzpaXdxQs0Gli2SN294GyYpCkUNjzHJnDFJ87TlRcCyBY2zK3n/Uj7rhdbyVB7rJDHbTuKCthEB22ZcMK+IlJ3Mh14ueESJEBHqISIk8vjMIoKO5I0pQ70oFdlF1egARows5pxebbACxEckd4kPRitOGyNoteGNoqYIrYo3BtBiNDGviFLFfDqZf6VOidL9YcPQ8kK+KdaJeMGoyTGylzOyZ542q4nhyLuaoESoEx3CTsTz2E7E3WB5GJ4dt2gCIxZVRSDF+GHO+CE9eXxewxHqRTy524swXF522QiuNvxStAzBFeOXOeOX5slORnm/71+kUyIX04vF/fxGHevtSJUdq8fqsKNmXjeyyJi2zBSZVqV/Z8TWHI8XXcYnzZPtPL1B7l2KH1OiXXY1VekNFi9Ia7DIhBpsWxBoMHNldh1XJn1Z72hxTImCX5acEODTipyJLeXFwZbGmGSXMUnztP20Xa+jiRIlfdpAi9swSFEsbHGMQnYZhTRPm14Llv2UKPxtYbcVORO/LS8OtjTmROwy4miett+25/+2nMNV85en22a4xW1YINXNBRi2OEYDu4wGmqftt/U65I8pUfDbEi1C/bYNgxPFwZbGKFyXUTjztP22fgpHiZL6baDFbbx1oljY4hiD6zIGZ57qbzvIwEjFudTO5JNjK27DwqhOSVYco2FdRsPM0+bbDtgmb+3xOKZEad8Wt7gNkRLFwm8bY1JdxqTM0/bbgpEqwqS6uN+28cxRnZK+bcw112NUyjxtv62fM1KipG+LWyzekDYLiUzo29omYC7VY1zKPG1dNn4uRYncFtf7L2pM52bGdFhH3S2IMM+IPK/LG/X85MpLmcVbnkcWmiI5OGdi6x9oLiNUPb4n7B2mjigRWmj20UJT5LELzQ1SjxMhakPERNEIqBgP6zEeRk+exeZhFyw2RXJ3sYkh824KN4asza6wqC6CjL86YGOM2PXY7nDP7wWlRC6+9Z5N0y7VbGtYFIqabbeG4XZGj7E687TdCfbODkeUCHWpAexSXpdeW6TakENRa4SX3RbGeDFu2PPvAvdQz8K7wBg57y5wW+TakEzRSIQcf3WggzGu2XO3hjs9/zqCErkWV28dNe1gzfaFRaGo2THC2WOE0zxtO5h3MXFEiVAHG8IO5vX2tUWqDV8VtUZ4xehqj9FVevJNXYeog/nPVuqzgRg57+HKtsi14b+ikQg5/mrcwQ4ZDTZPG1LY83sUKZEL82Z3ZdvD3i2NOzXGC8WLYrxQJActt00ItJjR4MPKL2d3pvr+40yUCPWxEepjIo/lhR6wkii0eFvawkFkQoDFPJOHjEjTk6eT9dFZJpHc5YcYOu8xy9bQtaHWotoIOv7qgM0xin1Y+QvtPl/P7zKhRC7Om72q5r2s2V65KBe1PObyPGTM2DxtZrKBf/VFiUAvo7NB4Ky7yGN7mQestF7WxmUqqoAAi3lMDxm3pidPLxugVZhI7vSyAHRelt0aujY0W1QbQcdfHehljG4fst30nt95RYkYzgN7oqp5L5NkOXyOUJSLWh5zxR4yemyetpvp/gUZJUK9DN4oEXlsLzvcBSutl7Vx6YoqIMBiBPuQEWx68vSyIVqRieRuL8PQeal2a+jacG1RbQQdf3WglzHOfVj5WfXJbrqSNPIvySgRs7k7uglMd0JV9rXXRk9E+igx5K9HDYx5hPuMCpunD0+rhg3zvneXjhKBhg3Vw7PFzf20WH7l9YmKnLEmiuSgibbO+Nv1GfelJ0bk628I1ksiuXMXLYP3P2Se2JcX6dN4rsiEwInx3D7jufS0Cw7ZAQIHuo4zeGZfFHG5BSdqPW04rSgOwRRzF/cZl6Unvw35V1IiuTuEYhsS946iNtTGMSwqhsBJdQz3Gf2lJ78tAZD8RJgcEwFbEnka2FIb0itahODirw6MS4z89iu/ar0+8m+iUyIX0zMdHUI5Xcg/s4hc0WGXF4KaGTsj0Od31zkxBTML55M7zfPuUJ5QQS4o0ebx5Kh5MRbaZyyUntjCtbrb3UNeOJHcnVXgmV+RJ/Xbt+GdoigEUYx39hnvpCcfRHBugS7eDB7oFEWk2k8bfimKQhDFfLl9xivpyW9FYMiEvtwMWxHPk2pFbZy2ojUIolSn7YAxVXryW5MfKpHcnYKhNYk8idYkcqWROZEJQCVS4ZllwBivedr6t/1HeCiRi+e5e8Rhcb++WSyW1Z2Vh1mfdveX65V59PN88bLYiCzLBjE27DlY5kdjIQAGjMuap9oxkPnd+keUCDkG4E6+yGMdA8mIKZ3SuzwSL040Hd4CZDoVwc0D4DGGO6hcnfVC0O9HokTMZsqb0oTdEDshJmpJktk0c9rK4oHZxJy2A8ZazdPGnTTydpYjSuSzGs1VsdX4zraq7/imUfOe1obHiuojk4k5bweMv5qnTVibvp+/UiIXuDfzm3JpreW8tp7L8xejvKMuJ+/mhY7BpC4pkFa5Ulc6+M1kPp7e35SqGI/L1Wph/lBFNhov5uNluS7VdbEC3Ut6a8PeSlFXhFKMBg8YDTZPm0HJf3D0iBIB8+pCl5vIYwelnxngNmcnRMURzDE6PmB03DxtjHHo39KlRGzdMLkm79xrCho2Vr9YTOZrdTxdrLQBGst7OX9MUfFKHW7pmGIx+QfwZucnRAVQ02M0e8Botnnazvb+Q52UqEXTKcILbrok0ZHOxZOjpsfo84DRZ/O0+eqH/tOdlKhF0+mcHm66JMeRpic5bm1bMMcbMjpsnjZNH/jna0rUoul0kAo2Xb4xFl+MVwB8dduWQNMZvR26oZ6ygf+rU6IWTacjLrjp8gJ+JLSa4Nd+smLbEmg6o7hDy+oqH/bQ71ShRC2arg8e4LZLT2yk7UkU1TYm0HbGUIfOeQK6NuznG5SoUdurEZ5uyOKmN+OoogLI4mMcdcg4qnnazm7+I7eUqEHTTxaz8NQmXxfr6bx01O4YxRwyimmetlMb6OmC5YVnddPu0LxGZTIYY+1O8p3ahgRMnZHGoeM7pRUGGNwF+Upod2hSozIbtTvJqWobEmg3Y3FD5+RrJxv4qQwlamrnwRmtGYUTpSM7j1G4IaNw5mk7mYNhvRGFM3YenM6a8TeqIlv6g+ksxt+GjL+Zpy1/A9+7EX8z7Q7PZc3YG9UxpeGxbfcRY2/madtw/wenRI0H9NBMJt8XGdlEcmDptiW4h48Yd6Mnt002Ju4QnasUyd1tE3g7QOR5vv8qT74xLU5sEv8z6IQiM44Y/6OnHfiyTm8ItutFchc+ePZb5EmELzSjyjemwZdEIQ06QfgYhRyJ06YVjT5EIUZEchc+eKpX5EmELzQxyzemwSc8tf7hml4djgs6YjSUnjyddwAOkh2L5IlQhOZq+cY0KJKYqWlp0JIYMx1x2mfHsQG0JHg/jI7aoADXjXht7a0KTfmi0onxrpMILr06YkmM4I44e6ynAbQ1LJI7HTHH00AjelzDF2QOohqJ+CURZXp1BD9GlEee0wd6TYzmAXgaNsfzQENnaXwlLSqdCB+vBmIhMb49YnybnnwsBIVXF8ld84NH9kWeyOAXXY3L16WNfEmc3UATHPl4OH//kQU48vmPLNA+Y45D+zdn/EH+0eYswyiJ9lOqSGDtjojr7znFoFf4MLC/P7K/hg9HI6cymy4cQgSEChOugaSguiIX6Ln65TEI2QqCZCG8LARw4GOZPqUnhjjIzvuSuqLMBdGIHd/NOmxBoB99nAzNpDK9O5bhqNuikBQIQzxEViJVNiJpSVABFAx232GLAopw74PwEC2qZHoXQngKS2ZKgTDIRXZemGiGSesC/fJYp2QrA1qDejHEZghvweXweJYsJGlODTlEdl6YiGHSgkK/PIYhW1JkHe+agnwjQH1ApHePuEFWLAv5DN6RnVcmopi0rpAvx36mrMNWGPrR7ypBaMIzGBnkyLKQz+As2XllIppJqwz58hCabL2RdfwLjhG0Tf+KQx+Ygq4TWchn8J3svDIRzaRFh3x5CE22/Mg63vUHOVL851Zk+s/gSdl5ZSIwSSsK+fIQMFwzoeNZXNBGEDYzeCAa+1UyUUgimmFC02aJISsCOWHqgWmSTGMaNkJoa+tlAUMglhHDbhZd5uc+bbDzzjTjFDVBcIpkAeOUCmPehQc5XRCc/ngaNAZirwvJ3jWDM+522XllIppJ+xfy5SE0+cJFqIzVxtmBaPoPbms0oRMmE4Xs64XZeV8ilGkLF1HZEJR8AZN5FjBEP+GoKdI7xBH7ZDKRKQXKkFNm532JUKatX0RlQ1DydUzmXceQfwb1cXgwPOCfEYWkQBn0z8j3JUKZtowRLw9ByZczQh2t3iIZIpIj0u/tp5HvS0QlbVkiXh5ChS9LhJhYjUofGhhcluTYXyMK2dtfI9+XCGXamkS8PAQlX5MIlTILJfltUF+Fa5Ic+22kBFvCKauw30a+MBHLtBWJeHkIS74iERJoNZbYLP2XNfVsjP03opD9/TfyhYlYpi1ixMtDWPJFjNGB+/CU4iPrf0ZAPlNor51P3q2LyVSLPV8eHzzzk9eTTOSKXbST6SEnju5jcI23jIm8USP9EqFCbe14MZ+X4zVdLAw1saGim65KwgGtusaBz8hF3TKhnbb5nGirRqR33cLY5SEyJVqByJUqrpvG6K2aXQgozuSF1poNaNBBO4KZSO8A1cVMXmRKNKU2AURk/VCfMaJ4wX0Grh6XCTm3rUGB+Uqkd2l6wKAEQU4bVnaU5pJ2+EQuCFRqOJHMiM05AtceMk77hXBdA2ToaFIKGRYvJNGw2tzC1M1LGql4slBH5Iyb5No2osWdQ7/C6lEG5Od0bFC4I7+TiQKvjCfzd95LcTuJ3xwdv/5eXZ5e/Prl8akvy6nMAk2pamGPHZjlqvBGXs6xIfcoPWlBe6XXMiESV3e1XhgTzm9fFdfl1IprPVzMTUT8R3R/Vb0o5je3dAXcf3V+p+wKrbNnV6cXL5+9Mnddw9Cl0Wwru8fYiECO02srWldRwRxEs8yAap1GLiAwz9lshRzSVdgpojFAadzZtjegEM716jLzWB/ipoiSwLT8TFkDhBdwQjHOAvTsSj2s70P/6oKuLN1AmxKFNoYsjSLXCOAgAxkXrase7XXobABiNOhU7kDpdke8UhOZKsxg6IWdMhpDlHRAqG5wj8UV4P2Oy91lQlTOxGzWozninSK9BcsEcsDjucwUHM9l4oTxXGRB47lIFpjqjCjedlgXInZ2kEIaxEeZSF9b1Cg4wItMbQd4WXZTSxP5IZac3oew5PxdKN1tBnxocH5PvMYSD/mikMiQL1JbSdL0OVHkh5DxloQg4x53obNnIRvCI0MivWt+eBIQmRpPArLQxnaX5msXxYRA5L52qxBYgTfsgFWQ0OpzocNzgchk5wIbcMRPWmU5jdFKI/e20SGUOKl3VQXpbBtgGULnz50x8VkgqUU4+UDxMCjA4GpF1F5df1KrcvpWre7vdPgngJq42NAYNRHf0FfK88xiEAKNe9mtQKF1RWTeeL00Kfh965qa4SM/ItPFDmjFel2Mb9V6ocqPk9XaQLmc3LzzRxDZqURjBNPov0UkBCGn/1bx0EpGjFDv9PvUNYTYRyGEC3chHMyUNkESXTOhtm+L+bsojKIiKRwlbV1ggQghx33proJi3hn5ZQAyoWXo9FgczF5m2kVu1AI571ohuOQUdYfTa+w2cMaFGKvHbSwuf2h3nQqsDrIA4eWZvMjdLO61Mna6zYmKpNhc2qrBSlIGbI4LO2ausmOeZX4RJZ3Kg5xZKmDmJjKdl+/L5W+KYq6uluV8fKs7Jsnzjmfqprxb3/qnCVlwAlIiy5l/ZojKQWZcD7J63Mbj8t68PtKpEFCYr0lFRw9Q5IccdlSNlXp4QxPEcnJ9v56Qw2gyX00o0hnFO1vpsPGL+fQTcBnJ6qXAyVuE4IxePzAKk9t1GBOczDJ/9DyKxQjhhNcfZaYjiv1GpjYr5vfFNMHiRJEpEPEsCKJY5PCMK01Wj1uLAwTOrzBpuiYmcCIThyhXl/fX6uR+vE5ASwQfT0GLZ0FoRS8SGDlJx6CqU/MV7aDtS79TTSg9uj4PzNxEJo5Wtwla4tBPClo8C0IremWAa1NmTJyStjgBWt7bx8a2MEkT4pAcrfNfHx+8ODk/VdSs+ODvXSMECYYoHMEVu3SckU7j028c46o2AyqX2tAfOYwGf++yQMMVYGYik4XrZrKkDW8a3GnI+k0xpUlzWXwqwBwpik2xKp4FwRS7W5wZxUkHpuo2re2DOVg9CTFIpw8GaJjIBGFaL6YApeY8X5SJUIpdIc64bGX1WLv/uyM0Unnd/8aYApSLZ7IoLWl7ibhXuUYG1JzIC4FKBE2Ux3OFysw82imPDsB4h6VjncolWbaVZyUt/cC3b863RRmogdHjL1xAkoa87U5rt+8fd491Kl8DKwZDS2RNBB/+opjeL9Uvn70EXE+8xLpCg6OnyAIaHZWQzIxq5HZYYCKSQ76jUac60pnQ6gwbvNRwLOfvihu9xrih/9+Vs2JFLgD94+SuqCjNHf1d/aacXRdrnUxnmDN/laLgrnf063uTYEY/vbufF+8p1bl+yTWZmfqumH5QmXnrI/rNnF71TvXVY5XlakahhZd+O5RNbOqyEfnR94lycS5ZmZnHmmgOe4BoAqlKMkcsu6JfzU4u/JzfJ2/+fbzcP333QLQWfZ/oQsAoYDr9hzPmanqlfoT2WYRGpTvB4kWoFLb8Z96TvOuOBl8qaREiMAl4VbieZyZEL+svBrd5/NqeZrLHexVSWfOfed/yrn0afLGkhZDAJPTF+G6J0RfdjoGAlPmFRM13wqeXpYwn+k5mejHL7npW+kT+nJmiCWp9r51laj3R84yegbazzurgRvU6wVlnz60WUX80qkXXVEZx1BnVnDVVlw7a+ZegQgnUHcuwB0jKh7ZFPIeI0ycPQu5dkjUw9qT1mZVwDRk535pxFVAJcr8aZSa1RBl651tOlWnPx412Ex12Dno022pmRT7whaFUJFdBEgvanAEXEmGiIkElZa2QFUYXY1ziNDOPtt8TJOAetdAI1fRwSzNbQPJIVeQEQNMswqpug0uyEDTRxZgRK3U6KF81mUlMWw2kHfBOdTfQVcW+xk9kbK2uWQudVYRrdA3IlVMzIVZa44pCmcr09YafPp2MPbyikJ/YYnfkWpOOgYtcAN+oamvGZVurx22XBtuofrlWM5UHUBU3ITY7NTQG0LJxNaEd52oc3A5/6pOetN8vVjQuqgN1sbi50fsRZqI/tRv9T/wDgaxjgmdPZEGYRpeIXO01M49bTP2BxXUqvtqrd/zOy+V1MdFr6IvNtEBYaE/wI7uORhDImxhhAQFZBdT+6BKMy7hmQv+07rMw/JpfyNVYF95nkCKrjnW1hk+KEKR1zKTVkYEoeI+Fy7xStDrH+dXpgr1moZ3qMj686SAFVzfQXS+uF+8V9TjT6c5eaKt7ochDQ66h+/fkIqqRfa9dQtMJ6Ih7Ll7S9GBrgAKEzmi/bqdoKwVbmyPYlBaqru6mF15dikznm4n5HaFEHZn8Z/e0qbqavCXf8g1tUdAGq8oeHVa0+BGNgvPf6IFvNplPZvczRX945OJeM6hZsXyvx8O7xQowRVn9lFEw6XSYRS8EN98GMkK121EQbFoLlVkXbrxIFJm2hyUs3BW5JrO9f6/dm9WeULG6o/0hu1qcltfFkhaL5usMzVkBSrj9NrRIHPNPQJ5Kehd9Ku3tXN5rR+jD62L8/u1kStFB3n1Fn3Lr/NyophmRvl/m+aH1sdKH1NUogO9dopHy9ZJWP/ZjhL4eX/0YHd7t1wNuTr/grh65A+FCpAyuM3LPCcl3usNc3a/v79RZMddXeg7quz1gvPFuVqWvH9NkeTOLSAhCfkLN6PRudFX8pzaOM654qz7Opl+ThYzLbx/cLctVufxQPni6OWJQsQDVzQ/yofqxWM4LdTtZF2SpXaXPuNwp8n3cK4BTw5WSlP31nw2yrQzBwi9uG23erYAoOH4mVG5dHo83OHgmP5ZWwu8zQdpqqSTahmhXdKnEpXsz81hvlY5GgHZKyd7F+H25NuSyMiu/7ezI70aYZZrybl3lgPFwzd3Miu7aTfMeYEVC+dY1HuzTF5lqO0nER8qWJVHHNIXeutEhlPhdE6vRW535B1dUjzMp0lufY7KDTK9z0O3ZQeYNjSp6jLk0Y8w5naQDRxB3tHZjdpJ0Gsy2JwQAvzlidXZrMwFUTyjdumaCqZ7IBIaTduC1Wn+ICoHRxCISgpDfG7GKvfWlG+Drk5K9Tm8xEIDRRF7sjllJ2q0PW+VQG/k+hhXXrc0EUFShWOuaCaaoIpNnNAnhI9XR0kaTtKBKttEhlDiRt+K6lSWQI82/9wDUdfV9tQBKjLz6J2x9fKCYT/6WjoxPSKqYTvMW5tQAuULomu7tV4DtyPo0PT4g8qNuFT31xRV0M1dCl8LuAvYjJGxrk9NgYu8wz9QQzDwM5p4UW7QHgRndj+CavKRdZ3wk9b1UcHtGqPK6YGLnJc/UEMxuGMw9r42L9iAwozsYXOWX4r86YMJL40Ln1wUT++p4poZg9sJgek+rpS/+RHsQmFEuznWDM1c4mLo5GDOFzq8LJvbe8Ux+MOnksHr2Q7MBU1am6YAp8gMkoyrEmRHr3XrtXB1iQhIckxVCxA6SgdPEPFMTJMOjpaxMYySTbpFERY0zI/3rIOnIGhOSYLQUusYukvjeF8/UBMnwUCkr0xjJpOVFVCI5M6rIDpKOSDIhCTbNhEixiyR2YfBMTZAMj5OyMo2RTNrlsOrRAW7JJZczV3OZ7l0jm/THfyU6FDhNKaSVjwt961BvJP6ff1RXx/7wlSe6QuxgbWOckk5e2WaHcOIrFa7RjHaDhKqya3F4QSsySZy8xk047XlOSuRHs0X0nJTRPnZ6phOIitzQqGf6r6Rre8JrFSHdLHHyMnnCac/DTSI/wim6DDGizw5OjgY09TuEk//eucYJL0OEdrPEyeuyI5z2XGGI/Ain6AqDS0Znrma03tbwr32FbLPb7/AKQ2SSOHl9VoTTnosHkR/hFF08cInpjGlMj/zRWI90Kt+tBW1PePEgMjGcso76SJul45nfzyTLazyY8/oisKKLAy5Lnbm61CRZBryTQhnaNSq8OBCZdsEixBBassCmaIn8AK2olHXGtayrx802WgfsRAINazKtwE6kyERRKW/uJ+tq2dQ9qAiV37JkcY2xSqL4Ud3qjAtXV4/bLUeEFQwtFVAmkPrSU4q7ohlVBKc9L4mIUpFNVSsbNgXz6G9Gg3k7/VlJ5spxNELBQ4AytbYpTOBlpltyiVMQwcqy6NR6FLE9L2uI8lHUkKgodcZVqavHjWmh66pCO9oZtAKB9EWmYzoec6feUgBr7caoobMBkmL2tid9F1WB6EWvkBupZsfgKiJbR20Gbl8hHe2ih/m7zLQudYBUgyG5yhXJg8zofMCPk5v1raJzKPphfTsZv5+XqxVt6c1m6i/Vzad5cb2YgjNFsoTGg13axkRU3Toz+s0OqG7U2dEIuIaArLXuxZjsi0zaEE+quDWn0+vFjyp7QgF+wdSwJ+OXRftF5SssggcKSQnavZZvHrfHA8Bml5CPdk0QU36RyYMWBGtP2i9LRmBFeT+Xrs6sXLPtrzkCyx91VpsW5v1S57raMn148vjk8iu7pwUMa0/qLwqGY1uU+xs9ZqcbOhsHtEZCdBZy/4BOl9CHflUs35XqoriZ3K/UVUmbgSePX119RfdDPoaR23OTIE2mOovqVJM5uV2yetzSNf+soFOBVVMOFwIyUzVqaSN7FTSyncIajvUyPzCyuuXYAZaTALUzelWP29HLf+VQpwJQ4dj0MtPJ4i5kTDtlNEYoSSCibnAIIXboJzeS1ZtA2X3/kKVTIYSge19muiro9Lc+hVzMdXj6A7qloQ8lLxd3/uiLO4U2hizFjV8DEEKMnRLKrf6yXQUM/BIyOhW7p0GxZooVnZijePNKn041TJXCqquvFUUd+1j9i+JLo3W2fGHTsLs7FfLPeXXrQnCwEz+5lVK2p2FApFidChkQXBTJTPY0zHkAylQs9+P7smJwuIrx/Zxkotlw5fD9zsAvEnGsM7lY0lJnTIFnyl3bSjat/bzyskIQjphbPucy0dXj5sxvx8/UdSpkWvDYpsxkTeszQLkfjZf1glDGPPc5iUIzy2Kee3AHV2dKGLRI6uBrfR/+Y/WvfmDM2o+n79QHjVkxnp6TqDNDoyKrNn5u3++f15mQYcF1tcwUHrMaILkfi5fVgnYVY/EUM5cj6R7/6WfeHSEasTiT9nSzajokPB4k29Z+7FzWCSISc9PnXNi5eqxVNLrATa9TIduC7gWZCQ9azdGU9WnKtdL0nWtwAuyC6zpTuFzneNnIH+n6WKeKj1t39+bKFfnp12Rjeeej+R/kW+KdjfmWrBMYu2wDQ4hwwk5KxBshKZouwKQo9IodhwwOFJCLTMGxqyma+3n0ZdVQb7XghNDkZJ7EiDWadTB6MH4JyWJAE5pCsp/LPheVgpDEorzmpCLsTo7mcRPou+/f5NeZ0AAGnVgyU5h1NYVzT04vGgThjHJ6kh9mcDrCbxSO0B8QJheaxTrWQRUcj25Sn7+8UuQh+6j/31NnahOV0Ovwk69qPnAl+d11MbrnhLoauxCQk5bwZuAiIPxXQ3QqHt9gE1/w6n5ZXOub5RoPHXvwqFhSvIMb4D0QL2oOQ5J2Q92qEAyciZMM8GbE6YwGwOEixILd8RueechFpkpZ5cULYCZ7cnNRGOwvUW5OWr6svzin87NOF3g3hQCwg08P6pTmIlN1zGE5oeuy+rD49J4CvBfjMW1wUTjzcgVg25OIizpA2KJE3CgZb9zpuRAfthNaF0XPkOlrAHWoyQABFTzeHI5sCuCevF1qMyNexYsJ9E+unpy76sl0wAiM10LY+Ae6VK2dmetC3zW6up9OKCgqBY49On2lzv5GXZyevzl69fLYH0ZWF+mOeTHhaJnef7yhbkio5dxZLpSLKwcnIQCCUeVAVFkHTcJjlMi0J3Bt7s/KeiP4YvF7ciMyvO2BVnPYwoacCkBgmTb/etiTLjKdOfE81EOKr0KOGBsrD1nYnvRbVACNWlHZ5ZzLLlePmzjdQBVXpwJcs4edxyJTNdhfUYj3zVivpYsIO+2Rjw77sgpN18ciP7K5KFfnKsy5edzub/kvtutUCD/sIRWZKvxeFTc3BFl7GPck6aJSCMYoRzdqzE7XdTh6Bx3sonAsEEbsDxSZjlisJN1115P15L2fa8gC43FdZB1hR42Sdy7VnFuF4tpjClgrkGrWgxsmFSJThdC4nOj4ODp4VBChxroMuiVsgQH4g21waPbkvN5KFNf7YMCTICSZz0nUajJWF5N3t2u1eKt+oOHp4cXrH746oHBuM3JUPSuXk8Ivt5DLd51enL288lnS6U5S1OgoWecyy7kVGa6vwoO9UKGY/MuSgh5ReDAdscc28zEt5qbFRB2VKwqw9LB6+J4iX2l1DjSnCR3kUPNTonfWjQl9cu78NvrJ25g04HSBTzC5S7cnQvK2uch0/sYiBQYKEeEzBEXK4XRdfGRxz4WRc58wMrl7Rog7+oSRCZRBEBQpXBwGRaYOgCKSgmklWRA554LI1ePGTvzXeo91KndgekW8ZLpY3gQ/vJQfDrUx5ex4XdVAHyD5Xne5bh63bQPLJaH5S4oa4WaJIJehZiUd8rC1DDWL+4VJOHfrZSeJBV+/o0/GienxYja7n0/Wn8KNEydDQo1LuT6paxHrrNzDa3Vu7ajtv4hKjRPhUyhQXTldTAt1tZiGGygyhhqYcu9R1yTWQO5ztZq0dQP9Ma1yIbGrwy9WDXx14Z1GT3ZyhFqWEh5RvzDWMu5EtVKxdcsA4feJ59opBy+YpHju1cHFD+HvLJhUCI2UcIO5bVy3y3gEu5KRG8XXLXMXcrUbXOC847mbaOedADI800UcGXGOIoQMT4qmHp4qNI5xbyrTu80OvadsqasLKkUEjQQd18WtOinKJf2noml+8iEzhxqbxMNsnUON5DzMKrLWDivv4oQaicKgXxnJh0EVVVUdPnikLq/Ib/fqMTmyFB3P9d+plO+Leu1k+YCDR+Vlcy4vWz1uZ2DQ+h2l2O2+Srvmixc+f5ESbUpXNmX1FZWOzbl0bPVYO0KyvpeKH+tUfHOpjpCrKhu4ojiYFN1VjZragnhx1BZkRYAtmEYG99i45GvuSr5muf+GGsEgY4gLW2gPg4wknmYTojoIjFhA8ZyLu1aPm7195MkXOq3bIMx8WOilDgvifXFTEHQQtT5K8rhYa24e62Gh59f5O9KpfK5BHQQBu6NFJkecoy1mMrJfmt0kxfergQhMJ1y4NbfCrdUWGtwBEjKoW7sBY0m6AckwfuEwh7q6SUNqRTXZAo1TKy7ImpvH2oC6XeDy8wuykhorCb08fnniXzeJTDsGJAegdOw4z0ydlNJcghaQEILcJWikWzcD0KGXrNNoLMjiJgo870xVyCM/BROviI85okg05lTuQB6DWtgMp5tGhnXLRYBHTCib1nvP2mbwfoTIBAedIE4ylnXaOCPYMUKr2rIPosV5q5ViNQNN3h8g+xC8VV8eObCDzJubYlmogTrTgpgqf+IPk5ILpdW4fYgiUYvt4V/2mbl9cG3W3NVm7Q1yQNOERqrzqc+9jccNF29KHA9k+aD5tjG8FaL5fIfdqrTaq0P+hcVxLiRIr4zujb45pbd631fHwiqZuo3GxkpvljyirczxLV2A14ejiversjrHAtznopSoTUgZWARKdNucK6PmVhm1vk6FTIJzxGpoXC3ea8GfhxUYoxlqaTN9G12jlAnVVjzAK7jGaG4etyFJwEJNCGL6WjqALZWhoCPUQQqSom8a5Z5cmzMXEpEbe0deGb82pyEReGdVZPqpukkraip1QxGwUcem0czcuruETuSm0/iBpaGEc8Nm3aZNOGlZItqAtlqgoc7DHZ9Cb9Keb+sAgV5qOed0zbpRKxYpFTHRN49e3TISks43dw+MUgQ9bzCxXChKuuf5Al2I88DqjMd0QVeZ6YoqHfQ4uJsW60oi5KD6Y3Vmxs9CZRWanpGRopgIQIsHA0JMu5yVCvHG2k94CIcjfPELX50XheyL5Z7HTEVtYDfkxYS6I6euRsdyM5d1vacdjnKh7ugaJQwLIjNVlwToKIC2Q31W68fJsjwwhIe8dKuV3rVZkjD8zB/NbacOja0yacPc4jHg3ZNbJdeozK3oorXGHlgrCQFHF0J8z0Jkyp6oc+rIJckaT9d08wlBJctqCpXIj4zONpzTXwEV581GfXLj3e35A5TlfrnJjj7SjM+jikz5E1LcXZFalz56osY6HpJ/sPPrRiZHEJeVhVhVdDqwwuIqkrlVkazHOASVP56DhgqPcEJ6sktQmahRt3rC0D1zeX9NCp3mEDQAbc9jqFL8EswQFoPAaMalJ3Mrs2hBG/jjCetUPl8pgYaDBcpMvSfq6H56s3in3pHuFkBpz4thUuMSoVSx+iwL7LZydcncqktWDgsKVuGnIn51Sd0PcaBA/WoX3MMnNhjZUD2sopJ9V0m0gzWeLLPx0JXmS7YAhCyLn0Sw+oe1ZQHy5lePNIhhh5jI1H9Cxwj1VRajd0Iqkf5AY7lf6rHBwJXmbo4KP+ZGXHBLc63WYIVUf4Bsyx/sQdsWprlCH3HwRG3UYW6IVejIdwa0HEG2Z1AHKTyJOmR0ZWDEBB3I2Mqg7w+hRY5FOGzhQ7ki07Ae60Mz4p4XytIEFXVz9JGZ0IzI2b8rqNjJgPhbLmQLHZ6VY6oqMo2IZ2kHpQFJrcglNwWXE2VpjYerNJe0bTq3DMG0OK+3Mot2vAIKJDmQWdS9MEBKOZHOOgytXI3NbbLJfLUupsRWiX+B6XHPC2Si8pB5RY/AciHF3AopmiNIGd1c9E+OUkjxjEJ2UWoNHSapIlOWPTFz4kqdlMWNKol6PZ+s12Rvq4PK9ujy8KfCD5+sQFPTS9NorNHIeKgZbntcpTG3Ko12BhiBGQCoNGoAMXUVmTKi+UeFFpI/pxC7em1pXMgAMBEdrjFgaZHbbOt7zC0s8OJHb129xgzoNR7lUq9xa3CBHX6RKSOu/+KULjA+e6VOTp+dqNPvT9TJy2dnp1enFyo7fEJuuMGTjjrzHl872alDYwjTPPJRwcfc6BluJ1Arb2idiYhzAMFHsrhAeECRKSPif/pxTfce9Lry2RI4K2RRjYFKO0wSlXXMjSiiA1TlqK69ruCUt5R1XNOiUF/iOZ4uVuUNReCkmNb6+u6PEwoq/N20oEjC58W8BNdbxNuiG1QiPRrXo3KPuZE6dNruXA6jgCtggwooN+4JQSsXvKgKBCJ69thIODpAcJUUsF3NFRCBMnUNy2vytGztAoy/0hsf2dQSEoyw+VG2zTUYc6vBaK/AwcHCz7b12QbMtlM0GGst0EbQ8cokbn8L2UUIYPSyGNddzIUC4QZI5IYHCoz6fnlgpmfMOCzm3QjKVkdHRBMglMlueK6+mLvqi0Q5wfFuoVb4A/Frdba4n6/ZyHyp4SjowpFviFZZrt6oh71Dumc+I9p5t74FnhhRWHzYTjtrYlsa8LxwLcXcSgbWVgamLCE/aDvZT4uQKDOxV4pcyJSiSok5V0qsHrenEsBNJyFO+PpuPRnTQemTCW3KTK5J2ZtYzfMlSew+vDw+eHZ+/JV6rLrDnrpbLNcr9fD1yXNgLuLFUXMR6SEM0WMoXOYwt2p+tbkAFzCQOdSjO/YSiEzWxj4jhK1OXYtaQSCjp665ymEuJAM3gKJRHugd0igfiOQjMn1+SNtoqsumQ0j5y0NDGr+3x2UPUXhgoUWYZGeHg35CV5VHYSJETFQEwhE9BGNkDLc81Koa1pYFbvcJqcAmJpKGRituLmoFMYlyc65gmJvHrc6O13dyrFO5myxJpjHIOwmm0ZCji4pAGKIcnQsU5uZxE7sQnagVqn9NTCMNjVa0W9QKYhKl3VyMMOdihP5z/GQanIImmcawnzLBS+IcGzXSXNu2VaHBk3u2rdZgPWqAzRKgNagneOzZFpma2FMahJwipxLGNGJtcQkAyXUIcyHYtwEUTfBAkVBP8PiKkcj02SEV70+EVORCXVQkC0HLz7lY5T7rZkPH44WKYFJX7Q3jg7h4b5SKi/QQjSgV53KEuXncDuLgnLTQ+GtiIUlgtCLVolIQkiip5sqDdJPTDVOcg3W+EPFLMoysn2IZDc+Vp6kJ1q0K9Q9OgI244NYyQLghocbXxDLS0Gh1JDxNIzC3TQxhwk+YWP27+ow9YH5CjC/JNIgTJAwaDW8riorAHhIlwFzVLzePW9MAXnqhktfENNLQaHVyPE27r25iyDT4/UNXu48UgxEmnK1+vvWiUNmLzydptxJtq0Iw8AMgVpXPEhX/TejjXOjcffb1onh/KuNIY8O2iSFMOBu2onNanFwfj0VzrDiyQYc0Fks631LcFWMKc+TfrxGyd/HPnsZTozJ5XS6TVz1+eFo30T9Z6FTukjjvJTRRZoo1UaYHg15dY/wVu1zernrcNtHfwXUqt4nEfuJfUWaKNzHplENd41AT2SGHrtWnq77iMPdPbzqV28RBiqHKTPEmJp1CqGscaiKLNda1gnJ1E/07FTqV28RRP+krNiNushBoqLELgV2jCrdx2elHt/KbYQesGWV694Y4PCW1U0iaEbQhc7IoiFPMtdnlAnD60YfTEOLkD0mr3RVwi1QWkjjktXF3yqIgTjG21+XKcPrRhxMKNynTO/aENxlkpsRxsw0DlEVBnGL+0C6XfdOPuzjRKArtCZ5ewL4aWUji4NvGRyqLgjjFfKRdLginH304odD3Mr1rTzDqicyUOIK3OXcgi4I4xVQmulzuTT/6x3H/fqlMX+OkfX8BnFqwTllSGr2WuSBOvEaBqZ2LwVFkQTCeA7yALJzGC14EkYWkjeeipES80mTeZI1CeDEHaFeIptU8YfQY4eVXbtZ4wesgspC0cV3ULBWvNFYsXh7Ci7Njodi2YckQL/+9P40XPJ3WFYWkje8iUypeaRRbvDyEF6faQjNtQ7khXv4rfxoveN6jKwpJG+dFplS8ks77yhqF8GIO167Vgquci+jsq07lzAr+M3fnxXp8S+vOG3V5pr570j8cPMtVdVaIjgrRD2/Oj+mHjM6XKa9fQZTyNLpcSxN0q5sYwoQ5XLtW+KzepPEfg9WpPhcmXQxJM+erqNRTNN1Fpd1I1tQNAl49bk8f+E9k6FT7Q0JnyyJm0uwogqgVxiRKvY2w23bJy3TeKKaTNxJel6uotew6FSYBM5FcO7wFLyqFIYmybC7t1hXaZ5tRBa1GgMgbjdohNpRwAPgzDEataHma+JvGKRwVvMvF3/Qjc2Iap24nRwFOZHp3+RKgTayQvUb5gKm2Oeog2wPHtdiVvi7XgtOPPlyRqJ5M7+KK6RUvZK8hIDB77mjMpYQ+lA1CwFrNvMAcSrJvjmpE1ys1R4MkGgmA1JweCTAP45l+orFV1CyRt4lcENfY0Yku16DTjzsG2yHehvgtUKPT/BYedxKF/FR0T1QtFdi0BYR4echw+QJC6LhtSCAEGC8gsCObF/ITjbSiJan4pi04xMtD+PIFhxB4q/HtQXz98au1AWMHOC/kJxtxW+0sCADgyMBfHgKYr16EplzFwfTICzwsQPSOAD6EQrtdnumnGnlb7UiI9kB8+ctD+PKlEBfKA9H6ukAoTx9RwH5RkemqXK3HpDikXl+dXKi/oEhVf6Ven5+p16+uaIVEN278y2hZdFxUUNYWYhZdIhmNvO0SSUjm+Y/kdoXMXU2qNFbYJyoyQaxOnhNWz0/8ej87RadglXSIRb85QvG50l6XK+11/NHYdSp3/nexwsReZEJYEUoaq2cIK2/EwsvTi1+/PD71meKprC20K6sIziJqsMgIXaNX5xgWX0XUJ3/gesgn0aejcoQiVukyXazbdkdvNJMwbEkHZmT9AkMY1/Lr+rX8On1Ezn1afhY9PJiJTC07qCw6oYNKWUB/MCYJQgg9vrQRknxWkQoFqj3qivS2yxrbw8ObyNSyy8qiU9BL29AQbw6hxzc0XDlBimniD4rZFXKC9TBHB6g68ErBTqb75Yfyk503T8rV5N1cPXy2ooi3c/9V1Z0XhIY2UUO/rJd+o54FukKehY9tRt5vO7YJzcEq3BAhBXunfwUyDGMlMjXHSrwgiBVPi7BKvTXZNXKBDmDVIZy6I/ojYupMvplTm1RgEOOZnq0ouOOEQqyeLIsfdeCTZ5l6mKlVuVYFBaxVf6m65mG8uPtED6vF2zUd4rr79NUj9cuzf/FIHb3+VW2Pi/G9FsnFYbB2KhxEOCm6YQUcHSd0Z2thi3wxIRQUNxBDW/QfV9IgB8Y6nunnBVlUOAhyUmDErpV8DIHMVxRCl3ELMlix+cQficvoDh+wZL5r8vOCLMIrBkFOkurpCgxC8w87Mk8yR2wY2ILtD+cm09czkQY7YNG8kJ8XbHHCKwh22lpGYBYCmx3MJ53PrchuNgSH0HUqd2Q+Xy5+U47X6qyYF++Mkrh/nSuykX/r+Qv/MkSkRPNPVdnIXM1O2XetlmM1Rw+B0K5OFd0X/aFYlreL+1UJNseFuGSotWmrB1v1wDDFZSi7VrKx7jHgLLrQf6RmUUx3Ekyu3BdnBamOlnOSsSnJgfG9Ir0b9akslv4vLN71/Wv15hKsNNNkJ+s2BCyYy052rUBj9YFpL8K/ySvUHusbNueLH6l9lXRTRmem5rNHiiTTq/8eHtJ/Qbs5Bw+2O42uR3Umu1xnsnq0t60oeKX/+ohO5fZc46d6WLf+ajIr1cliRp9cXZRvp9SnFzMtYwXYt3hZsNVpmwBWLDP0tbnz3+hLbtTNQPyNrhSAvDeh5C/vppNxiSyZc9tg29Ic8LaqobZxxzuTi6TVO7Bk4dO+unqhXi10JKCrJX1I4rzAZHm+YAPTuKqtb6iBnKtaVUc7PmX+OEddLqbo91w/162u26uOy+V68nZC8Q2vPynzp2MKrTWeTMFILeQag1CkMUrbshAUnFEyZcfOEI1anGr9ophNaGQmKegpaYvR5Gt+Xl4Xk/f001G5vL8rKKjq/D0F16YxfTUBlsDfGmx+GtezrQk1n3M8K8tYewXR4MXJz7b5s3I6uaUmNmo1f1mw1WmkyzbC0+qD1W1Zrk+KdfH0m82Pyvz08ubbB90HB97f98DvD/Xvt++8LNf0WFIo0+W8mB4tFu+dx1eT+fun/x8AAP//AwBQSwMEFAAGAAgAAAAhAFxuz01FAQAAuAUAABAAAAB4bC9jYWxjQ2hhaW4ueG1sbNTfSsMwHIbhc8F7KDl3Wf5OZd0ORK9ALyB02Vpo09EU0bs3ittkryeFfv1InoTyW28/hr56j1PuxlQLtViKKqZm3HXpUIu315e7e1HlOaRd6McUa/EZs9hubm/WTeibpzZ0qSorpFyLdp6Pj1Lmpo1DyIvxGFP5sh+nIczldTrIfJxi2OU2xnnopV4uvRzKAmKzbqqpFs9OiaorCFH13095ys3/8cr+5ufi6tS8JBqd0yaXTjnzz7bnxD8gKddw1Vkh8UgcEpg9zB5mD7OH2cHsYHYwO5gdzA5mB7OD2UFoIbQQWggthBZCC6GF0EJocasWZgOzgdnAbGA2MBuYDcwGZgOzgVnDrGHWMGuYNcwaZg2zhlnDrGFWMCuYFcwKZgWzglnBrGBWMCuYQYaYI+J6HvBXvm78ocnzzN18AQAA//8DAFBLAwQUAAYACAAAACEAMZdFw2ABAAB1AgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfJJdS8MwFIbvBf9DyH2XtJtlhq7DD6aIwsCJsruQnG3BJg1JZrd/b9p9uKF4mXPePHnOIcV4oyv0Bc6r2oxw2qMYgRG1VGY5wm+zSTLEyAduJK9qAyO8BY/H5eVFISwTtYOpqy24oMCjSDKeCTvCqxAsI8SLFWjuezFhYnNRO81DPLolsVx88iWQjNKcaAhc8sBJC0zskYj3SCmOSLt2VQeQgkAFGkzwJO2l5CcbwGn/54Wuc5LUKmxtnGmve8qWYtc8pjdeHYNN0/SafqcR/VPy8fL82o2aKNPuSgAuCymYcMBD7cr5Sj0pbtDjuiAn5XaFFffhJW57oUDebssHXqkVmjqlObqRUY+jiXKSr31BfqfjG91Iu4dAoijJdiMdOu/9u/vZBJcZzfoJzZMsn9GMpUNG6byVObvfSu8Keq/0PzFPaJbQwYwOWHrNrvonxAOg7LzPP0r5DQAA//8DAFBLAwQUAAYACAAAACEAKQ/1xn0BAAD+AgAAEAAIAWRvY1Byb3BzL2FwcC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACckstOwzAQRfdI/EPkPXUKCKHKMUI8xAJEpRZYG2fSWLh25Bmilq9nkoiSAit287i6Pr62utisfdZCQhdDIaaTXGQQbCxdWBXiaXl7dC4yJBNK42OAQmwBxYU+PFDzFBtI5AAztghYiJqomUmJtoa1wQmvA2+qmNaGuE0rGavKWbiO9n0NgeRxnp9J2BCEEsqjZmcoBsdZS/81LaPt+PB5uW0YWKvLpvHOGuJb6gdnU8RYUXazseCVHC8V0y3AvidHW50rOW7VwhoPV2ysK+MRlPweqDswXWhz4xJq1dKsBUsxZeg+OLZjkb0ahA6nEK1JzgRirE42NH3tG6SkX2J6wxqAUEkWDMO+HGvHtTvV017Axb6wMxhAeLGPuHTkAR+ruUn0B/F0TNwzDLwDzqLjG84c8/VX5pN+eN+78IZPzTJeG4Kv7PaHalGbBCXHvct2N1B3HFvynclVbcIKyi/N70X30s/Dd9bTs0l+kvMjjmZKfn9c/QkAAP//AwBQSwMEFAAGAAgAAAAhAGwfjagtAQAAEQIAABMACAFkb2NQcm9wcy9jdXN0b20ueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAApJFBS8MwGIbvgv8h5J4mzZatHW3H2q4gHhScu0pJ063QJCVJp0P872bMOTx40ePH+/LwvHzJ8k324CCM7bRKYRgQCITiuunULoXPmwpFEFhXq6butRIpPAoLl9ntTfJo9CCM64QFHqFsCvfODQuMLd8LWdvAx8onrTaydv40O6zbtuOi1HyUQjlMCZlhPlqnJRq+cfDMWxzcX5GN5ic7u90cB6+bJV/wI2il65oUvpesKEtGGKLruEAhCXMUT+I5IhEhNKdFFa/WHxAMpzKFQNXST78rtp51cIt+eLXOZFNW5VXJwmhFZ1M6pTGLShbNyZrEnkpXLyFN8LWe4IvGP4UmF6H7pwe/sxm5y8eub7bC/PCjhFEU0sA/NQhnlJHfbPD1l9knAAAA//8DAFBLAwQUAAYACAAAACEATEr2LzUBAACYAgAALQAAAHhsL2V4dGVybmFsTGlua3MvX3JlbHMvZXh0ZXJuYWxMaW5rMS54bWwucmVsc7xSQWvCMBi9D/YfSmCwHTTVwxhiFdQ4HavVWhkDL1n7tc1sk5DEUf+9H2OCgrDTdkjIeyHve++R/rCpK+8LjBVKBqTT9okHMlWZkEVANsm09UQ867jMeKUkBOQAlgwHtzf9GCru8JEthbYeqkgbkNI53aPUpiXU3LaVBok3uTI1dwhNQTVPd7wA2vX9R2rONcjgQtObZwEx86xLvOSgcfLv2irPRQoTle5rkO7KCAqNAyN59SrkbsldidrcFOACkosK0Dl97m2n8xGLZ1HIWss4emHjZLs06hNSd9f1w0MMev9RiXSbQK2xAkB2FU7O4ShaIdnCNZ3htn5fJyy8Z2GMIIlO9PdpvRmNo8UDcgv21m4q25wshSrD0OzHMKHX2+n8TztnWf8oHL34T4MjAAAA//8DAFBLAQItABQABgAIAAAAIQBYcSD4iQEAACIGAAATAAAAAAAAAAAAAAAAAAAAAABbQ29udGVudF9UeXBlc10ueG1sUEsBAi0AFAAGAAgAAAAhABNevmUCAQAA3wIAAAsAAAAAAAAAAAAAAAAAwgMAAF9yZWxzLy5yZWxzUEsBAi0AFAAGAAgAAAAhAEt69vaPAwAA4QgAAA8AAAAAAAAAAAAAAAAA9QYAAHhsL3dvcmtib29rLnhtbFBLAQItABQABgAIAAAAIQC2qBhKFwEAANkDAAAaAAAAAAAAAAAAAAAAALEKAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc1BLAQItABQABgAIAAAAIQCcbdIz+Q0AAPZVAAAYAAAAAAAAAAAAAAAAAAgNAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECLQAUAAYACAAAACEApI+SbIAGAACuGwAAEwAAAAAAAAAAAAAAAAA3GwAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQCwjU827gMAAJQMAAANAAAAAAAAAAAAAAAAAOghAAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAOarwgD3BAAAGxIAABQAAAAAAAAAAAAAAAAAASYAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgAAAAhAB9KVts6OwAAKYoBACIAAAAAAAAAAAAAAAAAKisAAHhsL2V4dGVybmFsTGlua3MvZXh0ZXJuYWxMaW5rMS54bWxQSwECLQAUAAYACAAAACEAXG7PTUUBAAC4BQAAEAAAAAAAAAAAAAAAAACkZgAAeGwvY2FsY0NoYWluLnhtbFBLAQItABQABgAIAAAAIQAxl0XDYAEAAHUCAAARAAAAAAAAAAAAAAAAABdoAABkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQApD/XGfQEAAP4CAAAQAAAAAAAAAAAAAAAAAK5qAABkb2NQcm9wcy9hcHAueG1sUEsBAi0AFAAGAAgAAAAhAGwfjagtAQAAEQIAABMAAAAAAAAAAAAAAAAAYW0AAGRvY1Byb3BzL2N1c3RvbS54bWxQSwECLQAUAAYACAAAACEATEr2LzUBAACYAgAALQAAAAAAAAAAAAAAAADHbwAAeGwvZXh0ZXJuYWxMaW5rcy9fcmVscy9leHRlcm5hbExpbmsxLnhtbC5yZWxzUEsFBgAAAAAOAA4AqgMAAEdxAAAAAA=="""

_BOQSVC_B64 = """UEsDBBQABgAIAAAAIQACMCm2kwEAADIHAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADMlc1qAjEUhfeFvsOQbXGiFkopji5suyqtUPsAaXJ1gpkk5Earb9878QcpVhEH2s2EmeSe8+XvTG+wrEy2gIDa2YJ18jbLwEqntJ0W7GP83LpnGUZhlTDOQsFWgGzQv77qjVceMKNqiwUrY/QPnKMsoRKYOw+WeiYuVCLSa5hyL+RMTIF32+07Lp2NYGMr1hqs33uEiZibmD0t6fOaJIBBlg3XA2uvggnvjZYiEilfWPXDpbVxyKkyjcFSe7whDMYPOtQ9vxts6t5oaYJWkI1EiK+iIgy+NPzLhdmnc7P8uMgBSjeZaAnKyXlFK5CjDyAUlgCxMnlq80pou+U+4p8GI09Np2GQen5J+EyO7j/huP0jDqADHKwwL9rOkO+/Nb1D+9onNinSpQSenpdTJJkThhhXBrDpu5FETzmXIoB6j4Hiq3GAfe0THFIYOSzpHje8CDvdY/4ULqPgPFLMBjgfYJujdXXLkxCEqGGXpIcSaedIEX3xjOuTbRWoc73lHKOrLrZfyxww5+mP1/8GAAD//wMAUEsDBBQABgAIAAAAIQATXr5lAgEAAN8CAAALAAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArJJNSwMxEIbvgv8hzL072yoi0mwvRehNZP0BMZn9YDeZkKS6/fdGQXShth56nK93nnmZ9Wayo3ijEHt2EpZFCYKcZtO7VsJL/bi4BxGTckaN7EjCgSJsquur9TONKuWh2PU+iqziooQuJf+AGHVHVsWCPblcaThYlXIYWvRKD6olXJXlHYbfGlDNNMXOSAg7cwOiPvi8+bw2N02vact6b8mlIyuQpkTOkFn4kNlC6vM1olahpSTBsH7K6YjK+yJjAx4nWv2f6O9r0VJSRiWFmgOd5vnsOAW0vKRFcxN/3JlGfOcwvDIPp1huL8mi9zGxPWPOV883Es7esvoAAAD//wMAUEsDBBQABgAIAAAAIQCsp2jopgMAAD4JAAAPAAAAeGwvd29ya2Jvb2sueG1spFZtb+I4EP5+0v2HnNWvaeK8FaLCipfkllPZZSnb3kmVkEmcxtckzjlOoVrtf79xAhSW6sR1EdixPX78zMwzDtcfNnmmPVNRMV70EL40kUaLiMeseOyhr4tQ7yCtkqSIScYL2kMvtEIf+r/+cr3m4mnF+ZMGAEXVQ6mUpW8YVZTSnFSXvKQFrCRc5ETCUDwaVSkoiauUUplnhmWanpETVqAWwRfnYPAkYREd86jOaSFbEEEzIoF+lbKy2qHl0TlwORFPdalHPC8BYsUyJl8aUKTlkT95LLggqwzc3mBX2wj4evDDJjTW7iRYOjkqZ5HgFU/kJUAbLekT/7FpYHwUgs1pDM5DcgxBn5nK4Z6V8N7Jyttjea9g2PxpNAzSarTiQ/DeiebuuVmof52wjN610tVIWX4iucpUhrSMVDKImaRxD13BkK/p0YSoy2HNMli1up7VQUZ/L+eZgAHkfpBJKgoi6YgXEqS2pf6zsmqwRykHEWtz+k/NBIXaAQmBO9CSyCerakZkqtUi66Hf/YdwMgzmHz9PA302//xHMFo8zAT/m0ZSm77MaVmvMhY9LGheQhFQ7ct0/HAgTHJaBf9DmiRSkTEgGi3j9vnHyABx4e/kN5NCg+fJ+AZScEueISGQ9nhbrxOIOLaXRSR8vPzm2sHADb2OfmU7lu4EVqgPh06oB05oeXhgD4Zd+zs4Izw/4qSW6TbXCrqHHEjsydKUbHYr2PRrFr/S+GZuP7rqf2h2a9+Vw+pWu2N0Xb2qQg21zT0rYr7uIR1b4NTL8XDdLN6zWKZKVqYDJu3cR8oeU2CM3Su1D9SvmPXQEaNxyyiEj66aI0bGAaXm/gRqTa8VjeZv1Z2K4aJWfRNkpAlfnSEmMVY+nVjD9bW3hue9tfWmtX1gDc97a7sRyI4S3TQ1k81pQgW8RSjQPJl73ew0m9/aFNOEFTRWBQ0QB6Otv8tNVuSXy5CpIh0TSVakoqrOI5I1sVAxgFCnLI6peq2hfhui3y4GF9i/CC8s17k2DoBB2seHAlI0E5rqmoB2sWl1FWHge1PJpocaZZDGodsZmnYXJBziUHdw1wQde47ujkPbvcLjUeCGSsfqPelvFGLyzuuvYzS7KZE13BvqymjGvmrD7ex+MmkntiE7Knt/PlaubHf/l+Et/A/I6JnG4d2ZhqNP08X0TNubYLG8D881HkyH48H59oP5fPDXIvhzd4TxZkANyDkIZJd5Y/fXp/8vAAAA//8DAFBLAwQUAAYACAAAACEAaqp6VCcBAADzBAAAGgAIAXhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAvJTBSsQwEIbvgu9Qcrdpu7qKbLoHRVjwpOsDhHTahKZJyUTdvr2hh3YLS7yUvQRmhvz/x0wmu/2p08kPOFTWMJKnGUnACFsp0zDydXy7eyIJem4qrq0BRgZAsi9vb3YfoLkPl1CqHpOgYpAR6X3/TCkKCR3H1PZgQqW2ruM+hK6hPRctb4AWWbal7lyDlAvN5FAx4g5V8D8OfXD+X9vWtRLwasV3B8ZfsKCCa/EiuTJBlLsGPCNTKg2khF6G2KwJ8WtdixLAzxBTCulY2cRgHteEQckdVJ/ehYHjDLRIx2CKK3emiMHkV4bJYzDbVcfkBx02b3q1OMYx+4c17X3YZ5jdx5COZ7QF92sywMmDM1y/K9POKOdZpOfRREYXX1X5BwAA//8DAFBLAwQUAAYACAAAACEAPEJEs/smAABcHwEAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbKSdbY8cx3WFvwfIf9jsFyeAxZ3qeSdIGpqhFhEsR0ZkOwmCfFgth+LCuxxmdijJMfzfc6vnpeucc6tK4xJskdo7VXP7bHX107er67z6zc9Pj1c/bnbPD9uPr6/Di9H11ebj/fbdw8cfXl//8Q+3Xyyur573dx/f3T1uP25eX/9l83z9mzf/+A+vftru/vz8YbPZX1kPH59fX3/Y7z+9vLl5vv+webp7frH9tPlokffb3dPd3v5z98PN86fd5u5d3+jp8aYbjWY3T3cPH68PPbzc/ZI+tu/fP9xv3m7vPz9tPu4Pnew2j3d7y//5w8On51NvT/e/pLunu92fP3/64n779Mm6+P7h8WH/l77T66un+5df//Bxu7v7/tGO++cwubu/+nln/+vs/+PT1/Q/l296erjfbZ+37/cvrOebQ856+Mub5c3d/bknPf5f1E2Y3Ow2Pz7EX+DQVff3pRSm5766obPx39nZ7NxZlGv38vPDu9fXfx0d//nC/gzxX6PhX6fY367fvOrHye93V+8fHveb3e+27+y3EK5v3rx692C/+ni4V7vN+9fXX4aXt910EiN9kz89bH56Tv5+tb/7/rvN4+Z+v7GvD9dXceR+v93+OX7wa/vRKH5Z/4HY5939/uHHzXrz+Pj6+isT4fl/D9/SvfzP27fdZBa/5+b8RenfT1962w95y/zd5v3d58f9evv4Hw/v9h9eXy+vTz/79+1P/7p5+OHD3jKaxi7vt4/W3v599fQQT0Qbf3c/H9I9tA2LF5NuOl+Ebnp9df/5eb99Ovbai3JuaSn3Le3Pnw4tx9Nf1tJ+SX1L+/PYshu9WEynk9liXv7OybGl/XlsGcYvwmQ0q+RqvfbfaH+e2nUvpvPRuHaQs2ND+3P4wpI8Nwd9+9/b27v93ZtXu+1PV3bGmdDPn+7i/BVeWm/xFzKevBifj/f8W7JBch8bfBlb9O3so8/20x/fjF7d/GgD4f74iZV+IuAn1vqJDj/xVj8xxk98pZ+Y4Cdu9RPT8ydu7PjPIsRhPohwOtCuP7hpPCfHixEfpLU5CzCjw0tjczqwNLagQ4p5vL6O3/v+zZ+++fbb3/7x9//8ZffrX/13+J8vv/nm6us/fPW7q6//7eq7//rO/varf/ry5e2vZ7++/fKb7776l1c37+OvgoS+9bOEg7fh7hz8OD34JR+8tTkf/JIOPo3xwacxPviYhxz8uOXg/Szh4O2MdQ5+Mhx8N5rywVub88EHCq4hSL+NtxCkEf9VTEWOf3LJ8Y9eLGfTZfoP6X+bSR0ksVMfZoSHd+82/Yx8Oi/sA8Px01m5giCLA0EWB4IsTkxKxJleJA5NDplEQQqbEMtSzGCg0BGtrPkgFE1OawiyFhBkLWJWosWsRYtMoqDFvKaFfWA42mGi7S8aKwiyFBBkKSDIUsSkRIp5ixSZREGKeEugl4sFjAXKdGVtBnX4egFBFgCCLEBMRQRYXDRpvph3yxN3xj/5EpLJHBQxsCufKPaB4fBpVlpBkLWBIGsDQdYmJiXaLFsGRyZRkCLYHWRZC/tEvFQfwGLCV5BV7GDQiq6Sa4yyHhhlQfrURJEwapEkly1qwpQp15RgnxgOmpBihVGRBNqKJBAVSQ6MSMgVQpMk8I1DtigJMufVB5Uk4c9uIQgWUrTr+EoL0SCapG2DaOJiaLiIQymf21y2qAmiqKdJinQdX28DREWSNKqSQFQkceE0XESnIkkmW5QEAdWTBGB1QSfHKqTI19FxrSGqmhR5NbbV+TVcRqx8x5bJFjWpEqqVFIbZpGNExahIUoRUaKtnjoupoYlTc9miJFVSDYCqC5lhUwTsGNCscTI7y2xSpNXY1hkmTbwK+STZoiZVYg0p7HWMrBgVSYrQCm11mLjYGpq4NZctSoLk6s0mSLFy6qQw2DGqhSLHQpRv/mPQGSUXkaxMsJlkUZIquoaU+jpmV4yKIkV6hbaiyAlf+zNvv8OiUGji2FzSoEyHIHuuiiX0asUROi9WsdUZ5DpGNYjK9QaiUhuL6Whx7CJ4lepYJlnUoQqvXQp8HU+tGBVFivAKbUWRE7u6I6RrYthc0qhMlWG7lGFHU55iLTwMljEDG0R1sBQLqX4ltQlhc8miJFWE7VLsGzPCYlQUKSIstJXBciJYf7A0kWwuaVSmSrIdkOyMKygWTgYLYxtEdbCkbUUaF2S7JpDNJYuSVEE2Pj86z6ZjvhpjVBQpgiy0FUVOHOsPliaezSWNylR5tgOenYs0KZOOGd6scYFnISrSuDjbNeEsfGGSLEpSxdkuRdKxzLUQFUWKOAs9iyInmvUHSxPV4iENSaMyVartgGoXwiwpKI7lYV6Raq3rYSiJNC7Vdk1UC1+YJIuSVKm2S8l0zFSLUVGkSLXQVhQpUm3XRLW5pEGZcbU8Gz8xzLlcT8GoPPuE2i3fKENbVqZPTJ9/NhVnc8miIlW+HaeMOpFnwRClY15DW7kuQ1QUcWuz4yauxUMZkkVFqlw7TuFzwqQC0bGMESi+yhgpYW3suL8HcqfacRPe5pJGZap4O04RdcKXZYzKWCniLbSVsVLE23ET3uaSRmWqeDtOGXTC1yCM0sV7DVE9i0p0G5s66yqa6DaXLCpSpdu4yug8007oKrPCKF2h1hBVRdKeZay4VdpxE9XmkkVFqlQ7TsF0ItceiMoYKUIt9CyKnKDWn1ea4BYPKbMAa1yF2/iJYaxwjQWjMlaKcAttRZki3I6b4DaXNI4Zf7HBOCXaoFeglEp5AdPaGhdufyAqerhEO76IaPmBzC18Y5ItClFF2vEyXYGxZKa18HDUU7n2FJkW2oom7kKDcRPL5pIFSSZVlrVPnJcadCHw5Th2cD6pplxAgKhMsRBlSfrMdDFbE8zmkkVJqjA7SXF1KopAlO+Soa0qkrYVRVyYnTTBLB5K5u54UoXZ+IlhFHDdAKN8KwhRVaQEs31iOkaaIDaXLI6RKsROUhCd8kSCUfpFryGqiqQ9yxhxVxlMmuA1lywqUoXXSYqYU778QnQmS2JhGQHf8EBbUcSF10kTvOaSRUWq8DpJEXPGdXyMysxaLM1CW1HEhddJE7zmkkVFqvBq74YM88hMZlaIysxahFfoWRRxK7KTJmjFQ8nNrFVonaTgOZOZFaIysxahFXoWRdz1BZMmWMVDGZLFMVKtxE5S9JzJzApRmVmL2Ao9iyIutk6aCrF4KMM3oiJVap2k4DmTmTWNzmVmLUIr9CyKuNA6aYJW+MIkWVBkWoXW+Ikzj8x5ZsUoz6wQlasvRFmRPjF9x6CJWXPJoiJVZp2mZCnPujDKMytEVZESs8amzlsXTcyaSxYVqTLrNCXLOc+sGOWZFaKqSIlZY1NHkSZmzSWLilSZdZqS5ZxnVojyop01RFWRErPGpvmS9LSJXXNJozJVdrV3Y5P5hE77FUSnMlaK7AptZT45satbVJs2MWwuaVSmyrDTlEPnfO2B6JSvxhDVMVMqwMamhTHT9lpX+sVJ0qhMlWWnKY8u+JEXROV+B6KqDPRML4DGpgVlmpg2lzQqU2XaacqlC7k6A9PK1bnItNgzK1MsxE6b2Ba+OFl3hMpU2XYKSwH4oSBEZ3KVLrIt9szKnNjWn2eaGDeXNCpTZdxpyqkLeWUUCFhm4CLjYs+sTHGxwbSJdeGLZ5n7oVmVdeMnzqwrC1MgOuMZGKIyz2DPpEyfWGRdd8zMmpg3lzSMmVmVeeMnBmWY8CAq90UQVWVKzNsnllemiX1zSaMyVfadAaHyY1OIznkGhqgqU2Lf2DR/bZo1MXAuaVSmysAzIFVmYIjO5X3s4uID7JnPpiIDz5oYOJc0KlNl4BksEWAGhuicZ2CI6pgpLT6ITQtjpomBc0mjMlUGngGpMgNDlN/BXENUlSkxcGyqd5KzJvbNJYuKVNl3lhLqkgkPozK/FOu40Jbvl2LQUaRtYwM4lCFZVKTKvLOUW5fMLxjlGh1EdYwATfO8UmTeWRPz5pJGZarMO0u5dcmLmjAqY6XIvNBWxkqReWdNzJtLGpVB5j29HDVLliCMpvJqv4UHplnyAh6M0iz8FqKih1vNnV1EuLwA6zaXLQgx998Ssx/HhQe2x539Y/tL4MhexVZnuOOVGmuM0rG+hSgL0aejO4JcBLTU520uWdQBgfY0IOYh1UFKtxZOdOCVohgVHUoYG5vqZDq/CF/5bblcsqgD4utZh8OLYMfxIG+pz2FNLO+1tYbwUoQoUWts6ghxEa3KgqVsuigF8upZimRrrTCXpVtz4FDejWUNYZWiVKqNTR0pLsLTCe8olU0XpagC6hwKrSOunkBYZk6M8swJUZkwioQ6byLUXNIoTZVQ4+aDw94oI8YPCOtcmjbW8VJC1NhxHt7nTaiaSxqlqaLqHHBzxPwB4aVMrwCHMquU6rSx44I0TcyaSxqlqTLrHGqtI9mQC8NMrdBah02JWmPTgjZtu3PlskZxqtg6PyyfPVyOFjrbAJmO+IbYWidXbRk4UAYmpI9NnTm4CVghnZBki6JUq7TzA7z2ooQRv7u/snA6DfEiawjriMHGMgu7/DpvqtBm8wVZFtUSrX3ijLJhxHdwq9hBMjvzJANhkQWiLEqfmm7u1lSczWaLolSrs4sD1/Z7mHWjJV+uLZyIIptkQnjJ12uIiigu2C6a6rLZbFGUamF2AbsdSJ3AwqkovAIdwipKCXJjU51VFk0l2Wy2KEq1JrtIcNdGCl+iLZyKwjUCCKsoJdyNTR1Rmqqx2WxRlCrtLtKtDkLgi7OFU1GI+NYQVlGQpOUEcpfULppAN5svylIl3cUUXlfgRxsWTmXhCj6EVRakaJHFrccumiA3my/KUqXcRbrXQQj8XMPCqSxcVoKwyoIELbK4RdlFE+Bm80VZqoS7mMNoodRXFk5lYYiDsMqCoCmynBgXdupeNLFtNl+Upcq2C3g1LPATDQunsjDGQVhlQTAWWVy6XTTRbTZflKVKt4v07bDAG2auLJzKQke2hrDKUqbb2Ni5EjXRbTZfkGVZpVv7xPkNsdFC1oDFDga65cG0hrDIgo1HPFr65HSD3ia+zeaLslT5dpnybZCdRi2cyCK7r0JYZYHGKotLuMsmws3mi7JUCXeZlnHHXFFZWTiVhQkXwioL1oBltLiMu2xi3Gy+KEuVcZcJ446WslzQwqkszLgQVlmwHiyyuJS7bKLcbL4oS5Vylwnljmcy5Vo4lYUpF8IqS5lyY2OdcpdNlJvNF2WpUu4ypdwFV9RXFk5lYcqFsBQUsLHOLS7lLpsoN5svylKl3GVKuQvZ/MDCqSxMuRBWWcqUGxs7o6WJcrP5oixVyl0CifKeqysKy2iBlQtcqsTGOlqKldxlE+1m80Z5qrS7BCKV7VkpLKOmWMnFxipPcRHCsol6s3mjPFXqXabUu5jwzYCF05OKqRfCelKVqTc2dk6qNgOGXL4gSxhVsTd+ZDBhGPHeo2bCAODLm7qaDQOsYOATi5rL0Dlk6FgxNHox5JImear4a3XudGA48mCcSY/a8zvfFHbkcRnYmrVZVeSSJnmqGGxPR0AeLn9znNGG4ipPGYX75npq2Y/b5MGDGpImeao4HEape9hI9i2N8WTake05MS7oR82d0eMysTVrkyeXNMlTxeIwAnaVnTo5zpUriuvoKaNx39wbPW3ODXRQmTciwqiKx/Ej6dQs+/LHLoaaBO9daVNzcckDhvl26pCfMzG3eThkU6axU4Vke7wL4nCJPMZTcfiKjnHn1CqDct/cGzttdg7ZpEmeKiyHUVITHo9kX9MYT+Xhp7UYd+Qpl4X75p48bdYO2aRJniosh1FSGzZ5uGQe44k8vNWnnVpFXsawnlpuddgatU3LuZRJnCoqhxFUiGWHxhhPxVHmwRIy+X5Sc+eq5fKyNWuTJ5c0yvMLXMtS27IuTGRaRicwXte4DhXnMgjL2MlYlzV6l2VTJnHqvGwOZclGa7KXZUAHM94X1cTBBRM8djCsYyfGnXmn1ccMkkqSJnnqvGz2X6k8YqKCdmV86pk8uHRC5Cktjugbe+K00XI2ZRKnTstmWzaslZ+J0bBt2wfzDtcwMC536dTcGTs+LTeam2WTJnnqtGw+Y8n6KzGcCehhxqeejZ0Uhx15KrScsTlr9DnLJk3y1GnZ/MHSNy3kVhQdy7jabPKUaRmbO6PHrShbr20OioD4SdIkT52Xj75nh5Weuto+oLUZV51NnuIiYWruyONWlq1ZmzxA6UnSJE+dl802LD25+D0dW2wCc4/wMsSdk6vCy7G5NzW38XI2aZKnzstmS5bKw+9CBnQ8k412Me7IU15O0Tf35Gkj5mzSJE+dmM1HLJVH5x6AT9lzN4APmSNPucDcN/fkaSNmSCokSaM8ZI/mGrjCymHZa9YWoaQnl+y/i3GVB5vr3NNnqKWMrq3GnE2a5Kkzs3mMJaOHvUlXtuoC5JEiKsQdecrLLPrundHT5puWT5rkqTPz0TvtuBxfq4RojyZ7egWIO/JUasyxuSdPow0wfGuSNMlTp2YzHUtGj7wVFcCULLD13BrjjjzlZRd9c0+ethpzNmmSp07NRzO108scYuSJfmmy71mAuCNPhZpjc0+ethpzNmmSp07N5kKWjh7hHnApC7wFtI2eMjVjc2dq9qm5a6PmbNIkT52aj+5qB2pe8q/fpmYAUNkbDuPO6KlUmWP33uhpo+Zs0iRPnZrNlix9JUiK8GBbFni7bBs9xVUZGLb6KL1D1sc9edqoOZs0yVOn5qPd2nHu4fxt9AD2yv55GHdGT4WaY/eePG3UnE2a5EFqPr30HcycLB0yUnkH87LA+4nbkElZ2NGkgsqxuafJZTtEsFt7NmnUpO65FuwjQxVjKa9OxXhSe2dv0DXGVZ6i81rf2BFn3AbK2ZRJnDoom2lZCspyF4oea7zzuomTkrAjThqWySY29sRpW4qRTZnEqWOy+ZalV3J+vyyAr1mQjRcx7ohTLC2fDdngPRDrs6m8k02ZxKlDslmXpaeVjhygXN6h3kZOGnfEScM6cvzCcpsZG6YUkpRJnDoim4tZKo48lEDTtWSPx/ur3etrE6dcWIawiuMD8rgNkLMpkzh1QDZDs3TOUXGgQss7+Zs4ZUCGsIrj43GbRRumFJKUSZw6HpunWTrnCP+B51mQjTsDxJ3TKsVjFceH43EbHGdTJnHqcGy2Zinp6IQMNWF2PLCRU4bjol1b39i7WrWhMXxnSFImcepofPRtO91Y8XrcgNZssrEpxp2Rk6KxjhwfjC9zb6Nb5dt8yiROvZx8NHA7bhIl7iABTNECO0PYyCkzMjbX26oY98ZOWzk5mzTKU/dyC0czt+PYmcstBPq1ydavsf3A0Dp2ioZufWNHnEkbI2dTJnHqjGxeaHHWOWxIYSthacO5gNZtvBHbGuOOOEVGjp174rQxcjZlEqfOyGaLNpDOXLYlNP/I9O5KtsbFuCNOkZF9nzfrs4mRsymTOHVGNoe0dOQII6OhG3uN2MgpMzKEZUqOUW/ktJWRsymTOHVGNrO05GI+1tMK6sCydXAAszVn5KTNVRyfkdvs3zClkKRM4tQZ2XzThpEz4jrWytbIwWnFr9pj3BEnba7i+Izc5gSXT5nEqTOyWagNc85MrEcDWKwFvprZaVVeeFG0hOsbe6dVGyNnUyZx6oxsbmrnkWPlGbmBALe1wG/lmzhlRi66w/WNPXHaGDmbMolTZ2QzVkvvy3XOgfovv21t4pQXKReN4vrGnjhtxWP4zpCkTOLUGdk81tI5R69WUAhmfxYTp8zIRc+4vrEnThshw3eGJGUUp24cF+wj6a2nPPREd7gFv4gf25cIuWgf1zd2xJm2EXI2ZRKnTsjmvJaOHBUH1kuwS8vaVsKUxSkSsm8lZ302QSB62yUpkzh1QjYTtrM4tlWBnFboGcdGLSZOisB6KYewXMp9Vznrs00cgPokZRKnTsjmw5aKIxMy2sctZIEyxB1xilXks8EcltjbnOVCNmUSp07IZsWWXK3EeC+AVVtYyMt8EHfEKRJybOzNOW1V5GzKJE6dkM3pLb19kFogmsgthJAh7ohTJOSzzRyNnLZFFtmUSZw6IaNPnOyJHiiu4pQJGS3seA1B2WnOvrtt7oEFHsnvlUSqkzJ4s4WFFEwxLg4ttnQnvQeTd2UxrDXBGPdOsDZWziZN8tRZ2Qzahtkn6C0oGLgFNnOx61aZlYvGc31jT5w2Vs6mTOLUWdk82gZxxlr2Ag+3wJvYmjhlVsbmztjx68ltvnOYVJo0ylO3nrPRAvdZ8kIf+svxmox1bF+iZWyu8pw96HB6bjOfw6RCkjTJU+dlc2s7U8/IWZKCJnO8sa3JA7zMmxxjWJDwbENH4rTxcjZlEqfOy2bYlt5MyONh9JkTw4YAcX3PGi3u+Np1dqIjcdp4OZsyiVPnZfNsG8SxbSf4WQR4ugU+ehs5sCqDLR3eYlyHjl9SbrOh45xy2zv0hm+f7j7aZSG8NAj48PDu3eaj/Ycp0i+b+NKW4KTqyD5fMZ4s9WLoM3UwTkPP1KksTT570tHoaWPmbNY0eurMbC5uyeiRLZtsyQvoI69kUVz1gfbOvOwXltus6fJZkz51bAYTOXsYKmcXsKe4fgRoH5yzq7I4+exT14/S/c5+W+/f/Ombb7/97R9//882vP9OciYd6mRsDm7JOOF3z1a2wCUZJ50YCWHc06Hy7l7s36G/Nlu6fNakTx2Nzcct1UdKOuDz1vETY5tnoMzsjJPKMuTY3tOnjY6zWZM+dTpOveo6KzbJeZTSb8dXMdMHKs2OPpUlybG9p09bMRmySrNGfci2zrtOHS3sjisK9O09sIXr2HtoHSDunF8Y13nYt7KzbpvuzLNZkz51Pj5a2530kdoO2MV1bEBk+kC9WccPxh19/EUXl3ncyWqmbNakDyGyDWTDrvvPz/vt0+1293RnG1QlzHN0oQMu5FfSVuH8qYF/OjEf8j7lja7D63uVt9h8dzz7ikvG2Jj3EbqtHQlpWSdqcJnrRnK7QXEpsqJLnTPWajqdmNq/5s/blmtksyed6sVoc5xLr3nyXBkc6To+p+ycrLA1ufTJvUeMO3N6m40eZpVmTfrU2dps51J9dBylbGxuNXhNNH2AnaWaCGEVxwfrNiM9TClNmcSpg7UZz53FsfeN5LYVjOk6MXWypTTFUiuEVRx/VfO8rQ6dTZnEqdO2udoNI2ciW7wHMMzr2CXMRk65Dg1hFcdH7XlbFTqbMomDqH2+jz966B2u/2Pdag9s6bogD9vRtk7PpeJSZt9Hz5YRXXLhou+0y1b6nWnKpAjC9aAIbCDHr33a9R2ImpdE2RgBYlZF0rCOER+nL/PQY28hkySXM0pCNnpnSY7mef0gmcx5G+WVLYVJ5oyx3qSiX51IUvTP6zt3rkaLiwiac77lnDOu8raQxa6F5VIYGM/Znnp8C0ZxebBOxnW8LVrRSO+Qn24P0mal1/d63pzVbivPOdF4qZeYwYCu0zVyFJfHo2RgJ+IUFy37jnqhzVLPmqclmeRkI3HqQAxGdF2QdwEoLiBDRnYiTnFJhu+sFxZtFJxNmcRBCh5mmnQdRuBFKDbTpGjbiWUPxnWpATSXyTdGvZnmoppyR+e3zTS5nEmSOviCFV2n65soznbApmfxWR+EVZ0T+Pp3T20me5ha+oslkRCAh3ED7/LpXuXgVtfxLkVrW9FTVqZSTY7NvZFzEffOebjb0Em/N02bVEHyHVRJXuIb2bpPuTJBcbmTvZnQXY9GxNsAYR0vPu5eZq/Hj9VMklzOJEmGd4++eofXr2ydsEgC8NixE0tAZz2VpMi7sbE3TC7iXfotmiK5lEmRejH56K53FEfXMoFdXccznZ1F5aUWENbx4qPvoq2SnE0ZxSGLvfMZdDTWOyrCb+OtAjjVdezrs8a4Xo+guSjiW+tZnxfdHjFW3OaTJk3q6Hs02Ovl6WwHDj6bwLGuY18fk6e8GhnCKo9fOm7z2MOU0pRJnDr6gs2eVXhEHOBIdmo3ccqrkSGs4hwqx3GBR/pcs81pD1PqkpRJnDr6Hs32DifWjF+RthMrZddOnXvQbk/mYQirOP7iija/vXzKJE4GfY8+e8ephjeqMEUAI/lcs+GSxp2pprgE2Xfasz4vmWrkypRNmRRB8h0m33RTZGevevCs68aylX/ZZC9AWMeIX+e9zGWPf4s290JtOsmZJKkXeo9Oe8fxoisHwLqu45c+bbyUkRfCqo4PvMu2Qm82ZRIng7vmRjeUvp0VW+BW1/FNgilSfh8PwqqIj7uX+evZnkN4lbABA7ybJE2aZHjXLOiG1bMzXlltswrAo5o4QdyZVYq8Gxs7vHuZqR5vdm6S5HImSerAezTWO+xxN9W1xWCd1/HurTZiysALYR0xPvAu24A3mzKIY0dcq2vGjySvfcgluu9iKBKO+XEtxmXoYJjFOeQndU37ccu6iXzKJE6VfG0lFiyd5VvrGE+e/4u9Fcb5QdNbDKs4LvlaozZxcimTOFXytWVYqTjOyAHy5YUH69g+WTzKzwswrOK45GuN2sTJpUziVMnXVjPByNCRg3E6PhMHVjmoOKWib9+4n5HdIp6F20TKpU4iVZdA2EIsGEG87C/Gk9NLLK4w7pxeJRjuG+tly37cJk4uZRKnWga2VVggDq/5i/FUHOZijDvipM319HK52PpsEyeXMolTxWJbagUHr+JgXMUprn/A7lWcExZnTq8mPOZDG1InkarrIGztFYwgfmQZ4+kI4sIwxp0RlDZXkVxStj7bRlAuZRIHOdlZMmpLk0AcrtvEeCoOl4gx7ohTQua+sTf3NK03zqdM4lSJ2ZYmgTi8Ni3GU3H4YTfGHXFKqyP6xp44TcScTxnFqZvr2WoqOHgRh+IiTtlcD7uX0+psrufPPW0me3xomRUBtmSqelsBJnsjedW+72K4rRCrK4zrCAKTPRXJJ+c2i718yjSC6uQMFnsj2R3QVjPBCBM4LFvsYXMVxyfn0EbO2ZRJnDo5Hy32DmvVRryf0soeQIA4vFwC487IKZJz7NyZe9oM9vIpkzh1Yj4a7B3FkaWNti6pfElHLzu5rYCwjhx35YR9Z9MlPZsyiVMnZvC/66RWaiucUnHE4grjzsgpEnPs3Bs5bcScTZnEqRMzmON18gqRLdoCcQQG0XxPR07aXEeOW0i272wbObmUSZw6KYMLXSe+lR3F9ZJeXDGMzVUcn5RDGylnUyZx6qQMHnQd2w/ZhAykLP47GHdOqyIpx86906qNlLMpkzh1UgYHuo7ddUwcIGWxl8G4I06RlGPnnjhtpJxNGcWpm+rZAq90TpEdTyiuEzL618mcA2E5rc6Wej4pt1nr5VMnkeqkDNZ4HbvJrGwBYHliRus9FSltriKdSDkjUlutOZs6iVQnZjDI69hTxkQCYhYHOYzraQbNVaQTMWdEaiPnbOokUp2cwZGuY28ZEwnIWSdqdLTTkVQk59i4UHPu2mrO2dRJpDpBgy9dxx4zJhIQtE7Y6GunIhVrzmervcxIaiPpbOokUp2kwZ2uY68ZEwlIWuxWMO6cbkWSjp07V7U2w718yiROnaTBm64T6wNbHwhXPbl7R287HUFFkvbt9uw7m0g6mzKJUydpcKbr2DHERg4UcMU0A+POyCnWnGPn3shpI+lsyiROnaTBl65jUwwTB0haXCEw7ohTJGnfas/6bBs5uZRJnDpJg0Fdx64PJg6QtDPnpHFHnCJJnz33MhNzG1FnU0eR6t57HRjVdez+sKK4uBlhXEUqeu/1jQuX+DYPvnzqJFKdqMGwrpMlubaGMJ2gdQ5CwzuZoCEssHj24PNH0riNqLOpk0h1ogbjuo7dIGwkAVHrXITGdypS2lxFKhN1mydfPnUSqU7UR0++U7lVHhKC517HTznWtvowGWmydR6GVSS/Ft3myYffmaZM4tRJ+ujJd7LxkedgYHDXiWtEh558vOkihlUcvxbd5smH35mmTOL4S5lt6WC6ZYVsfRLjyWNTsYrAuDNcitgcO3fg5zIjPn477DafM0lS52awrev02QXGnZMJljLreClyc+y8dO1q4+ds6iRSnZ/Bvq5jTwiblpGf5eYCHflUpCI/x8Ylkdo4Ops6iVTnaHDe69gbwkRCKJVHhBB3zrMiR8fGJZHaeDqbOomEPH16maI72vL1c7I5NsiyQ3C468Q4IrZPZij2TniLcbU77+PeJHQRP49lzxabhjCx4fEU6lK35uvAx66Tl885Lg/CqD0tDHqL7R2J+hTjmmcfENtc+vLZk05I0efxQ9Z8NL2sOvC568QhgeMqTrEa7XvzWactO8XkcyZJkJkHSQCE2ffAJMG4LJCnuEpSJGXfkc++tE2SXM4kSZ2Qwb6uk5eJO4qrOlCT1gmnaMnX916YiydtReds7qRSHZXBWq8TZ4SO4qoSFKUdlYpV59h7SaW2qnM2d1Ipw8ypMZ+95M7bvHVgzNeJaQTGeQNljHqTsQ/NlznzederbNqkCmLzMOmk+7stZEPTDuztOvGKwLiqUkTl2LdzDZ9chMhUELjNZ0yC1BE5NeSzFYvCOOBu14lVREdxusDbBbzIyDHqqdPGxtmcSZ06Gx8d+Q4EOJrIXTnY23VihdBRXNUpwnFs7anTBsXZnEmdepE5teQLI152ahdzwEzxQuC4qlOsMsfePXXaqsvZnFGduiefLZlL6xfyomSMJ3cHYobAcVGnaMrXt3bUaTPly+dM6tTLymDKF7h4vurA4a7jjWvXHFd1imzsu/JZp02Pb7I5kzr1enLqymdLpeWNHLDV68QqwlYkwthSdYqY7NvyWadt6mBOw90QqVPH5NSWz97TkFkZPO46MULoKK7qFNdm+L581mmbOoDuSc6kTh2Pj758pzK7MCCY3I3ZRsLOLMBj3g33LcallOwb81mjNnXSnNKcD+rcPH/YbPZv7/Z3b17dfd5vbx8e95vd1W7z/vX1l+HlbZ/zz7uXnx/evb7+6+j4zxf2Z4j/Gg3/OsX+FjeB6XtZbx8/P328ut8+fm2N4zuWh58/n/5y9ePd4+vr0Yul+WPc4A+D/OCFbQUqP7QFlNzQHgjxj+b8E1vGQ5+xxR78E/k227ScPxSRm382kR/I0Rlm0bepArZBG39Ivl5StD0trOObs87Hvx1+ExYYfsVvXn26+2Hzu7vdDw8fn68eN+/tLZjRC4PK3cMPH05/328/9T+1u5Pvt3vbp/70Xx82d+82u/hfdvK9325tzBz+w7499vvdZv/509Wnu0+b3XcP/7d5fW3Esd09bD7u7/YPWzP3+bTd7Xd3D/s+3Z+2uz/34/DN/wMAAP//AwBQSwMEFAAGAAgAAAAhALhvUnTXAQAA9AMAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWyck99v2yAQx98n7X9AvCfYadMtUZyqalW1D5OmdVufCT7bKMAxID+6af97z07tRspLVGTD8evD9+Bucb23hm0hRI2u4Pk44wycwlK7uuC/ft6PvnIWk3SlNOig4C8Q+fXy86fFDsM6NgCJEcHFgjcp+bkQUTVgZRyjB0czFQYrE3VDLaIPIMtukzVikmVXwkrt+IEwD+cwsKq0gjtUGwsuHSABjEykPzbax55m1Tk4K8N640cKrSfEShudXjooZ1bNH2uHQa4M+b3PL6Vi+0DfhP6L/phu/OQkq1XAiFUaE1kcNJ+6PxMzIdVAOvX/LEx+KQJsdfuA76jJxyTl04E1eYddfBB2NcDa6wrzjS4L/i97KyNq87bKRlneVkflP18uSk0v3HrFAlQFv8m5WC664PmtYRePbNbG4gpx3U480hkZbY9gQLVRwSQ1W7gFYwr+QErinw5IJgHFQDy2e/p9F73fAyuhkhuTbtE86zI1BZ/xfuwH7h5A102i9JkOGu9kkoT3soZvMtTaRWagojXZ+Atn4bC+sxP6bnTK2QpTQtv3GkoWoKDIxiS6Qkx95437BGnjmZcewpP+SzFKmjBoSosuGwruMaQgder8HPJ1+QoAAP//AwBQSwMEFAAGAAgAAAAhAJeBwNXAAQAAwgMAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0My54bWyck01v2zAMhu8D9h8E3WPZadMtQZyiWFGshwHD2q1nxaZtIZYoUHKTbth/H+3U7oBcghr+ICX6ISm9Wl8fbCuegYJBl8ssSaUAV2BpXJ3Ln493s89ShKhdqVt0kMsXCPJ68/HDeo+0Cw1AFExwIZdNjH6lVCgasDok6MHxTIVkdWSXahU8gS6Hn2yr5ml6paw2Th4JKzqHgVVlCrjForPg4hFC0OrI9YfG+DDSbHEOzmradX5WoPWM2JrWxJcBKoUtVve1Q9Lblvs+ZJe6EAfie87PxZhmGD/JZE1BGLCKCZPVsebT9pdqqXQxkU77PwuTXSqCZ9Nv4Btq/r6SssXEmr/BLt4Ju5pg/XLRqjNlLv+kr9eMv1n/Smcpa2Gwxrm/crMuDe9w35UgqHJ5k0m1WQ/i+WVgH/6zRa/FLeKun7jnHGkfqk5i7wYtfidRQqW7Nn7B9smUscnlUo5jP3D/FUzdRD4MiynjrY6akV7X8E1TbVwQLVQckyafpKBj/GBH9MPoQootxoh29BqWPvAWpwkvRoUYR+eV+wCx88JrD/RgfrPiuCYkwyIftJ1LjxRJmzj0Np2+zT8AAAD//wMAUEsDBBQABgAIAAAAIQCkj5JsgAYAAK4bAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxZT28TRxS/V+p3GO0dbCd2iCMcFDs2aSEQJYaK43g93h08u7OaGSf4huCIVKkqrbhUqnrpoWqLBFIrlX6ZhlJRKvEV+mZmbe/EY5KUSP1HFCX27G/e//fmvdmLl+4kDO0TISlPG0HlfDlAJA15n6ZRI7jR7ZxbDZBUOO1jxlPSCMZEBpfW33/vIl5TMUkIgv2pXMONIFYqWyuVZAjLWJ7nGUnh2YCLBCv4KqJSX+ADoJuw0lK5vFJKME0DlOIEyF4fDGhI0PMff3r51aNf7j6A32B9wqPNgFGqpF4ImdjTHIiz0WD7w4pGyLFsMYH2MWsEwK7PD7rkjgoQw1LBg0ZQNj9Baf1iCa/lm5hasLewr2N+8n35hv5wyfAUUW/KtNKp1i9sTukbAFPzuHa73WpXpvQMAIchaGplKdKsdlYrzQnNAsh+nKfdKtfKVRdfoL88J3O92WzW6rkslqgB2Y/VOfxqeaW6seTgDcjia3P4anOj1Vpx8AZk8Stz+M6F+krVxRtQzGg6nENrh3Y6OfUpZMDZlhe+CvDVcg6foSAaptGlWQx4qhbFWoJvc9EBgAYyrGiK1DgjAxxCMLdw0hMUawZ4jeDCE7sUyrklzQvJUNBMNYIPMwyJMaP3+tm3r589Qa+fPT689/Tw3g+H9+8f3vve0nI2buE0Km589fUnf3xxF/3+5MtXDz/z42UR/+t3D57//KkfCBk0k+jF549/e/r4xaOPX37z0APfELhXhHdpQiS6Rg7QLk9AN2MYV3LSE6fb0Y0xdXbgGGh7SLdV7ACvjTHz4ZrENd5NAcXDB7w8uu3IuheLkaIezlfixAFuc86aXHgNcEXzKli4O0ojP3MxKuJ2Md738W7h1HFte5RB1ZwEpWP7VkwcMXcYThWOSEoU0s/4kBCPdrcodey6TUPBJR8odIuiJqZek3Rpzwmk2aYtmoBfxj6dwdWObbZvoiZnPq03yb6LhITAzCN8lzDHjJfxSOHER7KLE1Y0+FWsYp+Qe2MRFnFtqcDTEWEctftESt+e6wL0LTj9CoZ65XX7NhsnLlIoOvTRvIo5LyI3+bAV4yTzykzTuIj9QA4hRDHa4coH3+Zuhujv4AecLnT3TUocdx9fCG7QyBFpFiD6yUh4fHmZcDcfx2yAiakyUNKdSp3Q9E1lm1Go2+/K9uQc24BDzJc8W0eK9SLcv7BEb+JRukMgK+aPqHcV+l2FDv7zFXpRLp99XZ6VYqjSs17bdN7JwsZ7QBnbU2NGrkrTe0s4gPodWDRDgZkMp4NYFsPHvM13cJHAZg8SXH1EVbwX4wz69ooZIyOZk44kyriEedEsm7mWHKFtRlQKrbuZNmt6DrGVQ2K1zft2ebk4b07JmOkzMjPthNGyJnBSZssX3o5ZxUq10GyuahUjmimKjmpTlcGH86rB4tSa0Nkg6IfAyisw/WvZYd7BjPS13e0sPnGLZn2mLpIx7pPcR1rveR9VjJMmsTIJI4+P9Ox4jI8K3Oqa7FtwO4mTiuyqC9hNvPc2XpoMzDMv6bw9ko4sLSYnS9FBI6jXlmoBCnHWCAYwKsPHJAOvS91MYhbBtVOohA37Y5PZhOvMm3V/WFbg9sPafU5hpw5kQqpNLGMbGuZRHgIsNYO9kX+pBmY9KwVspP8FKZZXIRj+NinAjq5ryWBAQlV0dmHF3HsYQF5K+UgRsRf3D1CPjcQuBvfrUAV9+lTCjYepCPoLXM9pa5tHbnHOk654KWZwdh2zLMZ5udUpOslkCzd5PJXBfLPSGvFAN6/sRrnTq2JS/oxUKYbx/0wVfZ7AFcRyX3sghEtigZHO10bAhYo5VKEspmFHwMWZqR0QLXDFC48hqOCq2vwXZF//tzlnaZi0hklS7dIICQrnkYoFITtQlkz0HUOskp9dliTLCZmIKogrMyt2j+wT1tU1cEWf7QGKIdRNNcnLgMEdjT/3e55BvUg3Of/Uzscm82nbA90d2BbL7j9hL1ItFP3CUVD3nn2mp5qWgzcc7Kc8am3FmtN4qXbiozaDiySk/8D5R0XIiAljfaB2+S7UVgTvL2x7hSCqz9nGA+kCactjDxonu2iDSZOyDUve3Z55GwW33HmnO+ULWfpXOt1TGnvanLnsnFx8c/d5OmPnFnZsXex0PaaGpD2aoro9mgwyxjHmhVnxZRbv3QZHb8JrgxFT0r4uuAPXhjBl2BcPkPzWuWbr+p8AAAD//wMAUEsDBBQABgAIAAAAIQDoDuWQTQQAACIPAAANAAAAeGwvc3R5bGVzLnhtbLxXy27jNhTdF+g/ENorelhybMPyII4jYIBpUSAp0C0tUTZRihQoOpWnmN3suuxHdNf1bPo3HaCf0UtKspSHYznTjhYSSZHnvs69JOdvqpyheyJLKnhkeReuhQhPREr5JrJ+vIvtiYVKhXmKmeAksvaktN4svv1mXqo9I7dbQhQCCF5G1lapYuY4ZbIlOS4vREE4/MmEzLGCrtw4ZSEJTku9KGeO77pjJ8eUWzXCLE+GgORY/rwr7ETkBVZ0TRlVe4NloTyZvd1wIfGagaqVF+AEVd5Y+qiSrRAz+kROThMpSpGpC8B1RJbRhDxVd+pMHZx0SID8OiQvdFz/ge2VfCVS4EhyT3X4rMU8E1yVKBE7riIrbAYW8/I9uscMwutZzmKeCCYkUhAlcJIZ4Tgn9YxrzOhaUjNti2UJ0a1XjgI9ZmLbTM0peFoPOlpqLft5UZSnpCJpZE309E7Y5z9/+/uv34/Jeh7WfQhxJSlmeijDOWX7Wll/kFYnHfAE8gzrH6nZc+tZqMYHJVhHGTvENdBxhYHFHBJAEclj6KCmfbcvIKoccrV2gpl3YvZG4r3nh8MXlILRVGuxue5zyRi9bsYOQR8b5jg9XTVlhuj1WIzcrCMrNo97tiwjEly5FjKFctcmiQ9m1EOLOSOZAipJutnqrxIFvNdCKSgJi3lK8UZwzDTj2xX9lVAmoSJGltpCRWuzDO+UaJLM0fAN+sm5RgejwsmpoGar5cm5tTHP29IYBS5KCGO32pifsoOfJmBSlSG+y+NcvYVcho1C533bhHg2zdondWcxh3qy4TnhUEeIVDTRZSiBLqlLR5Udh/VAwv8AOzoKC5E9ZuLXW+R/NfW+0L9OnyY1aXp80UXqfL6gKvsPiKOp+TxxDvgIFwXbX7Xk1Cn6IlfRLxIXd6SCbdXsmM5LxA06+RDNLh/A4bV90DDyY0ggDdj0QOeutzR1qusf1XUrJH0PQGdkVs9BQOx+wj5ykN5ZTvpGF0TtDkMIoECvgDwoHwe+IH0EiKzv9cmQtT4BndY7yhTlNQP0QaWtQw8WIB+BV2tHgvaVnO0o1KNf3eax4Rvql9u9RqP65wdz2DgGO+qAtYRBwHCI1c9J4A4a2DEMOjgNPULaAyZnwqGw4WlYraIB1Ywd5AZvACg6wF4Ohb18Afbzp0///PERTTvPjofCjnuwXRkD4qZVt+OZE4bSVwmzFx6oDERNSYZ3TN0dfkZW1/6OpHSXA4GaWT/Qe6EMRGR17Xf6kOGNNR+hqrwr4WQAX7STFKh8s7ycrm5i3564y4kdjEhoT8Plyg6D6+VqFU9d373+0LvQfMF1xty/oFR7waxkcOmRjbGN8rfdWGT1OrX65sQIavd1n/pj9yr0XDseuZ4djPHEnoxHoR2Hnr8aB8ubMA57uoevvPa4jufVFyitfDhTNCeM8jZWbYT6oxAk6L5ghNNGwukut4t/AQAA//8DAFBLAwQUAAYACAAAACEAPEb5KqgOAAC8NgAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1svFvNb9vIFb8HyP/woEMho7ZE6sOS08SFLMmxN1asleQN2tuIHEuMKFLlx8burb0URbHd0y7QS7HXXortpZe2KNC/pdm0/0V/MyQthzOULTtosJtYnnkf8+bN+9bzn14tXfqSB6Hjey9KZsUoEfcs33a82YvSxeR4r12iMGKezVzf4y9K1zws/fTw6ZPnYRgRYL3wRWkeRatn1WpozfmShRV/xT2sXPrBkkX4GMyq4SrgzA7nnEdLt1ozjP3qkjleiSw/9iLQxZ8SxZ7zi5h3k1/VjXrp8HnoHD6PDj/89k8/fPuHD3//9sN3v3r6BFh54DAXwDZ/Xo0On1fFtttbf/j66//88/unT3o8tAJnFeFw+Y3vf//r93/+a/XDX/72/o+/e/pkwLz4kllRHPCgGl2vFMTvv/rm3//46umTC8+J8rh++Ob7//7m66dPPo+ZFznRtbKesjOCfIJFmF8ecm/GlsxjHv2ILVc/oaG4D7bC5/IrHvI5wwWQjf/FJxdbxedXPHjLdvK4jk/2jk8UAt38byYB88KVH0QsdHZp4E8d1wnxc8qAzZc3v8qDnnrQB5eOzyH/gJOQFY0H9LKy36xVenud3nhMtYZcVE4K7sYXR93z1zTuj7447fbzyF8N8r9pGvhTM5rGnvkQTur7ek4ytObD0Dbam9HWHoa2VduAtt4wHyiEg41CaB88EK3ZKLjmVLjtdqEU7lCf9mbxtg8KEWcXQxbej4W3SCvf5bQURoXbdNybFClYvRBndivb42wU4syuZHuczUKcN/exNdL9Yn2tZXexNdJWsUSbrVQjt0ba1ojUFCbCMA80b3m8cp0I7oLMZ438xSdg7dq+KtBbYG0qD3w7dlmg2No13VoxbmUJ3g/MGg3NPQ4cK/BpE88ZsIZnBbidZ2pNuZCp/T1lKbX3w9vPyI8j2/eDxPiXzX08siAK5ccdEjsLTPXLAO/QvnmNm9AUYLjwbB7MEjRDbnPpi8bDk4OaQWNn5jHhwWmMAIGHNBG+yfEsN7Y5McviYejLhcTNWr5nBTziNGWhwrEUlmG2NS/j2JlCo84RVlj0me94EXVdPxR0JcFTb891PC6cYBcesugWNMbxnojh0woRG2ZTYxzviRjGcwPi1sMRw9ptQNzWaPM9ORYmb5MsNI/sTszJ5cHsbWJZ4zLvQNzzl3fqxPYSllg3K8S+xmTeh9c7tGH7S5O8blaF5vavQmK9Sw8eyKxwfpu0oKaay8zItguW5APV21/DxBMrWsIjKVpqNouhjGIOi2lBYwo53MDGhiNroOrSaxsHmhsfOrOIOa5IK8bdameopC+pkA3Ne+n6nsetCM5JD7omW+QEgVVZEj7ZQk6cN+abs5kzNuUuddkUEWjZ92QouguHRCdIqOdwk0pIkeIbdCb90WnnTPqpglQpQa4JaNOFzoTKme/7fAQ1tlVyyVadw5YhvNHUZRyJZ6zpsoZ0SRf5pUu6nCDlIvXmeRGPHJQnCIFRGOICaHpNyIEvKYxXIujI71byRxWcRRGz5hT5xK+cMJJIA8eeKZ5aBW0tSTCB65RZxZx5s23ADx4NbvuxUKZ7UB/yhSgQoFAwCVDRmQs+zaZhLcnmq2iulAg02+vGXtugDILKNqQVONNYlFOgxaGDqAqxVSjia99zrxVtPkK0BbIobcSoGGyg/vHGGo3jKfViK7o/TP0BMMMvutWT3rBPjdI9CNlOALuCo+Mob5kLwQbsmuXlmJ5E3Rz5bsHeADYCd8WjImQDDqUrABZ1KzDUMORFlT9jbhzQq86pcheyyGQLfRCFpJWo1EF7xY/OiiXCXjGb0Vu+nCb1JlGV8j56bIS4F8W9WSw3LPHTLPbYAruGAskUfNJL5n5JpsS6i994KFXNaJ/2yKzREsF2oGoe6Py/GKttwVhSxWg3Wmp+mC3pfJcQm+44UkRSqe1Ustd4OUuCkKNYPE6KHCErIcW15MKqjcvdUnJ3kaoVkkJKq6clT4yarCbH/lizhmvNMsWbtMUzbhrVBkQv9Aumy5eKhRwOqZcQiLYiA1qa1Eto31q6D6C1S4kKFNs/IIWahw6Me8LxmlG6Fhe08EOcgKo08m1RJ08eTz/1JZWi02ii8LWNRsV3yhzxkkY3YgEFYZ1209ekIF5DT/2pvyDwItkZnAjIE8JbwyOPF3jsGfKFeNyuU/wEZ1gFE7AAMcx26FzCMqH0L6rTZO42E93YhVy8t0IUS8dzlvGSsLB7m152SaLcLSS08kPlltfsp0QT9QDT8UKYqcTGsnAF85a+GJdPWYAHI3lEiRC+DBvXHOKhWB8zAosDXGBYWK0gFgatPGXW4tJxkebNdnCgtRG7qQfImvarWq2Z2kocR7DBFCN8+wI94JsJ4U3iKF4RDLOI8KpZqJeTueidPANSCz0VNEdCHnzJS4dpreXG/SUaQPVatdamdyzwGM2diOGMdXRl0NlAz2QWU0GtZOxbCx5JJUpQFO3LvG1KrmFU642U3AXwC2pjSW0I36+EDhnTt6hJDPc6sXAXzHN+iZjGQbEGMQaTXgJKj+B1vpM/m15uhVhqnwRL/ZNgaWyDBbEJdd48Sih6FFtJRI9iK3HoURTI4jDJlYRT+Nd3NOmqiVh+Qz2v0/kNStk3v6F5F4b9jRtMg65g/6zl3buwVbPN9+zYiZKbrlcTweRxHbtIVUTgX7RhPMfrQ46X4IL/L97ahTVfEbqzJK4mA0nTqWICEReprITFGyWUnJYwQ+8cO5oTDKX4EM0da+GhyAp7sVzSj8m+9tjUd1XTD8K9JJPpu1P/HZmVWrOkeCV1l7optXHl3l5vvJMajTyiMxbMOI2Y7cSoCXNYmd7e2WQHkcBVAUTClUB6VoS0568KgCcMTlY4PeaJkkEVYYLwgQH64soRkVIwNNKFkyDhLuS9IKenZ4TU6yr5y6zrFCery29AcScO3H5SAVcYuDf9x6AY6nlH5eKZCHqvkr/2tz/9fTBoGE+kD+DSvRl4JBa9BFaxjE5gEyIwUjOu5H8a47FRB+6NpeAKt+XisWhE4J1klgheh6cTQjJyJf5v0IBuEls1dr3JZCdxwKYi8hXAIjI8YgFianVIJCk4nShzEolrCBxEY7q2UZ50WoWsayL7rHukSZfeIIwT9iBiIt6ZxGLwAtMcnaP+GQ1+RqP+8OLo7LSrJPCDW7E0lRHTy6RfBkvK3uQgGPK4vjmHqFUBRhiaW50wvdM6Yza6bCgcbQ1/9FEmIpjEPIyzyJNJtlkcUyfCUBZtG6LYhR7byJnNI/Iv6Q34KY/O3+xUkS8t8Tg6chBIUYn+aHCq9PlfccT6cnII5YoEfA8a4jKHjniIvKKcfHiNtEcUedQKykUKl6eXVAPMhqYafQZ5u35gF0CiUFOw0vWXS8xDYX4h4VVVe1yn67uMJr57556zkSKO0aQ6erP5QOgrKDM3eKSQOao6c+oxHuCfIv7Wz1KWNVpJ7kjN0i6NJ9D1sz3DqJN5cKDEV7fS+omEnSDLQdZHB4/AkdFvfAL6D8GR0U+CZ60pMSALVeITWY8QAYV49IvEQialopsMPxTvaBeP25ojrhOmjy1CnjS7FT1ObNOBrp2biDv0F6LAUU6IHCz1GGCZNZUxHYaWiiExUK6PIAphD4xNdeWyKMlpq8liYrD0gjLbhVMwYogNqbGwc+9Qsa1K0UBvwhCZPgyNiFvzSM0KDUEegwDcjeBv1R21CmpZIaoIwjag6I0QOo+kji0ySJ6LUwn6QTwVgwLC0Oc3Nyp0FLu2P6MZsnjFoFTSSBvDJ0nI/TIpEio3sV+BdRTuSmbPqNAoUXSrIkuyMse2IRkRxsvNNWVrOzuB9nwHEJKIUOVi2gZRBGl8tAnTbLL3hV4BRgYhYc0cpmlW5BFD6mFEFMOnGNLCyA70OawmpBAUqHV2EzdyxESNboikR1y61FaFIdzKSR/utHNGvX6nR/3XPeqddgZ9NNhQ9K9AhVsVgwZK38jEFfWvIrgMceOdQFGarOwhhmOFFxMzIJgwG6G6BKtWgvYhM3rpMqRDQ+bxfPV/cwEoQ3oOlVuj1Dseo1U8v/cGYsd4J4bfPuJvLCd64ZZ0jIpK/QWVG02Ugpe4FXSLFL1LDy9nYGCYe7c7Q8cBSjjlpPu6g8J/vd1I54PK571jBVU6LNgqHkG7FxUxUpZMIWmpbMNwC7NEnwhVe/9eZ9eY/W0YxqTGJ+LX3P9kqMTsxAau0uG5tsaHQP/kaHE2wqlaGTlCp2slp4PHN2OFRaCaUC2dHL0DtF3TPLZ0QPROUE02kM6B3slwSy2FZRLUjF1kS5pJhmzqUTMYki2ZRdMfxr4qN70dG7LImuMS7Ww0vNVBV1fOcsAi4IeLYVf0BGFilOp1mlbVVBf/KFqoSRWSUjO4B5LCrMrGYxm1pqoEj6JVeCx8yaJowscEG4VzQbXC4R/BuzIZM8EIpoVkhc4nvVH69YHz4YDOzyaQBAx+/gkWAvSOAaCZZykCEGUyAHSKJrp1MWKaqhkag5staQzSOEZ35jo9HL5TgjlTKndChKyeOsGj/2KDxG42TY3h6YSIBR0Enr2AvRORRMeksonGe0QM0SvqqGj24IPlr67xIfQvk593dunV4Oe7dHT+ecaab8UiydUGWtn5ND3r9ZJyveslpdI+DPy3mIgQnS42k8l17qr1eo36B5/7cciV14gVxO1IeZPx3AG+IYT4C0kNhx69lmWCa86UAYLX53QxVr5JkgUNQ/8dwuOBmDwgs24aHgrVSNWTf5tN/JtXUKnI5Qx+4mBsEzOGYIZG/NIVs2X6ostxLJMDOSqmjjkcTyYndOZbiJZQV0HfVBkm04tLwmUQ1OVB5Fw6iDMxByWXxNelLMzK5U/xGUNzVoTH+MLQAhUm0VJdd5mPeBBjYuOIeQukB+LLQEpTeI1gyV10wwOB4R5wt7J3mS9UKckJL2ygoBZqebWK0nMZ9gedcef1y85rGp6f9dHAnpziI130OqOOClTFt88O/wcAAP//AwBQSwMEFAAGAAgAAAAhAB9KVts6OwAAKYoBACIAAAB4bC9leHRlcm5hbExpbmtzL2V4dGVybmFsTGluazEueG1szH1bjxw5lt57A/0fCD0s1FhJlRGZlZfebi1UF7U0o2rVVJWmsX6LygqpcpSXQmaWWto3+8WwF+N52gUMGMb60S+D9YP94hvg3+K5+F/4kMHI5DnJj2REqhvbmOmuqCKD5BeH5MdD8nzf/PXH2VR9KJeryWL+7YPsSeeBKufjxc1k/u7bB2+unj8ePlCrdTG/KaaLefntg0/l6sFfP/3yi2/Kj+tyOS+mrybz94peMl99++B2vb77+uBgNb4tZ8XqyeKunNNf3i6Ws2JNj8t3B6u7ZVncrG7Lcj2bHuSdTv9gVkzmD6o3fD0bp7xkVizf3989Hi9md8V6cj2ZTtafzLseqNn465fv5otlcT2lyn7Meurjx2me1e+nX+wUMJuMl4vV4u36Cb3wYPH27WRc7tZzdDBiNTWvbfeuPDsg8KaE2yrvUN2ebrA8Wiwsll8vU4CoKnuyGN/Pyvm6gnNZTgmUxXx1O7lbPVDLryc33z5YvrzR5ZhKf11MzZdbl2+W09Xml9erxfTe/G6bJ39w8PSbA28u8wm/L2YlvWHzs/pQTL99cPnm6Pj19+r8tc4t/nb0+lfq9OxCXb1Wz194/v7s1Sv18ur0TL38Xl3+zSX95El0eX89m6wVGVW5nBRTNZlXJkatxqlX5fIDfdlYYm2ZmWn2TgtPinVxWa5tm/STMmleEsCdTcH89+ZVJhX/fU5fY7n4UdFn1h9mXE6n+udnZKnrbx+s1kv65Yenf/79//rz7//TH//r3/3p3/znP/7Dv//T//iHP/3jv/zyi03DqZ+W3xx8oG+kX7B9zRF/zR//w2//8G9bvOaYv6aqxB9/97s//+9/+vKLk3I1Xk7utKn5qnDC8/7h3/2rP/z+vx386b/89z/8x7/78ouzYn7/thiv75fl8mD96c7bilPxit/+/f/9n7/98os388naV+Jz0ei//6f/969/9+UXv7ov5msaINwsBwT9Bv++i79++PD0sEP/dIedjq+c4z77Rufl/F0xK+bFXP1FMbv7K3WuB9PirvDDwjM/f/H4+QtfIaeikGNvi6vaZrBpA7dp+mHTtJG/aQPWtKtlMV/dLZbrYjV5pM4WeqRd0c+2oTflbPMrrwnwt+G28nTn/rZW1cdtHbpt1Q8fntL0Qv9k4DPSxOZ0teev1XixLJU2RnV5pr570j/Mn5w8fnZyeanynvnjyttK8Z4X6uzZ1enFy2evvN+Vp/7lmfe7VtXvPBn1D0fuPwP4pUdu6/WD/dJ559D7pY9GrPWXZ3knyzKNVSd7/PLEa/c8y8s50YKpYrjFweLvIJOwk8Xl6cWvXx6feiHjeQBkVZubQJZ12MBLxMcxBzMCaPAeM4urx+kjyuwmr/Hrh/ATedoBKF6ShqDIBCA0gHx4yuyFjZQZn6r0U93JRoNDr81QluRe1u0Hepl8UaSbieSoxVUTAi3W8/R2ctZP244FLCOHPStHPSvjeaBlhBHiL0m0DJ4J4VQ1PIBTl+HU9falDPUlntztSxgxnqclYvwliYjxTAgxDUiwL/UYYvrJ9qWMmIe/L/XS+1JvGOpL4kWxvsSToxZXTcieDPKRmXftv/CMnR0yCPTTtnPlPgiOKItv2NXTVhd2Lp7HZyphtHj+RCvhmRBmVZubYMbYasZZYj1p5aij8eRuR8Po8TzN0dshvAkzvmgXQq+iv4FRifHfrGKQhhRmo4GfAFOi5PlqkIf6mCS/YVooykUtrpoQaDFjwVnFI6vVTC8DLeaM1PeBww2V/DeF0lHNXJhRcy0Phqw3Y7SXnnYoXJea7e8NxyJ584a34bKiUNRwy2Zhw3PGXc1TPXv0u/7vTImSLXsUYmLyRZHZQyQHLbZNwJZt/GQbJmae6sliCBaz2u3nUHrfBw43VHDXpMWKKBQ1N0Y8c0Y86Wl3cULNBpYtkjdveBsmKQpFDY8xyZwxSfO05UXAsgWNsyt5/1I+64XW8lQe6yQx207igrYRAdtmXDCviJSdzIdeLnhEiRAR6iEiJPL4zCKCjuSNKUO9KBXZRdXoAEaMLOacXm2wAsRHJHeJD0YrThsjaLXhjaKmCK2KNwbQYjQxr4hSxXw6mX+lTonS/WHD0PJCvinWiXjBqMkxspczsmeeNquJ4ci7mqBEqBMdwk7E89hOxN1geRieHbdoAiMWVUUgxfhhzvghPXl8XsMR6kU8uduLMFxedtkIrjb8UrQMwRXjlznjl+bJTkZ5v+9fpFMiF9OLxf38Rh3r7UiVHavH6rCjZl43ssiYtswUmValf2fE1hyPF13GJ82T7Ty9Qe5dih9Tol12NVXpDRYvSGuwyIQabFsQaDBzZXYdVyZ9We9ocUyJgl+WnBDg04qciS3lxcGWxphklzFJ87T9tF2vo4kSJX3aQIvbMEhRLGxxjEJ2GYU0T5teC5b9lCj8bWG3FTkTvy0vDrY05kTsMuJonrbftuf/tpzDVfOXp9tmuMVtWCDVzQUYtjhGA7uMBpqn7bf1OuSPKVHw2xItQv22DYMTxcGWxihcl1E487T9tn4KR4mS+m2gxW28daJY2OIYg+syBmee6m87yMBIxbnUzuSTYytuw8KoTklWHKNhXUbDzNPm2w7YJm/t8TimRGnfFre4DZESxcJvG2NSXcakzNP224KRKsKkurjftvHMUZ2Svm3MNddjVMo8bb+tnzNSoqRvi1ss3pA2C4lM6NvaJmAu1WNcyjxtXTZ+LkWJ3BbX+y9qTOdmxnRYR90tiDDPiDyvyxv1/OTKS5nFW55HFpoiOThnYusfaC4jVD2+J+wdpo4oEVpo9tFCU+SxC80NUo8TIWpDxETRCKgYD+sxHkZPnsXmYRcsNkVyd7GJIfNuCjeGrM2usKgugoy/OmBjjNj12O5wz+8FpUQuvvWeTdMu1WxrWBSKmm23huF2Ro+xOvO03Qn2zg5HlAh1qQHsUl6XXluk2pBDUWuEl90Wxngxbtjz7wL3UM/Cu8AYOe8ucFvk2pBM0UiEHH91oIMxrtlzt4Y7Pf86ghK5FldvHTXtYM32hUWhqNkxwtljhNM8bTuYdzFxRIlQBxvCDub19rVFqg1fFbVGeMXoao/RVXryTV2HqIP5z1bqs4EYOe/hyrbIteG/opEIOf5q3MEOGQ02TxtS2PN7FCmRC/Nmd2Xbw94tjTs1xgvFi2K8UCQHLbdNCLSY0eDDyi9nd6b6/uNMlAj1sRHqYyKP5YUesJIotHhb2sJBZEKAxTyTh4xI05Onk/XRWSaR3OWHGDrvMcvW0LWh1qLaCDr+6oDNMYp9WPkL7T5fz+8yoUQuzpu9qua9rNleuSgXtTzm8jxkzNg8bWaygX/1RYlAL6OzQeCsu8hje5kHrLRe1sZlKqqAAIt5TA8Zt6YnTy8boFWYSO70sgB0XpbdGro2NFtUG0HHXx3oZYxuH7Ld9J7feUWJGM4De6KqeS+TZDl8jlCUi1oec8UeMnpsnrab6f4FGSVCvQzeKBF5bC873AUrrZe1cemKKiDAYgT7kBFsevL0siFakYnkbi/D0Hmpdmvo2nBtUW0EHX91oJcxzn1Y+Vn1yW66kjTyL8koEbO5O7oJTHdCVfa110ZPRPooMeSvRw2MeYT7jAqbpw9Pq4YN8753l44SgYYN1cOzxc39tFh+5fWJipyxJorkoIm2zvjb9Rn3pSdG5OtvCNZLIrlzFy2D9z9kntiXF+nTeK7IhMCJ8dw+47n0tAsO2QECB7qOM3hmXxRxuQUnaj1tOK0oDsEUcxf3GZelJ78N+VdSIrk7hGIbEveOojbUxjEsKobASXUM9xn9pSe/LQGQ/ESYHBMBWxJ5GthSG9IrWoTg4q8OjEuM/PYrv2q9PvJvolMiF9MzHR1COV3IP7OIXNFhlxeCmhk7I9Dnd9c5MQUzC+eTO83z7lCeUEEuKNHm8eSoeTEW2mcslJ7YwrW6291DXjiR3J1V4JlfkSf127fhnaIoBFGMd/YZ76QnH0RwboEu3gwe6BRFpNpPG34pikIQxXy5fcYr6clvRWDIhL7cDFsRz5NqRW2ctqI1CKJUp+2AMVV68luTHyqR3J2CoTWJPInWJHKlkTmRCUAlUuGZZcAYr3na+rf9R3gokYvnuXvEYXG/vlksltWdlYdZn3b3l+uVefTzfPGy2IgsywYxNuw5WOZHYyEABozLmqfaMZD53fpHlAg5BuBOvshjHQPJiCmd0rs8Ei9ONB3eAmQ6FcHNA+AxhjuoXJ31QtDvR6JEzGbKm9KE3RA7ISZqSZLZNHPayuKB2cSctgPGWs3Txp008naWI0rksxrNVbHV+M62qu/4plHzntaGx4rqI5OJOW8HjL+ap01Ym76fv1IiF7g385tyaa3lvLaey/MXo7yjLifv5oWOwaQuKZBWuVJXOvjNZD6e3t+UqhiPy9VqYf5QRTYaL+bjZbku1XWxAt1LemvD3kpRV4RSjAYPGA02T5tByX9w9IgSAfPqQpebyGMHpZ8Z4DZnJ0TFEcwxOj5gdNw8bYxx6N/SpURs3TC5Ju/cawoaNla/WEzma3U8Xay0ARrLezl/TFHxSh1u6ZhiMfkH8GbnJ0QFUNNjNHvAaLZ52s72/kOdlKhF0ynCC266JNGRzsWTo6bH6POA0WfztPnqh/7TnZSoRdPpnB5uuiTHkaYnOW5tWzDHGzI6bJ42TR/452tK1KLpdJAKNl2+MRZfjFcAfHXblkDTGb0duqGesoH/q1OiFk2nIy646fICfiS0muDXfrJi2xJoOqO4Q8vqKh/20O9UoUQtmq4PHuC2S09spO1JFNU2JtB2xlCHznkCujbs5xuUqFHbqxGebsjipjfjqKICyOJjHHXIOKp52s5u/iO3lKhB008Ws/DUJl8X6+m8dNTuGMUcMoppnrZTG+jpguWFZ3XT7tC8RmUyGGPtTvKd2oYETJ2RxqHjO6UVBhjcBflKaHdoUqMyG7U7yalqGxJoN2NxQ+fkaycb+KkMJWpq58EZrRmFE6UjO49RuCGjcOZpO5mDYb0RhTN2HpzOmvE3qiJb+oPpLMbfhoy/mactfwPfuxF/M+0Oz2XN2BvVMaXhsW33EWNv5mnbcP8Hp0SNB/TQTCbfFxnZRHJg6bYluIePGHejJ7dNNibuEJ2rFMndbRN4O0Dkeb7/Kk++MS1ObBL/M+iEIjOOGP+jpx34sk5vCLbrRXIXPnj2W+RJhC80o8o3psGXRCENOkH4GIUcidOmFY0+RCFGRHIXPniqV+RJhC80Mcs3psEnPLX+4ZpeHY4LOmI0lJ48nXcADpIdi+SJUITmavnGNCiSmKlpadCSGDMdcdpnx7EBtCR4P4yO2qAA1414be2tCk35otKJ8a6TCC69OmJJjOCOOHuspwG0NSySOx0xx9NAI3pcwxdkDqIaifglEWV6dQQ/RpRHntMHek2M5gF4GjbH80BDZ2l8JS0qnQgfrwZiITG+PWJ8m558LASFVxfJXfODR/ZFnsjgF12Ny9eljXxJnN1AExz5eDh//5EFOPL5jyzQPmOOQ/s3Z/xB/tHmLMMoifZTqkhg7Y6I6+85xaBX+DCwvz+yv4YPRyOnMpsuHEIEhAoTroGkoLoiF+i5+uUxCNkKgmQhvCwEcOBjmT6lJ4Y4yM77krqizAXRiB3fzTpsQaAffZwMzaQyvTuW4ajbopAUCEM8RFYiVTYiaUlQARQMdt9hiwKKcO+D8BAtqmR6F0J4CktmSoEwyEV2XphohknrAv3yWKdkKwNag3oxxGYIb8Hl8HiWLCRpTg05RHZemIhh0oJCvzyGIVtSZB3vmoJ8I0B9QKR3j7hBViwL+QzekZ1XJqKYtK6QL8d+pqzDVhj60e8qQWjCMxgZ5MiykM/gLNl5ZSKaSasM+fIQmmy9kXX8C44RtE3/ikMfmIKuE1nIZ/Cd7LwyEc2kRYd8eQhNtvzIOt71BzlS/OdWZPrP4EnZeWUiMEkrCvnyEDBcM6HjWVzQRhA2M3ggGvtVMlFIIpphQtNmiSErAjlh6oFpkkxjGjZCaGvrZQFDIJYRw24WXebnPm2w88404xQ1QXCKZAHjlApj3oUHOV0QnP54GjQGYq8Lyd41gzPudtl5ZSKaSfsX8uUhNPnCRaiM1cbZgWj6D25rNKETJhOF7OuF2XlfIpRpCxdR2RCUfAGTeRYwRD/hqCnSO8QR+2QykSkFypBTZud9iVCmrV9EZUNQ8nVM5l3HkH8G9XF4MDzgnxGFpEAZ9M/I9yVCmbaMES8PQcmXM0Idrd4iGSKSI9Lv7aeR70tEJW1ZIl4eQoUvS4SYWI1KHxoYXJbk2F8jCtnbXyPflwhl2ppEvDwEJV+TCJUyCyX5bVBfhWuSHPttpARbwimrsN9GvjARy7QViXh5CEu+IhESaDWW2Cz9lzX1bIz9N6KQ/f038oWJWKYtYsTLQ1jyRYzRgfvwlOIj639GQD5TaK+dT96ti8lUiz1fHh8885PXk0zkil20k+khJ47uY3CNt4yJvFEj/RKhQm3teDGfl+M1XSwMNbGhopuuSsIBrbrGgc/IRd0yoZ22+Zxoq0akd93C2OUhMiVagciVKq6bxuitml0IKM7khdaaDWjQQTuCmUjvANXFTF5kSjSlNgFEZP1QnzGieMF9Bq4elwk5t61BgflKpHdpesCgBEFOG1Z2lOaSdvhELghUajiRzIjNOQLXHjJO+4VwXQNk6GhSChkWLyTRsNrcwtTNSxqpeLJQR+SMm+TaNqLFnUO/wupRBuTndGxQuCO/k4kCr4wn83feS3E7id8cHb/+Xl2eXvz65fGpL8upzAJNqWphjx2Y5arwRl7OsSH3KD1pQXul1zIhEld3tV4YE85vXxXX5dSKaz1czE1E/Ed0f1W9KOY3t3QF3H91fqfsCq2zZ1enFy+fvTJ3XcPQpdFsK7vH2IhAjtNrK1pXUcEcRLPMgGqdRi4gMM/ZbIUc0lXYKaIxQGnc2bY3oBDO9eoy81gf4qaIksC0/ExZA4QXcEIxzgL07Eo9rO9D/+qCrizdQJsShTaGLI0i1wjgIAMZF62rHu116GwAYjToVO5A6XZHvFITmSrMYOiFnTIaQ5R0QKhucI/FFeD9jsvdZUJUzsRs1qM54p0ivQXLBHLA47nMFBzPZeKE8VxkQeO5SBaY6owo3nZYFyJ2dpBCGsRHmUhfW9QoOMCLTG0HeFl2U0sT+SGWnN6HsOT8XSjdbQZ8aHB+T7zGEg/5opDIkC9SW0nS9DlR5IeQ8ZaEIOMed6GzZyEbwiNDIr1rfngSEJkaTwKy0MZ2l+ZrF8WEQOS+dqsQWIE37IBVkNDqc6HDc4HIZOcCG3DET1plOY3RSiP3ttEhlDipd1UF6WwbYBlC58+dMfFZIKlFOPlA8TAowOBqRdReXX9Sq3L6Vq3u73T4J4CauNjQGDUR39BXyvPMYhACjXvZrUChdUVk3ni9NCn4feuamuEjPyLTxQ5oxXpdjG/VeqHKj5PV2kC5nNy880cQ2alEYwTT6L9FJAQhp/9W8dBKRoxQ7/T71DWE2EchhAt3IRzMlDZBEl0zobZvi/m7KIyiIikcJW1dYIEIIcd96a6CYt4Z+WUAMqFl6PRYHMxeZtpFbtQCOe9aIbjkFHWH02vsNnDGhRirx20sLn9od50KrA6yAOHlmbzI3SzutTJ2us2JiqTYXNqqwUpSBmyOCztmrrJjnmV+ESWdyoOcWSpg5iYynZfvy+VvimKurpblfHyrOybJ845n6qa8W9/6pwlZcAJSIsuZf2aIykFmXA+yetzG4/LevD7SqRBQmK9JRUcPUOSHHHZUjZV6eEMTxHJyfb+ekMNoMl9NKNIZxTtb6bDxi/n0E3AZyeqlwMlbhOCMXj8wCpPbdRgTnMwyf/Q8isUI4YTXH2WmI4r9RqY2K+b3xTTB4kSRKRDxLAiiWOTwjCtNVo9biwMEzq8wabomJnAiE4coV5f31+rkfrxOQEsEH09Bi2dBaEUvEhg5ScegqlPzFe2g7Uu/U00oPbo+D8zcRCaOVrcJWuLQTwpaPAtCK3plgGtTZkyckrY4AVre28fGtjBJE+KQHK3zXx8fvDg5P1XUrPjg710jBAmGKBzBFbt0nJFO49NvHOOqNgMql9rQHzmMBn/vskDDFWBmIpOF62aypA1vGtxpyPpNMaVJc1l8KsAcKYpNsSqeBcEUu1ucGcVJB6bqNq3tgzlYPQkxSKcPBmiYyARhWi+mAKXmPF+UiVCKXSHOuGxl9Vi7/7sjNFJ53f/GmAKUi2eyKC1pe4m4V7lGBtScyAuBSgRNlMdzhcrMPNopjw7AeIelY53KJVm2lWclLf3At2/Ot0UZqIHR4y9cQJKGvO1Oa7fvH3ePdSpfAysGQ0tkTQQf/qKY3i/VL5+9BFxPvMS6QoOjp8gCGh2VkMyMauR2WGAikkO+o1GnOtKZ0OoMG7zUcCzn74obvca4of/flbNiRS4A/ePkrqgozR39Xf2mnF0Xa51MZ5gzf5Wi4K539Ot7k2BGP727nxfvKdW5fsk1mZn6rph+UJl56yP6zZxe9U711WOV5WpGoYWXfjuUTWzqshH50feJcnEuWZmZx5poDnuAaAKpSjJHLLuiX81OLvyc3ydv/n283D9990C0Fn2f6ELAKGA6/Ycz5mp6pX6E9lmERqU7weJFqBS2/Gfek7zrjgZfKmkRIjAJeFW4nmcmRC/rLwa3efzanmayx3sVUlnzn3nf8q59GnyxpIWQwCT0xfhuidEX3Y6BgJT5hUTNd8Knl6WMJ/pOZnoxy+56VvpE/pyZoglqfa+dZWo90fOMnoG2s87q4Eb1OsFZZ8+tFlF/NKpF11RGcdQZ1Zw1VZcO2vmXoEIJ1B3LsAdIyoe2RTyHiNMnD0LuXZI1MPak9ZmVcA0ZOd+acRVQCXK/GmUmtUQZeudbTpVpz8eNdhMddg56NNtqZkU+8IWhVCRXQRIL2pwBFxJhoiJBJWWtkBVGF2Nc4jQzj7bfEyTgHrXQCNX0cEszW0DySFXkBEDTLMKqboNLshA00cWYESt1OihfNZlJTFsNpB3wTnU30FXFvsZPZGytrlkLnVWEa3QNyJVTMyFWWuOKQpnK9PWGnz6djD28opCf2GJ35FqTjoGLXADfqGprxmVbq8dtlwbbqH65VjOVB1AVNyE2OzU0BtCycTWhHedqHNwOf+qTnrTfL1Y0LqoDdbG4udH7EWaiP7Ub/U/8A4GsY4JnT2RBmEaXiFztNTOPW0z9gcV1Kr7aq3f8zsvldTHRa+iLzbRAWGhP8CO7jkYQyJsYYQEBWQXU/ugSjMu4ZkL/tO6zMPyaX8jVWBfeZ5Aiq451tYZPihCkdcyk1ZGBKHiPhcu8UrQ6x/nV6YK9ZqGd6jI+vOkgBVc30F0vrhfvFfU40+nOXmire6HIQ0Ouofv35CKqkX2vXULTCeiIey5e0vRga4AChM5ov26naCsFW5sj2JQWqq7uphdeXYpM55uJ+R2hRB2Z/Gf3tKm6mrwl3/INbVHQBqvKHh1WtPgRjYLz3+iBbzaZT2b3M0V/eOTiXjOoWbF8r8fDu8UKMEVZ/ZRRMOl0mEUvBDffBjJCtdtREGxaC5VZF268SBSZtoclLNwVuSazvX+v3ZvVnlCxuqP9IbtanJbXxZIWi+brDM1ZAUq4/Ta0SBzzT0CeSnoXfSrt7Vzea0fow+ti/P7tZErRQd59RZ9y6/zcqKYZkb5f5vmh9bHSh9TVKIDvXaKR8vWSVj/2Y4S+Hl/9GB3e7dcDbk6/4K4euQPhQqQMrjNyzwnJd7rDXN2v7+/UWTHXV3oO6rs9YLzxblalrx/TZHkzi0gIQn5Czej0bnRV/Kc2jjOueKs+zqZfk4WMy28f3C3LVbn8UD54ujliULEA1c0P8qH6sVjOC3U7WRdkqV2lz7jcKfJ93CuAU8OVkpT99Z8Nsq0MwcIvbhtt3q2AKDh+JlRuXR6PNzh4Jj+WVsLvM0Haaqkk2oZoV3SpxKV7M/NYb5WORoB2Ssnexfh9uTbksjIrv+3syO9GmGWa8m5d5YDxcM3dzIru2k3zHmBFQvnWNR7s0xeZajtJxEfKliVRxzSF3rrRIZT4XROr0Vud+QdXVI8zKdJbn2Oyg0yvc9Dt2UHmDY0qeoy5NGPMOZ2kA0cQd7R2Y3aSdBrMticEAL85YnV2azMBVE8o3bpmgqmeyASGk3bgtVp/iAqB0cQiEoKQ3xuxir31pRvg65OSvU5vMRCA0URe7I5ZSdqtD1vlUBv5PoYV163NBFBUoVjrmgmmqCKTZzQJ4SPV0dJGk7SgSrbRIZQ4kbfiupUlkCPNv/cA1HX1fbUASoy8+idsfXygmE/+lo6MT0iqmE7zFubUALlC6Jru7VeA7cj6ND0+IPKjbhU99cUVdDNXQpfC7gL2IyRsa5PTYGLvMM/UEMw8DOaeFFu0B4EZ3Y/gmrykXWd8JPW9VHB7RqjyumBi5yXP1BDMbhjMPa+Ni/YgMKM7GFzll+K/OmDCS+NC59cFE/vqeKaGYPbCYHpPq6Uv/kR7EJhRLs51gzNXOJi6ORgzhc6vCyb23vFMfjDp5LB69kOzAVNWpumAKfIDJKMqxJkR69167VwdYkISHJMVQsQOkoHTxDxTEyTDo6WsTGMkk26RREWNMyP96yDpyBoTkmC0FLrGLpL43hfP1ATJ8FApK9MYyaTlRVQiOTOqyA6SjkgyIQk2zYRIsYskdmHwTE2QDI+TsjKNkUza5bDq0QFuySWXM1dzme5dI5v0x38lOhQ4TSmklY8LfetQbyT+n39UV8f+8JUnukLsYG1jnJJOXtlmh3DiKxWu0Yx2g4SqsmtxeEErMkmcvMZNOO15TkrkR7NF9JyU0T52eqYTiIrc0Khn+q+ka3vCaxUh3Sxx8jJ5wmnPw00iP8Ipugwxos8OTo4GNPU7hJP/3rnGCS9DhHazxMnrsiOc9lxhiPwIp+gKg0tGZ65mtN7W8K99hWyz2+/wCkNkkjh5fVaE056LB5Ef4RRdPHCJ6YxpTI/80ViPdCrfrQVtT3jxIDIxnLKO+kibpeOZ388ky2s8mPP6IrCiiwMuS525utQkWQa8k0IZ2jUqvDgQmXbBIsQQWrLApmiJ/ACtqJR1xrWsq8fNNloH7EQCDWsyrcBOpMhEUSlv7ifratnUPagIld+yZHGNsUqi+FHd6owLV1eP2y1HhBUMLRVQJpD60lOKu6IZVQSnPS+JiFKRTVUrGzYF8+hvRoN5O/1ZSebKcTRCwUOAMrW2KUzgZaZbcolTEMHKsujUehSxPS9riPJR1JCoKHXGVamrx41poeuqQjvaGbQCgfRFpmM6HnOn3lIAa+3GqKGzAZJi9rYnfRdVgehFr5AbqWbH4CoiW0dtBm5fIR3toof5u8y0LnWAVIMhucoVyYPM6HzAj5Ob9a2icyj6YX07Gb+fl6sVbenNZuov1c2neXG9mIIzRbKExoNd2sZEVN06M/rNDqhu1NnRCLiGgKy17sWY7ItM2hBPqrg1p9PrxY8qe0IBfsHUsCfjl0X7ReUrLIIHCkkJ2r2Wbx63xwPAZpeQj3ZNEFN+kcmDFgRrT9ovS0ZgRXk/l67OrFyz7a85AssfdVabFub9Uue62jJ9ePL45PIru6cFDGtP6i8KhmNblPsbPWanGzobB7RGQnQWcv+ATpfQh35VLN+V6qK4mdyv1FVJm4Enj19dfUX3Qz6GkdtzkyBNpjqL6lSTObldsnrc0jX/rKBTgVVTDhcCMlM1amkjexU0sp3CGo71Mj8wsrrl2AGWkwC1M3pVj9vRy3/lUKcCUOHY9DLTyeIuZEw7ZTRGKEkgom5wCCF26Cc3ktWbQNl9/5ClUyGEoHtfZroq6PS3PoVczHV4+gO6paEPJS8Xd/7oizuFNoYsxY1fAxBCjJ0Syq3+sl0FDPwSMjoVu6dBsWaKFZ2Yo3jzSp9ONUyVwqqrrxVFHftY/YviS6N1tnxh07C7OxXyz3l160JwsBM/uZVStqdhQKRYnQoZEFwUyUz2NMx5AMpULPfj+7JicLiK8f2cZKLZcOXw/c7ALxJxrDO5WNJSZ0yBZ8pd20o2rf288rJCEI6YWz7nMtHV4+bMb8fP1HUqZFrw2KbMZE3rM0C5H42X9YJQxjz3OYlCM8tinntwB1dnShi0SOrga30f/mP1r35gzNqPp+/UB41ZMZ6ek6gzQ6MiqzZ+bt/vn9eZkGHBdbXMFB6zGiC5H4uX1YJ2FWPxFDOXI+ke/+ln3h0hGrE4k/Z0s2o6JDweJNvWfuxc1gkiEnPT51zYuXqsVTS6wE2vUyHbgu4FmQkPWs3RlPVpyrXS9J1rcALsgus6U7hc53jZyB/p+linio9bd/fmyhX56ddkY3nno/kf5FvinY35lqwTGLtsA0OIcMJOSsQbISmaLsCkKPSKHYcMDhSQi0zBsaspmvt59GXVUG+14ITQ5GSexIg1mnUwejB+CcliQBOaQrKfyz4XlYKQxKK85qQi7E6O5nET6Lvv3+TXmdAABp1YMlOYdTWFc09OLxoE4YxyepIfZnA6wm8UjtAfECYXmsU61kEVHI9uUp+/vFLkIfuo/99TZ2oTldDr8JOvaj5wJfnddTG654S6GrsQkJOW8GbgIiD8V0N0Kh7fYBNf8Op+WVzrm+UaDx178KhYUryDG+A9EC9qDkOSdkPdqhAMnImTDPBmxOmMBsDhIsSC3fEbnnnIRaZKWeXFC2Ame3JzURjsL1FuTlq+rL84p/OzThd4N4UAsINPD+qU5iJTdcxhOaHrsvqw+PSeArwX4zFtcFE483IFYNuTiIs6QNiiRNwoGW/c6bkQH7YTWhdFz5DpawB1qMkAARU83hyObArgnrxdajMjXsWLCfRPrp6cu+rJdMAIjNdC2PgHulStnZnrQt81urqfTigoKgWOPTp9pc7+Rl2cnr85evXy2B9GVhfpjnkx4WiZ3n+8oW5IqOXcWS6UiysHJyEAglHlQFRZB03CY5TItCdwbe7Pynoj+GLxe3IjMrztgVZz2MKGnApAYJk2/3rYky4ynTnxPNRDiq9CjhgbKw9Z2J70W1QAjVpR2eWcyy5Xj5s43UAVV6cCXLOHncciUzXYX1GI981Yr6WLCDvtkY8O+7IKTdfHIj+yuShX5yrMuXnc7m/5L7brVAg/7CEVmSr8XhU3NwRZexj3JOmiUgjGKEc3asxO13U4egcd7KJwLBBG7A8UmY5YrCTdddeT9eS9n2vIAuNxXWQdYUeNkncu1ZxbheLaYwpYK5Bq1oMbJhUiU4XQuJzo+Dg6eFQQoca6DLolbIEB+INtcGj25LzeShTX+2DAkyAkmc9J1GoyVheTd7drtXirfqDh6eHF6x++OqBwbjNyVD0rl5PCL7eQy3edXpy9vPJZ0ulOUtToKFnnMsu5FRmur8KDvVChmPzLkoIeUXgwHbHHNvMxLeamxUQdlSsKsPSwevieIl9pdQ40pwkd5FDzU6J31o0JfXLu/Db6yduYNOB0gU8wuUu3J0LytrnIdP7GIgUGChHhMwRFyuF0XXxkcc+FkXOfMDK5e0aIO/qEkQmUQRAUKVwcBkWmDoAikoJpJVkQOeeCyNXjxk7813qPdSp3YHpFvGS6WN4EP7yUHw61MeXseF3VQB8g+V53uW4et20DyyWh+UuKGuFmiSCXoWYlHfKwtQw1i/uFSTh362UniQVfv6NPxonp8WI2u59P1p/CjRMnQ0KNS7k+qWsR66zcw2t1bu2o7b+ISo0T4VMoUF05XUwLdbWYhhsoMoYamHLvUdck1kDuc7WatHUD/TGtciGxq8MvVg18deGdRk92coRalhIeUb8w1jLuRLVSsXXLAOH3iefaKQcvmKR47tXBxQ/h7yyYVAiNlHCDuW1ct8t4BLuSkRvF1y1zF3K1G1zgvOO5m2jnnQAyPNNFHBlxjiKEDE+Kph6eKjSOcW8q07vNDr2nbKmrCypFBI0EHdfFrTopyiX9p6JpfvIhM4cam8TDbJ1DjeQ8zCqy1g4r7+KEGonCoF8ZyYdBFVVVHT54pC6vyG/36jE5shQdz/XfqZTvi3rtZPmAg0flZXMuL1s9bmdg0Podpdjtvkq75osXPn+REm1KVzZl9RWVjs25dGz1WDtCsr6Xih/rVHxzqY6QqyobuKI4mBTdVY2a2oJ4cdQWZEWALZhGBvfYuORr7kq+Zrn/hhrBIGOIC1toD4OMJJ5mE6I6CIxYQPGci7tWj5u9feTJFzqt2yDMfFjopQ4L4n1xUxB0ELU+SvK4WGtuHuthoefX+TvSqXyuQR0EAbujRSZHnKMtZjKyX5rdJMX3q4EITCdcuDW3wq3VFhrcARIyqFu7AWNJugHJMH7hMIe6uklDakU12QKNUysuyJqbx9qAul3g8vMLspIaKwm9PH554l83iUw7BiQHoHTsOM9MnZTSXIIWkBCC3CVopFs3A9Chl6zTaCzI4iYKPO9MVcgjPwUTr4iPOaJINOZU7kAeg1rYDKebRoZ1y0WAR0wom9Z7z9pm8H6EyAQHnSBOMpZ12jgj2DFCq9qyD6LFeauVYjUDTd4fIPsQvFVfHjmwg8ybm2JZqIE604KYKn/iD5OSC6XVuH2IIlGL7eFf9pm5fXBt1tzVZu0NckDThEaq86nPvY3HDRdvShwPZPmg+bYxvBWi+XyH3aq02qtD/oXFcS4kSK+M7o2+OaW3et9Xx8IqmbqNxsZKb5Y8oq3M8S1dgNeHo4r3q7I6xwLc56KUqE1IGVgESnTbnCuj5lYZtb5OhUyCc8RqaFwt3mvBn4cVGKMZamkzfRtdo5QJ1VY8wCu4xmhuHrchScBCTQhi+lo6gC2VoaAj1EEKkqJvGuWeXJszFxKRG3tHXhm/NqchEXhnVWT6qbpJK2oqdUMRsFHHptHM3Lq7hE7kptP4gaWhhHPDZt2mTThpWSLagLZaoKHOwx2fQm/Snm/rAIFeajnndM26USsWKRUx0TePXt0yEpLON3cPjFIEPW8wsVwoSrrn+QJdiPPA6ozHdEFXmemKKh30OLibFutKIuSg+mN1ZsbPQmUVmp6RkaKYCECLBwNCTLuclQrxxtpPeAiHI3zxC1+dF4Xsi+Wex0xFbWA35MWEuiOnrkbHcjOXdb2nHY5yoe7oGiUMCyIzVZcE6CiAtkN9VuvHybI8MISHvHSrld61WZIw/MwfzW2nDo2tMmnD3OIx4N2TWyXXqMyt6KK1xh5YKwkBRxdCfM9CZMqeqHPqyCXJGk/XdPMJQSXLagqVyI+Mzjac018BFefNRn1y493t+QOU5X65yY4+0ozPo4pM+RNS3F2RWpc+eqLGOh6Sf7Dz60YmRxCXlYVYVXQ6sMLiKpK5VZGsxzgElT+eg4YKj3BCerJLUJmoUbd6wtA9c3l/TQqd5hA0AG3PY6hS/BLMEBaDwGjGpSdzK7NoQRv44wnrVD5fKYGGgwXKTL0n6uh+erN4p96R7hZAac+LYVLjEqFUsfosC+y2cnXJ3KpLVg4LClbhpyJ+dUndD3GgQP1qF9zDJzYY2VA9rKKSfVdJtIM1niyz8dCV5ku2AIQsi59EsPqHtWUB8uZXjzSIYYeYyNR/QscI9VUWo3dCKpH+QGO5X+qxwcCV5m6OCj/mRlxwS3Ot1mCFVH+AbMsf7EHbFqa5Qh9x8ERt1GFuiFXoyHcGtBxBtmdQByk8iTpkdGVgxAQdyNjKoO8PoUWORThs4UO5ItOwHutDM+KeF8rSBBV1c/SRmdCMyNm/K6jYyYD4Wy5kCx2elWOqKjKNiGdpB6UBSa3IJTcFlxNlaY2HqzSXtG06twzBtDivtzKLdrwCCiQ5kFnUvTBASjmRzjoMrVyNzW2yyXy1LqbEVol/gelxzwtkovKQeUWPwHIhxdwKKZojSBndXPRPjlJI8YxCdlFqDR0mqSJTlj0xc+JKnZTFjSqJej2frNdkb6uDyvbo8vCnwg+frEBT00vTaKzRyHioGW57XKUxtyqNdgYYgRkAqDRqADF1FZkyovlHhRaSP6cQu3ptaVzIADARHa4xYGmR22zre8wtLPDiR29dvcYM6DUe5VKvcWtwgR1+kSkjrv/ilC4wPnulTk6fnajT70/UyctnZ6dXpxcqO3xCbrjBk4468x5fO9mpQ2MI0zzyUcHH3OgZbidQK29onYmIcwDBR7K4QHhAkSkj4n/6cU33HvS68tkSOCtkUY2BSjtMEpV1zI0oogNU5aiuva7glLeUdVzTolBf4jmeLlblDUXgpJjW+vrujxMKKvzdtKBIwufFvATXW8TbohtUIj0a16Nyj7mROnTa7lwOo4ArYIMKKDfuCUErF7yoCgQievbYSDg6QHCVFLBdzRUQgTJ1Dctr8rRs7QKMv9IbH9nUEhKMsPlRts01GHOrwWivwMHBws+29dkGzLZTNBhrLdBG0PHKJG5/C9lFCGD0shjXXcyFAuEGSOSGBwqM+n55YKZnzDgs5t0IylZHR0QTIJTJbniuvpi76otEOcHxbqFW+APxa3W2uJ+v2ch8qeEo6MKRb4hWWa7eqIe9Q7pnPiPaebe+BZ4YUVh82E47a2JbGvC8cC3F3EoG1lYGpiwhP2g72U+LkCgzsVeKXMiUokqJOVdKrB63pxLATSchTvj6bj0Z00HpkwltykyuSdmbWM3zJUnsPrw8Pnh2fvyVeqy6w566WyzXK/Xw9clzYC7ixVFzEekhDNFjKFzmMLdqfrW5ABcwkDnUozv2EohM1sY+I4StTl2LWkEgo6euucphLiQDN4CiUR7oHdIoH4jkIzJ9fkjbaKrLpkNI+ctDQxq/t8dlD1F4YKFFmGRnh4N+QleVR2EiRExUBMIRPQRjZAy3PNSqGtaWBW73CanAJiaShkYrbi5qBTGJcnOuYJibx63Ojtd3cqxTuZssSaYxyDsJptGQo4uKQBiiHJ0LFObmcRO7EJ2oFap/TUwjDY1WtFvUCmISpd1cjDDnYoT+c/xkGpyCJpnGsJ8ywUviHBs10lzbtlWhwZN7tq3WYD1qgM0SoDWoJ3js2RaZmthTGoScIqcSxjRibXEJAMl1CHMh2LcBFE3wQJFQT/D4ipHI9NkhFe9PhFTkQl1UJAtBy8+5WOU+62ZDx+OFimBSV+0N44O4eG+Uiov0EI0oFedyhLl53A7i4Jy00PhrYiFJYLQi1aJSEJIoqebKg3ST0w1TnIN1vhDxSzKMrJ9iGQ3PlaepCdatCvUPToCNuODWMkC4IaHG18Qy0tBodSQ8TSMwt00MYcJPmFj9u/qMPWB+QowvyTSIEyQMGg1vK4qKwB4SJcBc1S83j1vTAF56oZLXxDTS0Gh1cjxNu69uYsg0+P1DV7uPFIMRJpytfr71olDZi88nabcSbatCMPADIFaVzxIV/03o41zo3H329aJ4fyrjSGPDtokhTDgbtqJzWpxcH49Fc6w4skGHNBZLOt9S3BVjCnPk368Rsnfxz57GU6MyeV0uk1c9fnhaN9E/WehU7pI47yU0UWaKNVGmB4NeXWP8Fbtc3q563DbR38F1KreJxH7iX1Fmijcx6ZRDXeNQE9khh67Vp6u+4jD3T286ldvEQYqhykzxJiadQqhrHGoiizXWtYJydRP9OxU6ldvEUT/pKzYjbrIQaKixC4Fdowq3cdnpR7fym2EHrBlleveGODwltVNImhG0IXOyKIhTzLXZ5QJw+tGH0xDi5A9Jq90VcItUFpI45LVxd8qiIE4xttflynD60YcTCjcp0zv2hDcZZKbEcbMNA5RFQZxi/tAul33Tj7s40SgK7QmeXsC+GllI4uDbxkcqi4I4xXykXS4Ipx99OKHQ9zK9a08w6onMlDiCtzl3IIuCOMVUJrpc7k0/+sdx/36pTF/jpH1/AZxasE5ZUhq9lrkgTrxGgamdi8FRZEEwngO8gCycxgteBJGFpI3noqREvNJk3mSNQngxB2hXiKbVPGH0GOHlV27WeMHrILKQtHFd1CwVrzRWLF4ewouzY6HYtmHJEC//vT+NFzyd1hWFpI3vIlMqXmkUW7w8hBen2kIzbUO5IV7+K38aL3jeoysKSRvnRaZUvJLO+8oahfBiDteu1YKrnIvo7KtO5cwK/jN358V6fEvrzht1eaa+e9I/HDzLVXVWiI4K0Q9vzo/ph4zOlymvX0GU8jS6XEsTdKubGMKEOVy7Vvis3qTxH4PVqT4XJl0MSTPnq6jUUzTdRaXdSNbUDQJePW5PH/hPZOhU+0NCZ8siZtLsKIKoFcYkSr2NsNt2yct03iimkzcSXperqLXsOhUmATORXDu8BS8qhSGJsmwu7dYV2mebUQWtRoDIG43aITaUcAD4MwxGrWh5mvibxikcFbzLxd/0I3NiGqduJ0cBTmR6d/kSoE2skL1G+YCptjnqINsDx7XYlb4u14LTjz5ckaieTO/iiukVL2SvISAwe+5ozKWEPpQNQsBazbzAHEqyb45qRNcrNUeDJBoJgNScHgkwD+OZfqKxVdQskbeJXBDX2NGJLteg0487Btsh3ob4LVCj0/wWHncShfxUdE9ULRXYtAWEeHnIcPkCQui4bUggBBgvILAjmxfyE420oiWp+KYtOMTLQ/jyBYcQeKvx7UF8/fGrtQFjBzgv5CcbcVvtLAgA4MjAXx4CmK9ehKZcxcH0yAs8LED0jgA+hEK7XZ7ppxp5W+1IiPZAfPnLQ/jypRAXygPR+rpAKE8fUcB+UZHpqlytx6Q4pF5fnVyov6BIVX+lXp+fqdevrmiFRDdu/MtoWXRcVFDWFmIWXSIZjbztEklI5vmP5HaFzF1NqjRW2CcqMkGsTp4TVs9P/Ho/O0WnYJV0iEW/OULxudJelyvtdfzR2HUqd/53scLEXmRCWBFKGqtnCCtvxMLL04tfvzw+9ZniqawttCurCM4iarDICF2jV+cYFl9F1Cd/4HrIJ9Gno3KEIlbpMl2s23ZHbzSTMGxJB2Zk/QJDGNfy6/q1/Dp9RM59Wn4WPTyYiUwtO6gsOqGDSllAfzAmCUIIPb60EZJ8VpEKBao96or0tssa28PDm8jUssvKolPQS9vQEG8Oocc3NFw5QYpp4g+K2RVygvUwRweoOvBKwU6m++WH8pOdN0/K1eTdXD18tqKIt3P/VdWdF4SGNlFDv6yXfqOeBbpCnoWPbUbebzu2Cc3BKtwQIQV7p38FMgxjJTI1x0q8IIgVT4uwSr012TVygQ5g1SGcuiP6I2LqTL6ZU5tUYBDjmZ6tKLjjhEKsniyLH3Xgk2eZepipVblWBQWsVX+puuZhvLj7RA+rxds1HeK6+/TVI/XLs3/xSB29/lVtj4vxvRbJxWGwdiocRDgpumEFHB0ndGdrYYt8MSEUFDcQQ1v0H1fSIAfGOp7p5wVZVDgIclJgxK6VfAyBzFcUQpdxCzJYsfnEH4nL6A4fsGS+a/LzgizCKwZBTpLq6QoMQvMPOzJPMkdsGNiC7Q/nJtPXM5EGO2DRvJCfF2xxwisIdtpaRmAWApsdzCedz63IbjYEh9B1KndkPl8uflOO1+qsmBfvjJK4f50rspF/6/kL/zJEpETzT1XZyFzNTtl3rZZjNUcPgdCuThXdF/2hWJa3i/tVCTbHhbhkqLVpqwdb9cAwxWUou1ayse4x4Cy60H+kZlFMdxJMrtwXZwWpjpZzkrEpyYHxvSK9G/WpLJb+Lyze9f1r9eYSrDTTZCfrNgQsmMtOdq1AY/WBaS/Cv8kr1B7rGzbnix+pfZV0U0ZnpuazR4ok06v/Hh7Sf0G7OQcPtjuNrkd1JrtcZ7J6tLetKHil//qITuX2XOOneli3/moyK9XJYkafXF2Ub6fUpxczLWMF2Ld4WbDVaZsAViwz9LW589/oS27UzUD8ja4UgLw3oeQv76aTcYksmXPbYNvSHPC2qqG2ccc7k4uk1TuwZOHTvrp6oV4tdCSgqyV9SOK8wGR5vmAD07iqrW+ogZyrWlVHOz5l/jhHXS6m6PdcP9etrturjsvlevJ2QvENrz8p86djCq01nkzBSC3kGoNQpDFK27IQFJxRMmXHzhCNWpxq/aKYTWhkJinoKWmL0eRrfl5eF5P39NNRuby/Kyio6vw9BdemMX01AZbA3xpsfhrXs60JNZ9zPCvLWHsF0eDFyc+2+bNyOrmlJjZqNX9ZsNVppMs2wtPqg9VtWa5PinXx9JvNj8r89PLm2wfdBwfe3/fA7w/177fvvCzX9FhSKNPlvJgeLRbvncdXk/n7p/8fAAD//wMAUEsDBBQABgAIAAAAIQAdgWOzvAIAAPsSAAAQAAAAeGwvY2FsY0NoYWluLnhtbHTYy27TQACG0T0S7xB5T+Pfd6OmXSB4AniAKDVtpVyqJELw9gRE06onbCLFntifZ+Ezk+vbn5v17Me0PzzutosiV2Uxm7ar3d3j9n5RfPv65cNQzA7H5fZuud5tp0XxazoUtzfv312vluvVp4fl43Z2usL2sCgejsenj/P5YfUwbZaHq93TtD2d+b7bb5bH09f9/fzwtJ+Wd4eHaTpu1vOqLLv55nSB4uZ6Ndsvis8ZTnd/PFUUs/Wfz/nziapt/nei/nfi1djKQ6cr/r3uq1HPt3o51IyMak5P/+aHTe+hzkOth56f4dUdrW+sb6xvrK+tr62vra+tr62vra+tr62vra+tr6yvrK+sr6yvrK+sr6yvrK+sr6yP9bE+1sf6WB/rY32sj/WxvrS+tL60vrS+tL60vrS+tL60vqQ+I/UZqc9IfUbqM1KfkfqM1GekPiP1Ga0frB+sH6wfrB+sH6wfrB+sH6zvTe1N7U3tTe1N7U3tTe1N7S+kOtGd9Z31nfWd9Z31nfWd9Z31nfWd9a31rfWt9a31rfVnXs80pbW+tb61vrVeWCOsEdY01jfWN859Y72wRlgjrBHWCGuENcIaYY2wRlgjrBHWCGuENcIaYY2wRlgjrBHWCGuENcIaYY2wRlgjrBHWCGuENcIaYY2wRlgjrBHWCGuENcIaYY2wRlgjrLoqq6oqqpoqqYoqqHoqp2oqploqpUoqpDoqoyp63uGc37SyqqqiqqmSqqiCqqdy2vNGF1MtlVIlFVIdlVEVFVENlVAFFVD9lE/1FE/tlE7lFE7dlE3VFE3NlEzFFEy9lEu1FEutlEqlFEqdlMkL2093QG6A3P+4/bmw6+d/gAv7TsZc2Pm8HXOBxgurcX51YT3IGNdTLqdcC7oU5L3BNFPMJNPLFFPrBL884fz8t9XNbwAAAP//AwBQSwMEFAAGAAgAAAAhAJAHUitgAQAAdQIAABEACAFkb2NQcm9wcy9jb3JlLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHySXUvDMBSG7wX/Q8h9m7SbY4a2ww+miMLAibK7kJxtwSYtSWa3f2/abXVD8TLnvHnynEOyyVaX6AusU5XJcRJTjMCISiqzyvHbfBqNMXKeG8nLykCOd+DwpLi8yETNRGVhZqsarFfgUCAZx0Sd47X3NSPEiTVo7uKQMKG5rKzmPhztitRcfPIVkJTSEdHgueSekxYY1T0RH5BS9Mh6Y8sOIAWBEjQY70gSJ+Qn68Fq9+eFrnOS1Mrv6jDTQfeULcW+2ae3TvXBpmniZtBpBP+EfLw8v3ajRsq0uxKAi0wKJixwX9lisVZPihv0uMnISbldYcmdfwnbXiqQt7vigZdqjWZWaY5uZNDjaKqs5BuXkd/p8EY30v4hkChIsv1Ix8774O5+PsVFStNBREdROprTlCVjRumilTm730rvC/qg9D9xFNE0osM5HbLkml2lJ8QjoOi8zz9K8Q0AAP//AwBQSwMEFAAGAAgAAAAhANpNOYyKAQAANAMAABAACAFkb2NQcm9wcy9hcHAueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAnJNBT+MwEIXvSPyHyHfqtF0hVDlGK1jEAUSlFjh7nUlj4dqRZ4hafj2TRJR0gcvexvOenj7P2Opyt/VZCwldDIWYTnKRQbCxdGFTiMf1zdmFyJBMKI2PAQqxBxSX+vRELVNsIJEDzDgiYCFqomYhJdoatgYnLAdWqpi2hviYNjJWlbNwHe3rFgLJWZ6fS9gRhBLKs+YQKIbERUv/G1pG2/Hh03rfMLBWv5vGO2uIb6nvnU0RY0XZn50Fr+RYVEy3AvuaHO11ruT4qFbWeLjiYF0Zj6DkZ0PdgumGtjQuoVYtLVqwFFOG7o3HNhPZX4PQ4RSiNcmZQIzV2YZDX/sGKennmF6wBiBUkg1Dsy/H3nHtful5b+Di2NgFDCAsHCOuHXnAh2ppEn1DPB8T9wwD74Cz6vimY74DaS/NfpYG0vGt+kEx3z9Edy684GOzjteG4GPix021qk2Ckpd02MihoW552Ml3IVe1CRsoPzxfhe59PA2fQE/PJ/k859WPekp+Pnf9DgAA//8DAFBLAwQUAAYACAAAACEAbB+NqC0BAAARAgAAEwAIAWRvY1Byb3BzL2N1c3RvbS54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACkkUFLwzAYhu+C/yHknibNlq0dbcfariAeFJy7SknTrdAkJUmnQ/zvZsw5PHjR48f78vC8fMnyTfbgIIzttEphGBAIhOK66dQuhc+bCkUQWFerpu61Eik8CguX2e1N8mj0IIzrhAUeoWwK984NC4wt3wtZ28DHyietNrJ2/jQ7rNu246LUfJRCOUwJmWE+WqclGr5x8MxbHNxfkY3mJzu73RwHr5slX/AjaKXrmhS+l6woS0YYouu4QCEJcxRP4jkiESE0p0UVr9YfEAynMoVA1dJPvyu2nnVwi354tc5kU1blVcnCaEVnUzqlMYtKFs3JmsSeSlcvIU3wtZ7gi8Y/hSYXofunB7+zGbnLx65vtsL88KOEURTSwD81CGeUkd9s8PWX2ScAAAD//wMAUEsDBBQABgAIAAAAIQBMSvYvNQEAAJgCAAAtAAAAeGwvZXh0ZXJuYWxMaW5rcy9fcmVscy9leHRlcm5hbExpbmsxLnhtbC5yZWxzvFJBa8IwGL0P9h9KYLAdNNXDGGIV1Dgdq9VaGQMvWfu1zWyTkMRR/70fY4KCsNN2SMh7Ie9775H+sKkr7wuMFUoGpNP2iQcyVZmQRUA2ybT1RDzruMx4pSQE5ACWDAe3N/0YKu7wkS2Fth6qSBuQ0jndo9SmJdTctpUGiTe5MjV3CE1BNU93vADa9f1Has41yOBC05tnATHzrEu85KBx8u/aKs9FChOV7muQ7soICo0DI3n1KuRuyV2J2twU4AKSiwrQOX3ubafzEYtnUchayzh6YeNkuzTqE1J31/XDQwx6/1GJdJtArbECQHYVTs7hKFoh2cI1neG2fl8nLLxnYYwgiU7092m9GY2jxQNyC/bWbirbnCyFKsPQ7Mcwodfb6fxPO2dZ/ygcvfhPgyMAAAD//wMAUEsBAi0AFAAGAAgAAAAhAAIwKbaTAQAAMgcAABMAAAAAAAAAAAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54bWxQSwECLQAUAAYACAAAACEAE16+ZQIBAADfAgAACwAAAAAAAAAAAAAAAADMAwAAX3JlbHMvLnJlbHNQSwECLQAUAAYACAAAACEArKdo6KYDAAA+CQAADwAAAAAAAAAAAAAAAAD/BgAAeGwvd29ya2Jvb2sueG1sUEsBAi0AFAAGAAgAAAAhAGqqelQnAQAA8wQAABoAAAAAAAAAAAAAAAAA0goAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhADxCRLP7JgAAXB8BABgAAAAAAAAAAAAAAAAAOQ0AAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQItABQABgAIAAAAIQC4b1J01wEAAPQDAAAYAAAAAAAAAAAAAAAAAGo0AAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWxQSwECLQAUAAYACAAAACEAl4HA1cABAADCAwAAGAAAAAAAAAAAAAAAAAB3NgAAeGwvd29ya3NoZWV0cy9zaGVldDMueG1sUEsBAi0AFAAGAAgAAAAhAKSPkmyABgAArhsAABMAAAAAAAAAAAAAAAAAbTgAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECLQAUAAYACAAAACEA6A7lkE0EAAAiDwAADQAAAAAAAAAAAAAAAAAePwAAeGwvc3R5bGVzLnhtbFBLAQItABQABgAIAAAAIQA8RvkqqA4AALw2AAAUAAAAAAAAAAAAAAAAAJZDAAB4bC9zaGFyZWRTdHJpbmdzLnhtbFBLAQItABQABgAIAAAAIQAfSlbbOjsAACmKAQAiAAAAAAAAAAAAAAAAAHBSAAB4bC9leHRlcm5hbExpbmtzL2V4dGVybmFsTGluazEueG1sUEsBAi0AFAAGAAgAAAAhAB2BY7O8AgAA+xIAABAAAAAAAAAAAAAAAAAA6o0AAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYACAAAACEAkAdSK2ABAAB1AgAAEQAAAAAAAAAAAAAAAADUkAAAZG9jUHJvcHMvY29yZS54bWxQSwECLQAUAAYACAAAACEA2k05jIoBAAA0AwAAEAAAAAAAAAAAAAAAAABrkwAAZG9jUHJvcHMvYXBwLnhtbFBLAQItABQABgAIAAAAIQBsH42oLQEAABECAAATAAAAAAAAAAAAAAAAACuWAABkb2NQcm9wcy9jdXN0b20ueG1sUEsBAi0AFAAGAAgAAAAhAExK9i81AQAAmAIAAC0AAAAAAAAAAAAAAAAAkZgAAHhsL2V4dGVybmFsTGlua3MvX3JlbHMvZXh0ZXJuYWxMaW5rMS54bWwucmVsc1BLBQYAAAAAEAAQADYEAAARmgAAAAA="""

# ---- BOQ Constants -----
_BOQ_SHEET_NAMES = ["BoQ NRO Cluster", "BoQ NRO All Feeder"]  # Aliases yang didukung
_BOQ_COL_DESC    = "Description/Item"
_BOQ_COL_QTY_MAT = "Material Qty"
_BOQ_COL_QTY_SVC = "Service Qty"
_BOQ_SVC_HP_CODE = "500002155"
_BOQ_DESC_ALIASES: dict = {
    "pengamanan perizinan dan k3":
        "Pengamanan & Persiapan (Kesehatan dan Keselamatan Kerja)",
}

# ---- Helper functions ----
def _boq_find_col(df, keyword):
    key = keyword.lower()
    for c in df.columns:
        if key in str(c).lower():
            return c
    raise KeyError(f"Template column containing '{keyword}' not found. Columns: {list(df.columns)}")

def _boq_find_col_opt(df, keyword):
    key = keyword.lower()
    for c in df.columns:
        if key in str(c).lower():
            return c
    return None

def _boq_is_special_svc_skip(desc: str) -> bool:
    d = desc.lower()
    return "pengamanan" in d and "persiapan" in d

def _boq_excel_val(v):
    try:
        if isinstance(v, float) and v == int(v):
            return int(v)
    except (ValueError, OverflowError):
        pass
    return v

def _boq_load_wb(b64_str: str):
    data = base64.b64decode(b64_str.encode("ascii"))
    return openpyxl.load_workbook(BytesIO(data))

# ---- Main BOQ functions ----
def _boq_process_boq(boq_path: str):
    df = None
    matched_sheet = None
    last_err = None
    for sheet in _BOQ_SHEET_NAMES:
        try:
            df = pd.read_excel(boq_path, sheet_name=sheet, header=5,
                               engine="openpyxl", dtype=str)
            matched_sheet = sheet
            break
        except ValueError:
            last_err = sheet
            continue
        except Exception as e:
            raise RuntimeError(f"Failed to read BoQ: {e}")
    if df is None:
        raise ValueError(
            f"Sheet tidak ditemukan. Dicari: {_BOQ_SHEET_NAMES}. "
            f"Pastikan file BoQ memiliki salah satu sheet tersebut."
        )

    for c in [_BOQ_COL_DESC, _BOQ_COL_QTY_MAT, _BOQ_COL_QTY_SVC]:
        if c not in df.columns:
            raise KeyError(
                f"Required column '{c}' not found in BoQ sheet '{matched_sheet}'. "
                f"Found: {list(df.columns)}"
            )

    df[_BOQ_COL_DESC] = df[_BOQ_COL_DESC].astype(str).str.strip()
    df = df[df[_BOQ_COL_DESC].notna() & (df[_BOQ_COL_DESC] != "") & (df[_BOQ_COL_DESC] != "nan")].copy()
    df[_BOQ_COL_QTY_MAT] = pd.to_numeric(df[_BOQ_COL_QTY_MAT], errors="coerce").fillna(0)
    df[_BOQ_COL_QTY_SVC] = pd.to_numeric(df[_BOQ_COL_QTY_SVC], errors="coerce").fillna(0)

    boq_mat = df.groupby(_BOQ_COL_DESC, as_index=False)[_BOQ_COL_QTY_MAT].sum()
    boq_svc = df.groupby(_BOQ_COL_DESC, as_index=False)[_BOQ_COL_QTY_SVC].sum()
    return boq_mat, boq_svc

def _boq_build_outputs(boq_mat, boq_svc):
    def _process(tpl_df, boq_agg, boq_qty_col):
        tpl = tpl_df.copy()
        col_desc = _boq_find_col(tpl, "description")
        col_qty  = _boq_find_col(tpl, "quantity")
        col_unit = _boq_find_col_opt(tpl, "unit")
        col_no   = (_boq_find_col_opt(tpl, "no material")
                    or _boq_find_col_opt(tpl, "material code")
                    or _boq_find_col_opt(tpl, "no."))

        tpl[col_desc] = tpl[col_desc].astype(str).str.strip()
        if col_no:
            tpl[col_no] = tpl[col_no].astype(str).str.strip()

        boq = boq_agg.copy()
        boq_desc_col = boq.columns[0]
        boq[boq_desc_col] = boq[boq_desc_col].astype(str).str.strip()

        if _BOQ_DESC_ALIASES:
            _alias_lower = {k.lower(): v for k, v in _BOQ_DESC_ALIASES.items()}
            boq[boq_desc_col] = boq[boq_desc_col].apply(
                lambda x: _alias_lower.get(str(x).strip().lower(), x)
            )

        boq = boq.rename(columns={boq_qty_col: "_boq_qty"})
        m = tpl.merge(boq[[boq_desc_col, "_boq_qty"]],
                      left_on=col_desc, right_on=boq_desc_col, how="left")
        m["_boq_qty"] = pd.to_numeric(m["_boq_qty"], errors="coerce").fillna(0)

        def _calc_qty(row):
            q = float(row["_boq_qty"])
            u = str(row.get(col_unit, "")).strip().upper() if col_unit else ""
            if u == "KM":
                q = round(q / 1000.0, 3)
            if abs(q - round(q)) < 1e-9:
                return int(round(q))
            return q

        m[col_qty] = m.apply(_calc_qty, axis=1)
        extra = [c for c in [boq_desc_col, "_boq_qty"] if c in m.columns and c != col_desc]
        m = m.drop(columns=extra, errors="ignore")
        m[col_qty] = pd.to_numeric(m[col_qty], errors="coerce").fillna(0)
        m = m[m[col_qty] > 0].copy()

        if col_no and col_no in m.columns:
            m = m[~m[col_no].astype(str).str.contains(r"-\d+$", regex=True, na=False)].copy()
            m = (m.sort_values(by=col_qty, ascending=False)
                  .drop_duplicates(subset=[col_no], keep="first"))
        return m

    mat_tpl = pd.read_excel(BytesIO(base64.b64decode(_BOQMAT_B64)), engine="openpyxl", dtype=str)
    svc_tpl = pd.read_excel(BytesIO(base64.b64decode(_BOQSVC_B64)), engine="openpyxl", dtype=str)

    mat_out = _process(mat_tpl, boq_mat, _BOQ_COL_QTY_MAT)
    svc_out = _process(svc_tpl, boq_svc, _BOQ_COL_QTY_SVC)

    tpl_mat_desc = set(mat_tpl[_boq_find_col(mat_tpl, "description")].astype(str).str.strip())
    tpl_svc_desc = set(svc_tpl[_boq_find_col(svc_tpl, "description")].astype(str).str.strip())

    desc_col    = boq_mat.columns[0]
    qty_col_mat = boq_mat.columns[1]
    qty_col_svc = boq_svc.columns[1]

    missing_mat = sorted([
        d for d, v in zip(boq_mat[desc_col], boq_mat[qty_col_mat])
        if float(v or 0) > 0 and d not in tpl_mat_desc
    ])
    missing_svc = sorted([
        d for d, v in zip(boq_svc[desc_col], boq_svc[qty_col_svc])
        if float(v or 0) > 0 and d not in tpl_svc_desc and not _boq_is_special_svc_skip(d)
    ])

    return mat_out, svc_out, missing_mat, missing_svc

def _boq_inject_homepass(svc_out, homepass: float):
    svc_tpl = pd.read_excel(BytesIO(base64.b64decode(_BOQSVC_B64)), engine="openpyxl", dtype=str)
    col_no  = (_boq_find_col_opt(svc_tpl, "no material")
               or _boq_find_col_opt(svc_tpl, "material code"))
    col_qty = _boq_find_col(svc_tpl, "quantity")

    if col_no is None:
        return svc_out

    svc_tpl[col_no] = svc_tpl[col_no].astype(str).str.strip()
    row = svc_tpl[svc_tpl[col_no] == _BOQ_SVC_HP_CODE].copy()
    if row.empty:
        return svc_out

    qty_val = int(homepass) if abs(homepass - round(homepass)) < 1e-9 else homepass
    row[col_qty] = qty_val

    if col_no in svc_out.columns:
        mask = svc_out[col_no].astype(str).str.strip() == _BOQ_SVC_HP_CODE
        if mask.any():
            svc_out = svc_out.copy()
            svc_out.loc[mask, col_qty] = qty_val
            return svc_out

    return pd.concat([svc_out, row], ignore_index=True)

def _boq_write_value_only(template_b64: str, out_df, out_path: str):
    wb = _boq_load_wb(template_b64)
    while len(wb.worksheets) > 1:
        del wb[wb.worksheets[-1].title]
    wb._external_links = []
    ws = wb.worksheets[0]
    ws.auto_filter.ref = None
    for rd in ws.row_dimensions.values():
        rd.hidden = False

    headers = [cell.value for cell in ws[1]]
    last_col = 0
    for i, h in enumerate(headers, start=1):
        if h is not None and str(h).strip() != "":
            last_col = i
    if last_col == 0:
        raise RuntimeError("Template header row appears empty; cannot write output.")

    header_vals = headers[:last_col]
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    df_cols = list(out_df.columns)
    df_map = {str(c).strip(): c for c in df_cols}

    aligned_rows = []
    for _, row in out_df.iterrows():
        out_row = []
        for h in header_vals:
            key = str(h).strip()
            col = h if h in out_df.columns else df_map.get(key)
            out_row.append(row[col] if col is not None else "")
        aligned_rows.append(out_row)

    for r_idx, row in enumerate(aligned_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_boq_excel_val(val))

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)

_BOQ_OK = True


def main():
    root = tk.Tk()
    app = FTTHAutomationApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
